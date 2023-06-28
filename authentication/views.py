import calendar
from django.utils.dateparse import parse_datetime
import json
from django.contrib.auth.models import User
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import logout
from django.shortcuts import HttpResponseRedirect
from django.template import RequestContext
from django.views.decorators.csrf import csrf_protect
from django.db import connection
from django.contrib import messages
from datetime import datetime
from django.contrib.auth.forms import UserCreationForm
import sys
from array import *
from django.contrib.auth import login as authlogin
import io
import xlsxwriter
from xlsxwriter.utility import xl_range
from django.contrib.auth.decorators import login_required
from fairways import settings
import openpyxl
# Create your views here.
# generatebutton = "INACTIVE"
@csrf_protect

def home(request): 
    if request.user.is_authenticated:
        return redirect('payroll_main')
    return render(request, 'login.html')

@login_required
def user(request):
    company = request.session['company']
    role1 = request.session['role1']
    if role1 != "PAYROLL":
        cursor = connection.cursor()
        if role1 == "ADMIN":
            cursor = connection.cursor()
            cursor.execute(f"SELECT * FROM TBL_USER WHERE COMPANY = '{company}' and ROLE = 'PAYROLL'")
            data = cursor.fetchall()
            return render(request, 'user.html',{'data': data, 'company': company, 'role1': request.session['role1']})
        elif role1 == "SUPERUSER":
            cursor = connection.cursor()
            cursor.execute(f"SELECT * FROM TBL_USER where ROLE = 'ADMIN'")
            data = cursor.fetchall()
            return render(request, 'user_admin.html',{'data': data, 'company': company, 'role1': request.session['role1']})
    return redirect('payroll_main')

@login_required
def change_password(request):
    username = request.session['username']
    if request.method == "POST":
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        password3 = request.POST['password3']
        
        if password2 != password3:
            messages.error(request, "Passwords did not match.")
            return redirect("change_password")
        
        user = authenticate(username=username, password=password1)
        
        if user is not None:
            user.set_password(password3)
            user.save()
            messages.error(request, "New password saved!")
            return redirect("change_password")
        else:
            messages.error(request, "Wrong current password!")
            return redirect("change_password")

    return render(request,'change_password.html',{'role1': request.session['role1']})

@login_required
def signup(request):
    username = request.session['username']
    company = request.session['company']
    # role1 = request.session['role1']
    if request.method == "POST":
        username = request.POST['username']
        fname = request.POST['fname']
        lname = request.POST['lname']
        password = request.POST['password']
        designation = request.POST['role'].upper()
        email = ""
        
        if User.objects.filter(username=username):
            messages.error(request, "Username already exist! Please try some other username.")
            return redirect("signup")
        
        if len(username)>20:
            messages.error(request, "Username must be under 20 charcters!!")
            return redirect('signup')
        
        if not username.isalnum():
            messages.error(request, "Username must be Alpha-Numeric!!")
            return redirect('signup')

        if designation == "ROLE":
            messages.error(request, "Select role of the user.")
            return redirect('signup')
        
        myuser = User.objects.create_user(username, email, password)
        myuser.first_name = fname
        myuser.last_name = lname
        myuser.is_active = True
        myuser.save()
        cursor = connection.cursor()
        cursor.execute(f"SELECT * FROM auth_user where username = '{username}'")
        data = cursor.fetchall()
        for r in data:
            id = r[0]
        cursor.close()
        cursor = connection.cursor()
        query = ("insert into TBL_USER (username, lastname, firstname, company, role, id)"
                            "VALUES (%s, %s, %s, %s, %s, %s)"
                        )
        data_values = (username, lname.upper(), fname.upper(), company, designation, id)
        cursor.execute(query,data_values)
        r = cursor.fetchall()
        return redirect("user")   
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_ROLE ORDER BY ROLE1")
    user_role = cursor.fetchall()
    return render(request, "add_user.html", {'user_role': user_role, 'role1': request.session['role1']})

@login_required
def signup_admin(request):
    username = request.session['username']
    # company = request.session['company']
    # role1 = request.session['role1']
    if request.method == "POST":
        username = request.POST['username']
        fname = request.POST['fname']
        lname = request.POST['lname']
        password = request.POST['password']
        company = request.POST['company'].upper()
        email = ""
        
        if User.objects.filter(username=username):
            messages.error(request, "Username already exist! Please try some other username.")
            return redirect("signup_admin")
        
        if len(username)>20:
            messages.error(request, "Username must be under 20 charcters!!")
            return redirect('signup_admin')
        
        if not username.isalnum():
            messages.error(request, "Username must be Alpha-Numeric!!")
            return redirect('signup_admin')

        if company == "COMPANY":
            messages.error(request, "Select company of the user.")
            return redirect('signup_admin')
        
        myuser = User.objects.create_user(username, email, password)
        myuser.first_name = fname
        myuser.last_name = lname
        myuser.is_active = True
        myuser.save()
        cursor = connection.cursor()
        cursor.execute(f"SELECT * FROM auth_user where username = '{username}'")
        data = cursor.fetchall()
        for r in data:
            id = r[0]
        cursor.close()
        cursor = connection.cursor()
        query = ("insert into TBL_USER (username, lastname, firstname, company, role, id)"
                            "VALUES (%s, %s, %s, %s, %s, %s)"
                        )
        data_values = (username, lname.upper(), fname.upper(), company, "ADMIN", id)
        cursor.execute(query,data_values)
        r = cursor.fetchall()
        return redirect("user")   
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_COMPANY ORDER BY COMPANY")
    data = cursor.fetchall()
    return render(request, "add_admin.html", {'data': data, 'role1': request.session['role1']})

def signout(request):
    logout(request)
    return redirect('login')

@login_required
def payroll_main(request):
    username = request.session['username']
    company = request.session['company']
    role1 = request.session['role1']
    return render(request, 'payroll_main.html', {'username':  username, 'company':  company, 'role1':  role1})

def guest_main(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where EMPNO = {username} order by periodfrom desc limit 25")
    data = cursor.fetchall() 
    return render(request, 'guest_main.html', {'username':  username, 'company':  company, 'data': data, 'role1': request.session['role1']})

def guest_password(request):
    username = request.session['username']
    if request.method == "POST":
        password1 = request.POST['password1'].upper()
        password2 = request.POST['password2'].upper()
        password3 = request.POST['password3'].upper()
        
        if password2 != password3:
            msg = "Passwords did not match"
            return render(request,'guest_password.html', {'msg': msg})
        
        # user = authenticate(username=username, password=password1)
        cursor = connection.cursor()
        cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE username = {username} and password = '{password1}'")
        if cursor.rowcount > 0:
            cursor.execute(f"UPDATE TBL_MASTERFILE SET PASSWORD = '{password2}' WHERE empno = {username}")
            messages.info(request, 'New password saved!')
            msg = "New password saved!"
            return redirect('guest_main')
        else:
            msg = 'Wrong current password!'
            return render(request, 'guest_password.html', {'msg': msg})

    return render(request,'guest_password.html')

def show_guest_payslip(request, rowid):
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where ROWID = {rowid}")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        dateformat = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        dateformat.set_align('left')
        bold1 = workbook.add_format({"bold": True})
        money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        money2 = workbook.add_format({'num_format': '#,##0.00'})
        money2.set_bottom(3)
        worksheet.set_column('H:H',10,None)
        allborder = workbook.add_format({'border':1})
        row2 = 0 # row count for left part of payslip
        row1 = 1 # use to check if top or bottom of the page
        row3 = 0 # row count for right part of payslip
        for data in data:
            company = data[25]
            row4 = row2
            period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            periodfrom = data[3]
            periodto = data[4]
            #Left part of payslip
            worksheet.write(row2, 0, company, bold1)
            row2+=1
            worksheet.write(row2, 0, "Payroll Period:")
            worksheet.write(row2, 2, period)
            row2+=3
            worksheet.write(row2, 0, "Employee No/Name:")
            worksheet.write(row2, 2, str(data[1]) + " - " + data[2])
            row2+=1
            worksheet.write(row2, 0, "Earnings:")
            row2+=1
            worksheet.write(row2, 0, "No Of Trips")
            worksheet.write(row2, 3, data[9], money2)
            row2+=1
            worksheet.write(row2, 0, "Basic Pay")
            worksheet.write(row2, 3, data[10], money1)
            row2+=1
            if float(data[26]) > 0:
                worksheet.write(row2, 0, "Other Earnings")
                worksheet.write(row2, 3, data[26], money1)
                row2+=1
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money1)
            row2 = row4 + 19
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money2)
            row2+=1
            worksheet.write(row2, 0, "Net Pay")
            worksheet.write(row2, 3, data[22], money2)
            row2+=1
            worksheet.write(row2, 0, "Received By:")
            
            #right part of payslip
            row3+=1
            worksheet.write(row3, 5, "Pay Date:")
            worksheet.write(row3, 7, datetime.now(), dateformat)
            row3+=3
            cursor.execute(f"SELECT * FROM TBL_MASTERFILE where empno = {data[1]}")
            emp = cursor.fetchall()
            for emp in emp:
                dept = emp[23]
            worksheet.write(row3, 6, "Dept:")
            worksheet.write(row3, 7, dept)
            row3+=1
            worksheet.write(row3, 5, "Deductions")
            row3+=1
            if data[16] != 0:
                worksheet.write(row3, 5, "SSS CONTRIBUTION")
                worksheet.write(row3, 8, data[16], money1)
                row3+=1
            if data[18] != 0:
                worksheet.write(row3, 5, "PH CONTRIBUTION")
                worksheet.write(row3, 8, data[18], money1)
                row3+=1
            if data[19] != 0:
                worksheet.write(row3, 5, "PAGIBIG CONTRI")
                worksheet.write(row3, 8, data[19], money1)
                row3+=1
            if data[20] != 0:
                worksheet.write(row3, 5, "TAX")
                worksheet.write(row3, 8, data[20], money1)
                row3+=1
            cursor.execute(f"SELECT dedname FROM TBL_FINAL_DEDUCTIONS where empno ={data[1]} and periodfrom = '{periodfrom}' and periodto = '{periodto}' order by dedname" )
            deductions = cursor.fetchall()
            count = cursor.rowcount
            if count > 0:
                for deductions in deductions:
                    dedname = deductions[0]
                    cursor.execute(f"SELECT sum(amount) as amount FROM TBL_FINAL_DEDUCTIONS where dedname = '{dedname}' and empno = {data[1]} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                    dedamount = cursor.fetchall()
                    for dedamount in dedamount:
                        if dedamount != None:
                            worksheet.write(row3, 5, dedname)
                            worksheet.write(row3, 8, dedamount[0], money1)
                            row3+=1
            row3 = row4 + 19
            worksheet.write(row3, 5, "Total Deductions")
            worksheet.write(row3, 8, data[21], money2)
            row3 += 2
            worksheet.write(row3, 5, "Date Received:")        
            if row1 % 2 == 0: #check if top or bottom part of payslip for new line spacing
                row2 += 1
                row3 += 1
            else:
                row2 += 2
                row3 += 2
            row1 += 1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='Payslip.xlsx')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def login(request):
    c = {}
    cursor = connection.cursor()
    if request.user.is_authenticated:
        return redirect('payroll_main')
        # return render(request, 'payroll_main.html', {'username':  username, 'company':  company, 'role1':  role1})
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        cursor.execute(f"SELECT * FROM TBL_USER WHERE USERNAME = '{username}'")
        data = cursor.fetchall()
        for row in data:
            role1 = row[5]
            company = row[4]
        user = authenticate(request, username=username, password=password)
        if user is not None:
            authlogin(request, user)
            request.session['username'] = username
            request.session['company'] = company
            request.session['role1'] = role1
            return redirect('payroll_main')
            # return render(request, 'payroll_main.html', {'username':  username, 'company':  company, 'role1':  role1})
        else:
            password = password.upper()
            cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE username = '{username}' and password = '{password}' and showpayslip = 'YES'")
            if cursor.rowcount > 0:
                data = cursor.fetchall()
                for data in data:
                    company = data[2] + ", " + data[3] + ", " + data[4]
                request.session['username'] = username
                request.session['company'] = company
                return redirect('guest_main')
            else:
                msg = 'Error Login!'
                form = AuthenticationForm(request.POST)
                return render(request, 'login.html', {'form': form, 'msg': msg})
    
    form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})
    # return render_to_response('login.html', context_instance=RequestContext(request))
    
def confirm_delete_user(request, rowid):
    empno = rowid
    source = 'delete_user'
    message = 'Are you sure you want to delete user with ID '
    return render(request, 'confirm_delete.html',{'empno': empno, 'source': source, 'message': message})

def delete_user(request, rowid):
    if request.method == 'POST':  
        cursor = connection.cursor()
        cursor.execute(f"delete FROM auth_user where id = {rowid}")
        cursor.close()
        cursor = connection.cursor()
        cursor.execute(f"delete FROM TBL_USER where id = {rowid}")
        return redirect("user")
            
########################### MASTERFILE ADD/EDIT/DELETE ##################################   
@login_required  
def masterfile(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE COMPANY = '{company}' and status = 'ACTIVE' order by empno desc")
    data = cursor.fetchall()
    return render(request, 'masterfile.html',{'data': data, 'company': company})

@login_required 
def employee_image(request,rowid):
    company = request.session['company']
    cursor = connection.cursor()
    if request.method == 'POST':
        image_data = request.POST['captured_image_data']
        cursor.execute(f"update TBL_MASTERFILE set imagepath = '{image_data}' where rowid = {rowid}")
        return redirect("masterfile")
    cursor.execute(f"Select imagepath from TBL_MASTERFILE where rowid = {rowid}")
    if cursor.rowcount > 0:
        data = cursor.fetchall()
        for data in data:
            image_data = data[0]
    else:
        image_data = 1
    return render(request, 'employee_image.html',{'rowid': rowid, 'image_data': image_data}) 

@login_required 
def add_employee(request):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        empno=0
        # Get all values from registration form
        lname = request.POST['lname'].upper()
        fname = request.POST['fname'].upper() 
        mname = request.POST['mname'].upper()
        birthdate = request.POST['birthdate']
        address1 = request.POST['address1'].upper()
        address2 = request.POST['address2'].upper()
        contact = request.POST['contact'].upper()
        gender = request.POST['gender'].upper()
        civil = request.POST['civil'].upper()
        spouse = request.POST['spouse'].upper()
        child1 = request.POST['child1'].upper()
        child2 = request.POST['child2'].upper()
        sss = request.POST['sss'].upper()
        phic = request.POST['phic'].upper()
        hdmf = request.POST['hdmf'].upper()
        tin = request.POST['tin'].upper()
        sssd = request.POST['sssd'].upper()
        phicd = request.POST['phicd'].upper()
        hdmfd = request.POST['hdmfd'].upper()
        taxd = request.POST['taxd'].upper()
        hdmfpay = request.POST['hdmfpay']
        department = request.POST['department'].upper()
        position = request.POST['position'].upper()
        location = request.POST['location'].upper()
        if request.POST['triprate'] == "": triprate = 0
        else: triprate = request.POST['triprate']
        if request.POST['salary'] == "": salary = 0
        else: salary = request.POST['salary']
        if request.POST['allowance'] == "":allowance = 0
        else: allowance = request.POST['allowance']
        datehired = request.POST['datehired']
        datepermanent = request.POST['datepermanent'].upper()
        status = request.POST['status'].upper()
        password = request.POST['password'].upper()
        showpayslip = request.POST['showpayslip'].upper()
        payrolltype = request.POST['payrolltype'].upper()
        atm = request.POST['atm']
        mother = request.POST['mother'].upper()
        emergency = request.POST['emergency'].upper()
        userlastupdate = username
        datelastupdated  = datetime.now() # save the date and time
        company = company

        cursor = connection.cursor()
        # Get last Employee Number from active company then add 1 for new empno 
        cursor.execute(f"SELECT max(EMPNO) as empno FROM TBL_MASTERFILE where COMPANY = '{company}'")
        r  = cursor.fetchall()
        for row in r:
            empno = row[0]
        empno = empno + 1
        cursor.close
        cursor = connection.cursor()
        # insert new data to TBL_MASTERFILE table
        query = ("insert into TBL_MASTERFILE (empno,lname, fname, mname, birthdate, address1, address2, \
                contact, gender, civil, spouse, child1, child2, sss, phic, hdmf, tax, sssd, phicd, hdmfd, \
                taxd, hdmfpay, department, position, location, triprate, salary, allowance, datehired, datepermanent,\
                status, password, showpayslip, payrolltype, atm, mother, emergency, userlastupdate, datelastupdated, company,username)"
                        "VALUES (%s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,\
                            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                    )
        data_values = (empno, lname, fname, mname, birthdate, address1, address2, contact, gender, \
                    civil, spouse, child1, child2, sss, phic, hdmf, tin, sssd, phicd, hdmfd, taxd, hdmfpay, \
                        department, position, location, triprate, salary, allowance, datehired, datepermanent, status, \
                            password, showpayslip, payrolltype, atm, mother, emergency, userlastupdate, datelastupdated, company, empno)
        cursor.execute(query,data_values)
        r = cursor.fetchall()
        messages.info(request, 'New Employee successfully added!.')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))    
    cursor = connection.cursor()
    cursor.execute(f"select * from TBL_LOCATION where company = '{company}' order by location")
    data = cursor.fetchall()
    return render(request, 'add_employee.html',{'data': data})

@login_required 
def edit_employee(request, empno):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        # Get all values from registration form
        lname = request.POST['lname'].upper()
        fname = request.POST['fname'].upper() 
        mname = request.POST['mname'].upper()
        birthdate = request.POST['birthdate']
        address1 = request.POST['address1'].upper()
        address2 = request.POST['address2'].upper()
        contact = request.POST['contact'].upper()
        gender = request.POST['gender'].upper()
        civil = request.POST['civil'].upper()
        spouse = request.POST['spouse'].upper()
        child1 = request.POST['child1'].upper()
        child2 = request.POST['child2'].upper()
        sss = request.POST['sss'].upper()
        phic = request.POST['phic'].upper()
        hdmf = request.POST['hdmf'].upper()
        tin = request.POST['tin'].upper()
        sssd = request.POST['sssd'].upper()
        phicd = request.POST['phicd'].upper()
        hdmfd = request.POST['hdmfd'].upper()
        taxd = request.POST['taxd'].upper()
        hdmfpay = request.POST['hdmfpay']
        department = request.POST['department'].upper()
        position = request.POST['position'].upper()
        location = request.POST['location'].upper()
        triprate = request.POST['triprate']
        salary = request.POST['salary']
        allowance = request.POST['allowance']
        datehired = request.POST['datehired']
        datepermanent = request.POST['datepermanent'].upper()
        status = request.POST['status'].upper()
        password = request.POST['password'].upper()
        showpayslip = request.POST['showpayslip'].upper()
        payrolltype = request.POST['payrolltype'].upper()
        atm = request.POST['atm']
        mother = request.POST['mother'].upper()
        emergency = request.POST['emergency'].upper()
        userlastupdate = username
        datelastupdated  = datetime.now() # save the date and time
        company = company

        cursor = connection.cursor()
        query = f"update TBL_MASTERFILE set lname = '{lname}', fname = '{fname}', mname = '{mname}', birthdate = '{birthdate}', address1 = '{address1}', address2 = '{address2}', contact = '{contact}', gender = '{gender}', civil = '{civil}', spouse = '{spouse}', child1 = '{child1}', child2 = '{child2}', sss = '{sss}', phic = '{phic}', hdmf = '{hdmf}', tax = '{tin}', sssd = '{sssd}', phicd = '{phicd}', hdmfd = '{hdmfd}', taxd = '{taxd}', hdmfpay = {hdmfpay}, department = '{department}', position = '{position}', location = '{location}', triprate = {triprate}, salary = {salary}, allowance = {allowance}, datehired = '{datehired}', datepermanent = '{datepermanent}', status = '{status}', password = '{password}', showpayslip = '{showpayslip}', payrolltype = '{payrolltype}', atm = '{atm}', mother = '{mother}', emergency = '{emergency}', userlastupdate = '{userlastupdate}', datelastupdated = '{datelastupdated}', company = '{company}' where empno = {empno}"          
        cursor.execute(query)
        return redirect("masterfile")

    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE where EMPNO = {empno}")
    data = cursor.fetchall() 
    for row in data:
        birthdate = str(row[5])
        birthdate = birthdate[0:10]
        datehired = str(row[29])
        datehired = datehired[0:10]
        datepermanent = str(row[30])
        datepermanent = datepermanent[0:10]
        locationval = row[25]
    cursor.execute(f"select * from TBL_LOCATION where company = '{company}' order by location")
    location = cursor.fetchall()
    return render(request, 'edit_employee.html', {'locationval': locationval, 'data': data, 'birthdate': birthdate, 'datehired': datehired, 'datepermanent': datepermanent, 'location': location})

def confirm_delete_employee(request, empno):
    empno = empno
    source = 'delete_employee'
    message = 'Are you sure you want to delete employee with ID '
    return render(request, 'confirm_delete.html',{'empno': empno, 'source': source, 'message': message})  

def delete_employee(request, empno):
    if request.method == 'POST':  
        cursor = connection.cursor()
        cursor.execute(f"delete FROM TBL_MASTERFILE where EMPNO = {empno}")
        return redirect("masterfile")  
########################### END OF MASTERFILE ADD/EDIT/DELETE ##################################

########################### DEDUCTIONS LIST ADD/EDIT/DELETE ##################################
@login_required 
def deductions(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE WHERE COMPANY = '{company}' order by PRIORITY, DEDUCTION_TYPE")
    data = cursor.fetchall()
    return render(request, 'deductions.html',{'data': data, 'company': company})

@login_required 
def add_deductions(request): 
    company = request.session['company']
    if request.method == 'POST':
        
        ded_name = request.POST['ded_name'].upper()
        priority = request.POST['priority']
        ded_type = request.POST['ded_type'].upper()
        
        cursor = connection.cursor()
        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_name = '{ded_name}' and deduction_type = '{ded_type}' and company = '{company}'")
        count  = cursor.rowcount
        if count > 0 :
            cursor.close
            messages.info(request, 'Duplicate Deductions name and type!')
            return redirect("add_deductions")
        else:   
            cursor.close
            cursor = connection.cursor()
            query = ("insert into TBL_DEDUCTION_TYPE (deduction_name, priority, deduction_type, company)"
                            "VALUES (%s, %s, %s, %s)"
                        )
            data_values = (ded_name, priority, ded_type, company)
            cursor.execute(query,data_values)
            r = cursor.fetchall()
            messages.info(request, 'New deduction name successfully added!.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    return render(request, 'add_deductions.html')

@login_required 
def edit_deductions(request, rowid):
    company = request.session['company']
    if request.method == 'POST':
        
        ded_name_new = request.POST['ded_name'].upper()
        priority_new = request.POST['priority']
        ded_type_new = request.POST['ded_type'].upper()
        
        cursor = connection.cursor()
        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where rowid = {rowid}")
        r  = cursor.fetchall()
        for row in r:
            ded_name = row[1]
            ded_type = row[3]
        if ded_name != ded_name_new or ded_type != ded_type_new:
            cursor.close
            cursor = connection.cursor()
            cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_name = '{ded_name_new}' and deduction_type = '{ded_type_new}' and company = '{company}'")
            count  = cursor.rowcount
            if count > 0 :
                #cursor.close
                messages.info(request, 'Duplicate Deductions name and type!')
                return redirect("edit_deductions")
            else:   
                cursor.close
                cursor = connection.cursor()
                query = f"update TBL_DEDUCTION_TYPE set deduction_name = '{ded_name_new}', priority = {priority_new}, deduction_type = '{ded_type_new}' where rowid = {rowid}"          
                cursor.execute(query)
                return redirect("deductions")
        else:
            cursor.close
            cursor = connection.cursor()
            query = f"update TBL_DEDUCTION_TYPE set deduction_name = '{ded_name_new}', priority = {priority_new}, deduction_type = '{ded_type_new}' where rowid = {rowid}"          
            cursor.execute(query)
            return redirect("deductions")
    ###############################
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where rowid = {rowid} and company = '{company}'")
    data = cursor.fetchall() 
    return render(request, 'edit_deductions.html', {'data': data}) 


def delete_deductions(request, rowid):
    company = request.session['company']
    if request.method == 'POST':  
        cursor = connection.cursor()
        cursor.execute(f"delete FROM TBL_DEDUCTION_TYPE where ROWID = {rowid} and company = '{company}'")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        # return redirect("deductions")

########################### END OF DEDUCTIONS LIST ADD/EDIT/DELETE ##################################

########################### SSS ADD/EDIT/DELETE ##################################
@login_required 
def sss(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM TBL_SSS order by sssfrom")
    data = cursor.fetchall()
    return render(request, 'sss.html',{'data': data, 'company': company})

@login_required 
def add_sss(request):
    if request.method == 'POST':
        sssfrom = request.POST['sssfrom']
        sssto = request.POST['sssto']
        ee = request.POST['ee']
        er = request.POST['er']
        ec = request.POST['ec']
        wispee = request.POST['wispee']
        wisper = request.POST['wisper']
        totee = request.POST['totee']
        toter = request.POST['toter']

        cursor = connection.cursor()
        # insert new data to TBL_MASTERFILE table
        query = ("insert into TBL_SSS (sssfrom, sssto, ee, er, ec, wispee, wisper, totee, toter)"
                        "VALUES (%s, %s, %s, %s,%s, %s, %s, %s, %s)"
                    )
        data_values = (sssfrom, sssto, ee, er, ec, wispee, wisper, totee, toter)
        cursor.execute(query,data_values)
        r = cursor.fetchall()
        messages.success(request, 'New SSS successfully added!.')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    return render(request, 'add_sss.html')

@login_required 
def edit_sss(request, rowid):
    if request.method == 'POST':
        sssfrom = request.POST['sssfrom']
        sssto = request.POST['sssto']
        ee = request.POST['ee']
        er = request.POST['er']
        ec = request.POST['ec']
        wispee = request.POST['wispee']
        wisper = request.POST['wisper']
        totee = request.POST['totee']
        toter = request.POST['toter']
        
        cursor = connection.cursor()
        query = f"update TBL_SSS set sssfrom = {sssfrom}, sssto = {sssto}, ee = {ee}, er = {er}, ec = {ec}, wispee = {wispee}, wisper = {wisper}, totee = {totee}, toter = {toter} where rowid = {rowid}"          
        cursor.execute(query)
        #r = cursor.fetchall()
        # messages.info(request, 'SSS record successfully updated!.')
        return redirect("sss")
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_SSS where rowid = {rowid}")
    data = cursor.fetchall() 
    return render(request, 'edit_sss.html', {'data': data})

@login_required 
def delete_sss(request, rowid):
    if request.method == 'POST':  
        cursor = connection.cursor()
        cursor.execute(f"delete FROM TBL_SSS where ROWID = {rowid}")
        return redirect("sss")

########################### END OF SSS ADD/EDIT/DELETE ##################################

########################### PHIC AND HDMF ADD/EDIT ##################################
@login_required 
def phichdmf(request):
    role1 = request.session['role1']
    if request.method == 'POST':
        phicto1 = request.POST['phicto1']
        phicfrom2 = request.POST['phicfrom2']
        phicto2 = request.POST['phicto2']
        phicfrom3 = request.POST['phicfrom3']
        amt1 = request.POST['amt1']
        amt2 = request.POST['amt2']
        amt3 = request.POST['amt3']
        hdmf = request.POST['hdmf']
        
        cursor = connection.cursor()
        query = f"update TBL_PHICHDMF set phicto1 = {phicto1}, phicfrom2 = {phicfrom2}, phicto2 = {phicto2}, phicfrom3 = {phicfrom3}, amt1 = {amt1}, amt2 = {amt2}, amt3 = {amt3}, hdmf = {hdmf}"          
        cursor.execute(query)
        #r = cursor.fetchall()
        # messages.info(request, 'PHIC/HDMF record successfully updated!.')
        return redirect("phichdmf")
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PHICHDMF")
    data = cursor.fetchall() 
    return render(request, 'phichdmf.html', {'data': data, 'role1': role1})

########################### END OF PHIC AND HDMF ADD/EDIT ##################################

########################### PAYROLL PERIOD ADD/EDIT/DELETE ##################################
@login_required 
def payroll_period(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where status = 'ACTIVE' and company = '{company}'")
    data = cursor.fetchall()
    return render(request, 'payroll_period.html',{'data': data, 'company': company, 'role1': request.session['role1']})
@login_required 
def add_payroll_period(request,rowid):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        if rowid == 1:
            # Get all values from registration form
            paytype = request.POST['paytype']
            employeetype = request.POST['employeetype']
            periodfrom = request.POST['periodfrom']
            periodto = request.POST['periodto']
            status = "ACTIVE"
            paymonth = request.POST['paymonth1']
            payyear = request.POST['payyear1']
            datemaintained = datetime.now()
            periodfrom = parse_datetime(periodfrom)
            periodto = parse_datetime(periodto)
            user = username
            company = company
            payrollperiod = periodfrom.strftime("%m/%d/%Y") + "-" + periodto.strftime("%m/%d/%Y")
            cursor = connection.cursor()
            query = ("insert into TBL_PAYROLL_PERIOD (paytype, employeetype, periodfrom, periodto, status, paymonth, payyear, datemaintained, user, company, payrollperiod)"
                            "VALUES (%s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s)"
                        )
            data_values = (paytype, employeetype, periodfrom, periodto, status, paymonth, payyear, datemaintained, user, company, payrollperiod)
            cursor.execute(query,data_values)
            r = cursor.fetchall()
            messages.success(request, 'New Payroll Period successfully added!.')
            return redirect("payroll_period")
    if rowid == 0:
        return render(request, 'add_payroll_period.html')
@login_required 
def edit_payroll_period(request, rowid):
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where rowid = {rowid}")
    data = cursor.fetchall() 
    for row in data:
        periodfrom = str(row[3])
        periodfrom = periodfrom[0:10]
        periodto = str(row[4])
        periodto = periodto[0:10]
        paymonth = row[6]
        payyear = row[7]
    return render(request, 'edit_payroll_period.html', {'data': data, 'periodto': periodto, 'periodfrom': periodfrom, 'paymonth': paymonth, 'payyear': payyear, 'role1': request.session['role1']})

def save_edited_period(request, rowid):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        # Get all values from registration form
        paytype = request.POST['paytype']
        employeetype = request.POST['employeetype']
        periodfrom = request.POST['periodfrom']
        periodto = request.POST['periodto']
        status = "ACTIVE"
        paymonth = request.POST['paymonth1']
        payyear = request.POST['payyear1']
        datemaintained = datetime.now()
        periodfrom = parse_datetime(periodfrom)
        periodto = parse_datetime(periodto)
        payrollperiod = periodfrom.strftime("%m/%d/%Y") + "-" + periodto.strftime("%m/%d/%Y")
        user = username
        company = company
        cursor = connection.cursor()
        query = f"update TBL_PAYROLL_PERIOD set paytype = '{paytype}', employeetype = '{employeetype}', periodfrom = '{periodfrom}',\
              periodto = '{periodto}', status = '{status}', paymonth = {paymonth}, payyear = {payyear}, \
            datemaintained = '{datemaintained}', user = '{user}', company = '{company}', payrollperiod = '{payrollperiod}' where rowid = {rowid}"          
        cursor.execute(query)
        return redirect("payroll_period")

def delete_payroll_period(request, rowid):
    if request.method == 'POST':  
        cursor = connection.cursor()
        cursor.execute(f"delete FROM TBL_PAYROLL_PERIOD where ROWID = {rowid}")
        return redirect("payroll_period")

########################### END OF PAYROLL PERIOD ADD/EDIT/DELETE ##################################
@login_required  
def other_list(request):
    company = request.session['company']
    if request.method == 'POST':
        rowid = request.POST['rowid']
        selected = request.POST['otherlist']
        listname = request.POST['otherlistname'].upper()
        cursor = connection.cursor()
        if rowid == "ADD":
            if selected == "LOCATION":
                query = ("insert into TBL_LOCATION (location, COMPANY) VALUES (%s, %s)")
                data_values = (listname, company )
                cursor.execute(query,data_values)
            elif selected == "VTR EXPENSES":
                query = ("insert into TBL_EXPENSES_LIST (expenses_name, COMPANY) VALUES (%s, %s)")
                data_values = (listname, company )
                cursor.execute(query,data_values)
            else:
                query = ("insert into TBL_EARNINGS_TYPE (earningsname, COMPANY) VALUES (%s, %s)")
                data_values = (listname, company )
                cursor.execute(query,data_values)
        elif rowid == "":
            selected = request.POST['otherlist']
        else:
            if selected == "LOCATION":
                cursor.execute(f"delete from TBL_LOCATION where rowid = {rowid}")
            elif selected == "VTR EXPENSES":
                cursor.execute(f"delete from TBL_EXPENSES_LIST where rowid = {rowid}")
            else:
                cursor.execute(f"delete from TBL_EARNINGS_TYPE where rowid = {rowid}")
        if selected == "LOCATION":
            cursor.execute(f"SELECT * FROM TBL_LOCATION where company = '{company}' order by location")
        elif selected == "VTR EXPENSES":
            cursor.execute(f"SELECT * FROM TBL_EXPENSES_LIST  where company = '{company}' order by expenses_name")
        else:
            cursor.execute(f"SELECT * FROM TBL_EARNINGS_TYPE  where company = '{company}' order by earningsname")
        locationlist = cursor.fetchall()        
        return render(request, 'other_list.html',{'locationlist': locationlist, 'selected': selected, 'role1': request.session['role1']})
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_LOCATION WHERE COMPANY = '{company}' order by location")
    locationlist = cursor.fetchall()
    return render(request, 'other_list.html',{'locationlist': locationlist, 'role1': request.session['role1']}) 

@login_required 
def checkduplicate(request):
    company = request.session['company']
    #Get the variable text
    request.session['vtrcheck'] = 0
    selected = request.POST['selected']
    text = request.POST['text']
    #Do whatever with the input variable text
    cursor = connection.cursor()
    if selected == "LOCATION":
        cursor.execute(f"SELECT * FROM TBL_LOCATION where company = '{company}' and location = '{text}'")
        if cursor.rowcount > 0:
            otherlistname = "YES"
        else:
            otherlistname = "NO"
            request.session['vtrcheck'] = 1
    elif selected == "VTR EXPENSES":
        cursor.execute(f"SELECT * FROM TBL_EXPENSES_LIST  where company = '{company}' and expenses_name = '{text}'")
        if cursor.rowcount > 0:
            otherlistname = "YES"
        else:
            otherlistname = "NO"
            request.session['vtrcheck'] = 1
    else:
        cursor.execute(f"SELECT * FROM TBL_EARNINGS_TYPE  where company = '{company}' and earningsname  = '{text}'")
        if cursor.rowcount > 0:
            otherlistname = "YES"
        else:
            otherlistname = "NO"
            request.session['vtrcheck'] = 1     
   #Send the response 
    return JsonResponse({'otherlistname': otherlistname})

@login_required  
def emp_deductions(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE COMPANY = '{company}' and status = 'ACTIVE' order by empno desc")
    data = cursor.fetchall()
    return render(request, 'emp_deductions.html',{'data': data, 'company': company, 'role1': request.session['role1']})

@login_required 
def nonfixed_deductions(request, empno):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        if request.POST['rowid'] == "":
            dedname = request.POST['dedtype']
            amount = request.POST['amount']
            datemaintained = request.POST['datemaintained']
            datelastupdate = request.POST['datelastupdate']
            payrollperiod = request.POST['period']
            periodfrom = payrollperiod[6:10]+"-"+payrollperiod[0:2]+"-"+payrollperiod[3:5]
            periodto = payrollperiod[17:21]+"-"+payrollperiod[11:13]+"-"+payrollperiod[14:16]
            user = request.POST['user']
            cursor = connection.cursor()
            cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE periodfrom = '{periodfrom}' and periodto = '{periodto}' and company = '{company}'")
            data = cursor.fetchall()
            for r in data:
                paymonth = r[6]
                payyear = r[7]
            query = ("insert into TBL_NONFIXED_DEDUCTIONS (empno, dedtype, amount, periodfrom, periodto, paymonth, payyear, datemaintained, datelastupdate, status, USER, COMPANY)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
            data_values = (empno, dedname, amount, periodfrom, periodto, paymonth, payyear, datemaintained, datelastupdate,0, user, company )
            cursor.execute(query,data_values)
            messages.success(request, 'New deduction successfully added.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            if request.POST['rowid1'] == "":
                rowid = request.POST['rowid']
                dedname = request.POST['dedtype']
                amount = request.POST['amount']
                payrollperiod = request.POST['period']
                periodfrom = payrollperiod[6:10]+"-"+payrollperiod[0:2]+"-"+payrollperiod[3:5]
                periodto = payrollperiod[17:21]+"-"+payrollperiod[11:13]+"-"+payrollperiod[14:16]
                datelastupdate = datetime.now()
                user = username
                cursor = connection.cursor()
                cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE periodfrom = '{periodfrom}' and periodto = '{periodto}' and company = '{company}'")
                data = cursor.fetchall()
                for r in data:
                    paymonth = r[6]
                    payyear = r[7]
                query = (f"update TBL_NONFIXED_DEDUCTIONS set dedtype = '{dedname}', amount = {amount}, periodfrom = '{periodfrom}', periodto = '{periodto}', paymonth = {paymonth}, payyear = {payyear}, datelastupdate = '{datelastupdate}', USER = '{user}' where rowid = '{rowid}'")
                cursor.execute(query)
                messages.success(request, 'Deduction successfully updated.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            else:
                rowid = request.POST['rowid']
                cursor = connection.cursor()
                query = (f"delete from TBL_NONFIXED_DEDUCTIONS where rowid = '{rowid}'") 
                cursor.execute(query)
                messages.success(request, 'Deduction deleted!')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_NONFIXED_DEDUCTIONS WHERE EMPNO = '{empno}' and status = 0")
    data = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE EMPNO = '{empno}'")
    data1 = cursor.fetchall()
    for r in data1:
        name = str(r[1])+ " - " + r[2] + ", " + r[3]
        position = r[24]
    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE WHERE deduction_type = 'NONFIXED' and company = '{company}'")
    dedtype = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE employeetype = '{position}' and company = '{company}' and status = 'ACTIVE'")
    dataperiod = cursor.fetchall()
    period = [""]
    for r in dataperiod:
        period.append(r[3].strftime("%m/%d/%Y") + "-" + r[4].strftime("%m/%d/%Y"))
    return render(request, 'nonfixed_deductions.html',{'period': period, 'empno': empno, 'name': name, 'username': username, 'data': data,'data1': data1, 'company': company, 'dedtype': dedtype, 'role1': request.session['role1']})
@login_required 
def checkdeductions1(request):
    #Get the variable text
    rowid = request.POST['text']
    #Do whatever with the input variable text
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_NONFIXED_DEDUCTIONS where rowid = '{rowid}'")
    data = cursor.fetchall()
    for r in data:
            period = r[4].strftime("%m/%d/%Y") + "-" + r[5].strftime("%m/%d/%Y")
            dedtype = r[2]
            dedamount = r[3]
            datemaintained = r[11]
            datelastupdate = r[12]
            user = r[9]

   #Send the response 
    return JsonResponse({'period': period, 'user': user, 'datelastupdate': datelastupdate, 'datemaintained': datemaintained, 'dedtype': dedtype, 'dedamount': dedamount})#amount, noofperiod, dedamount, balance)

@login_required 
def fixed_deductions(request, empno):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        if request.POST['rowid'] == "":
            dedname = request.POST['dedtype']
            amount = request.POST['amount']
            noofperiod = request.POST['noofperiod']
            perioddedamount = request.POST['perioddedamount']
            datestart = request.POST['datestart']
            deferredpay = request.POST['deferredpay']
            periodremaining = request.POST['periodremaining']
            balance = request.POST['balance']
            datemaintained = request.POST['datemaintained']
            datelastupdate = request.POST['datelastupdate']
            user = request.POST['user']
            cursor = connection.cursor()
            query = ("insert into TBL_FIXED_DEDUCTIONS (empno, dedtype, dedamount, noofperiod, perioddedamount, datestart, deferredpay, periodremaining, amtremaining, datemaintained, datelastupdate, status, USER, COMPANY)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
            data_values = (empno, dedname, amount, noofperiod, perioddedamount, datestart, deferredpay, periodremaining, balance, datemaintained, datelastupdate,0, user, company )
            cursor.execute(query,data_values)
            messages.success(request, 'New deduction successfully added.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            if request.POST['rowid1'] == "":
                rowid = request.POST['rowid']
                dedname = request.POST['dedtype']
                amount = request.POST['amount']
                noofperiod = request.POST['noofperiod']
                perioddedamount = request.POST['perioddedamount']
                datestart = request.POST['datestart']
                deferredpay = request.POST['deferredpay']
                periodremaining = request.POST['periodremaining']
                balance = request.POST['balance']
                # datemaintained = request.POST['datemaintained']
                datemaintained = datetime.now()
                user = username
                cursor = connection.cursor()
                query = (f"update TBL_FIXED_DEDUCTIONS set dedtype = '{dedname}', dedamount = {amount}, noofperiod = '{noofperiod}', perioddedamount = '{perioddedamount}', datestart = '{datestart}', deferredpay = '{deferredpay}', periodremaining = '{periodremaining}', amtremaining = '{balance}', datelastupdate = '{datemaintained}', USER = '{user}' where rowid = '{rowid}'")
                cursor.execute(query)
                messages.success(request, 'Deduction successfully updated.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            else:
                rowid = request.POST['rowid']
                query = (f"delete from TBL_FIXED_DEDUCTIONS where rowid = '{rowid}'")
                cursor = connection.cursor()
                cursor.execute(query)
                messages.success(request, 'Deduction deleted!')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_FIXED_DEDUCTIONS WHERE EMPNO = '{empno}' and status = 0")
    data = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE EMPNO = '{empno}'")
    data1 = cursor.fetchall()
    for r in data1:
        name = str(r[1])+ " - " + r[2] + ", " + r[3]
    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE WHERE deduction_type = 'FIXED' and company = '{company}'")
    dedtype = cursor.fetchall()
    return render(request, 'fixed_deductions.html',{'empno': empno, 'name': name, 'username': username, 'data': data,'data1': data1, 'company': company, 'dedtype': dedtype, 'role1': request.session['role1']})

@login_required 
def checkdeductions(request):
    #Get the variable text
    rowid = request.POST['text']
    #Do whatever with the input variable text
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_FIXED_DEDUCTIONS where rowid = '{rowid}'")
    data = cursor.fetchall()
    for r in data:
            dedtype = r[2]
            dedamount = r[3]
            noofperiod = r[4]
            dedamount = r[5]
            balance = r[6]
            periodremaining = r[7]
            datestart = r[8]
            datemaintained = r[9]
            datelastupdate = r[10]
            user = r[12]
            deferredpay = r[13]

   #Send the response 
    return JsonResponse({'deferredpay': deferredpay, 'user': user, 'datelastupdate': datelastupdate, 'datemaintained': datemaintained, 'datestart':datestart, 'periodremaining': periodremaining, 'dedtype': dedtype, 'dedamount': dedamount, 'noofperiod': noofperiod, 'dedamount': dedamount, 'balance': balance})#amount, noofperiod, dedamount, balance)

@login_required 
def other_earnings(request, empno):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        if request.POST['rowid'] == "":
            dedname = request.POST['dedtype']
            amount = request.POST['amount']
            datemaintained = request.POST['datemaintained']
            datelastupdate = request.POST['datelastupdate']
            payrollperiod = request.POST['period']
            periodfrom = payrollperiod[6:10]+"-"+payrollperiod[0:2]+"-"+payrollperiod[3:5]
            periodto = payrollperiod[17:21]+"-"+payrollperiod[11:13]+"-"+payrollperiod[14:16]
            user = request.POST['user']
            cursor = connection.cursor()
            cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE periodfrom = '{periodfrom}' and periodto = '{periodto}' and company = '{company}'")
            data = cursor.fetchall()
            for r in data:
                paymonth = r[6]
                payyear = r[7]
            query = ("insert into TBL_OTHER_EARNINGS (empno, EARNINGSTYPE, amount, periodfrom, periodto, paymonth, payyear, datemaintained, datelastupdate, status, USER, COMPANY, vtrno)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
            data_values = (empno, dedname, amount, periodfrom, periodto, paymonth, payyear, datemaintained, datelastupdate,0, user, company,0 )
            cursor.execute(query,data_values)
            messages.success(request, 'New Other earnings successfully added.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        else:
            if request.POST['rowid1'] == "":
                rowid = request.POST['rowid']
                dedname = request.POST['dedtype']
                amount = request.POST['amount']
                payrollperiod = request.POST['period']
                periodfrom = payrollperiod[6:10]+"-"+payrollperiod[0:2]+"-"+payrollperiod[3:5]
                periodto = payrollperiod[17:21]+"-"+payrollperiod[11:13]+"-"+payrollperiod[14:16]
                datelastupdate = datetime.now()
                user = username
                cursor = connection.cursor()
                cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE periodfrom = '{periodfrom}' and periodto = '{periodto}'  and company = '{company}'")
                data = cursor.fetchall()
                for r in data:
                    paymonth = r[6]
                    payyear = r[7]
                query = (f"update TBL_OTHER_EARNINGS set earningstype = '{dedname}', amount = {amount}, periodfrom = '{periodfrom}', periodto = '{periodto}', paymonth = {paymonth}, payyear = {payyear}, datelastupdate = '{datelastupdate}', USER = '{user}' where rowid = '{rowid}'")
                cursor.execute(query)
                messages.success(request, 'Other earnings successfully updated.')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
            else:
                rowid = request.POST['rowid']
                query = (f"delete from TBL_OTHER_EARNINGS where rowid = '{rowid}'")
                cursor = connection.cursor()
                cursor.execute(query)
                messages.success(request, 'Other earnings deleted!')
                return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_OTHER_EARNINGS WHERE EMPNO = '{empno}' and status = 0")
    data = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE EMPNO = '{empno}'")
    data1 = cursor.fetchall()
    for r in data1:
        name = str(r[1])+ " - " + r[2] + ", " + r[3]
        position = r[24]
    cursor.execute(f"SELECT * FROM TBL_EARNINGS_TYPE WHERE company = '{company}'")
    dedtype = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE employeetype = '{position}' and company = '{company}' and status = 'ACTIVE'")
    dataperiod = cursor.fetchall()
    period = [""]
    for r in dataperiod:
        period.append(r[3].strftime("%m/%d/%Y") + "-" + r[4].strftime("%m/%d/%Y"))
    return render(request, 'other_earnings.html',{'period': period, 'empno': empno, 'name': name, 'username': username, 'data': data,'data1': data1, 'company': company, 'dedtype': dedtype, 'role1': request.session['role1']})

@login_required 
def checkearnings1(request):
    #Get the variable text
    rowid = request.POST['text']
    #Do whatever with the input variable text
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_OTHER_EARNINGS where rowid = '{rowid}'")
    data = cursor.fetchall()
    for r in data:
            period = r[5].strftime("%m/%d/%Y") + "-" + r[6].strftime("%m/%d/%Y")
            dedtype = r[1]
            dedamount = r[3]
            datemaintained = r[7]
            datelastupdate = r[8]
            user = r[9]

   #Send the response 
    return JsonResponse({'period': period, 'user': user, 'datelastupdate': datelastupdate, 'datemaintained': datemaintained, 'dedtype': dedtype, 'dedamount': dedamount})#amount, noofperiod, dedamount, balance)

@login_required 
def generate(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where status = 'ACTIVE' and company = '{company}'")
    data = cursor.fetchall()
    return render(request, 'generate.html',{'data': data, 'company': company, 'role1': request.session['role1']})  

@login_required 
def generate_payroll(request, rowid):
    company = request.session['company']
    request.session['rowid1'] = rowid
    rowid1 = request.session['rowid1']
    cursor = connection.cursor()
    if request.method == 'POST':
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where rowid = {rowid}")
        data = cursor.fetchall()
        for r in data:
            period = r[11]
            rowid=r[0]
            position = r[2]
        cursor.execute(f"SELECT * FROM TBL_MASTERFILE where position = '{position}' and company = '{company}'")
        data = cursor.fetchall()
        cursor.execute(f"SELECT * FROM TBL_LOCATION where company = '{company}' ORDER BY LOCATION")
        location = cursor.fetchall()
        return render(request, 'generate_payroll.html',{'position': position, 'rowid': rowid, 'period': period, 'data': data, 'company': company, 'location': location, 'role1': request.session['role1']})
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where rowid = {rowid1}")
    data = cursor.fetchall()
    for r in data:
        period = r[11]
        rowid=r[0]
        position = r[2]
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE where position = '{position}' and company = '{company}'")
    data = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_LOCATION where company = '{company}' ORDER BY LOCATION")
    location = cursor.fetchall()
    return render(request, 'generate_payroll.html',{'position': position, 'rowid': rowid1, 'period': period, 'data': data, 'company': company, 'location': location, 'role1': request.session['role1']})


@login_required 
def generate_payroll_driver(request, rowid):
    username = request.session['username']
    company = request.session['company']
    location = request.POST['location']
    if request.method == 'POST':
        cursor = connection.cursor()
        # cursor.execute(f"delete from TBL_OTHER_EARNINGS where amount = 0")
        cursor.execute(f"delete from TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_DEDUCTIONS where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_UNAPPLIED where user = '{username}' and company = '{company}'")
        
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where rowid = {rowid}")
        
        payrollperiod = cursor.fetchall()
        empno = request.POST['empno']
        dateprocess = datetime.now()
        for row in payrollperiod:
            periodfrom = row[3]
            periodto = row[4]
            paymonth = row[6]
            payyear = row[7]
            position = row[2]
            paytype = row[1]
        if request.POST['range'] == "ALL":
            if location == "ALL":
                cursor.execute(f"SELECT driverempno FROM TBL_VTR where periodfrom = '{periodfrom}' and periodto = '{periodto}' and company = '{company}' group by driverempno")
                request.session['generatebutton'] = "ACTIVE"
            else:
                cursor.execute(f"SELECT TBL_VTR.driverempno FROM TBL_VTR inner join TBL_MASTERFILE on TBL_VTR.driverempno = TBL_MASTERFILE.empno\
                                where location = '{request.POST['location']}' and periodfrom = '{periodfrom}' and periodto = '{periodto}' \
                            and TBL_VTR.company = '{company}' group by TBL_VTR.driverempno")
        else:
            cursor.execute(f"SELECT driverempno FROM TBL_VTR where periodfrom = '{periodfrom}' and periodto = '{periodto}' and driverempno = {empno}")
        count  = cursor.rowcount
        if count > 0:
            emplist = cursor.fetchall()
            for row in emplist:
                empno = row[0]
                #collect data per employee, and get total of trips in the selected payroll period
                cursor.execute(f"SELECT lname,fname,mname,tripamount,driverallowance,excesstrip,truckvale,tripeq,atm,sssd,phicd,hdmfd,taxd FROM TBL_VTR \
                            inner join TBL_MASTERFILE on TBL_VTR.driverempno = TBL_MASTERFILE.empno \
                            where TBL_VTR.driverempno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                result = cursor.fetchall()
                basicpay = 0
                allowance = 0
                excesstrip = 0
                truckvale = 0
                trips = 0
                otherearnings = 0
                # otherearnings1 = 0
                for result in result:
                    empname = result[0] + ", " + result[1] + ", " + result[2]
                    atm = result[8]
                    basicpay += float(result[3])
                    allowance += float(result[4])
                    excesstrip += float(result[5])
                    truckvale += float(result[6])
                    trips += float(result[7])
                ############driver as helper###########
                cursor.execute(f"SELECT sum(tripamount) as tripamount, sum(helperallowance) as helperallowance, sum(excesstrip) as excesstrip, \
                            sum(truckvale) as truckvale, sum(tripeq) as tripeq from TBL_HELPER where helperempno = {empno} and periodfrom = '{periodfrom}' and \
                            periodto = '{periodto}'")
                helper = cursor.fetchall()
                for helper in helper:
                    if helper[0] != None:
                        helpertripamount = float(helper[0])
                        helpertruckvale = float(helper[3])
                        helpertripeq = float(helper[4])
                    else:
                        helpertripamount = 0
                        helpertruckvale = 0
                        helpertripeq = 0
                truckvale += helpertruckvale
                basicpay += helpertripamount
                trips += helpertripeq
                ######Other earnings computation##########
                cursor.execute(f"SELECT sum(amount) as earnings from TBL_OTHER_EARNINGS where empno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                earnings = cursor.fetchall()
                for earnings in earnings:
                    if earnings[0] != None:
                        otherearnings = float(earnings[0])
                ######end of Other earnings computation##########
                # otherearnings = allowance + excesstrip + otherearnings1
                grosspay = basicpay + otherearnings
                totaldeductions = 0
                other_deductions = 0
                netpay = grosspay
                cursor.execute(f"SELECT sum(sssee) as totalsss, sum(phic) as totalphic, sum(hdmf) as totalhdmf, sum(grosspay) as \
                            totalgross, sum(ssser) as totalssser, sum(ec) as totalec FROM TBL_FINAL_PAYROLL where empno = {empno} \
                            and paymonth = {paymonth} and payyear = {payyear}")
                govded = cursor.fetchall()
                for govded in govded:
                    if govded[0] == None: totalsss = 0
                    else: totalsss = float(govded[0])
                    if govded[4] == None: totalssser = 0
                    else: totalssser = float(govded[4])
                    if govded[5] == None: totalec = 0
                    else: totalec = float(govded[5])
                    if govded[1] == None: totalphic = 0
                    else: totalphic = float(govded[1])
                    if govded[3] == None: totalgross = 0
                    else: totalgross = float(govded[3])
                totalgross += grosspay
                #sss computation
                if result[9] == "YES":
                    cursor.execute(f"SELECT * FROM TBL_SSS where sssfrom <= {totalgross} and sssto >= {totalgross}")
                    sss = cursor.fetchall()
                    for sss in sss:
                        sssee = float(sss[3]) + float(sss[6]) - float(totalsss)
                        ssser = float(sss[4]) + float(sss[7]) - float(totalssser)
                        ec = float(sss[5]) - float(totalec)
                else:
                    sssee = 0
                    ssser = 0
                    ec = 0
                netpay -= sssee
                totaldeductions += sssee
                #phic vomputation
                if result[10] == "YES":
                    cursor.execute("SELECT * FROM TBL_PHICHDMF")
                    phichdmf = cursor.fetchall()
                    for phichdmf in phichdmf:
                        phicto1 = float(phichdmf[1])
                        phicfrom2 = float(phichdmf[2])
                        phicto2 = float(phichdmf[3])
                        phicfrom3 = float(phichdmf[4])
                        amt1 = float(phichdmf[5])
                        amt2 = float(phichdmf[6])
                        amt3 = float(phichdmf[7])
                        hdmfpercentage = float(phichdmf[8])
                    if totalgross <= phicto1:
                        phicee = round((amt1/2),2) - float(totalphic)
                    elif totalgross >= phicfrom2 and totalgross <= phicto2:
                        phicee = round((((totalgross * amt2)/100)/2),2) - totalphic
                    elif totalgross >= phicfrom3:
                        phicee = round((amt3/2),2)
                else:
                    phicee = 0
                netpay  -= phicee
                totaldeductions += phicee
                #HDMF computation
                if result[11] == "YES":
                    hdmf = round((((grosspay * hdmfpercentage) / 2)/100),2)
                else:
                    hdmf = 0
                netpay -= hdmf
                totaldeductions += hdmf
                #tax
                if result[12] == "YES":
                    tax = 0
                else:
                    tax = 0
                netpay  -= tax
                totaldeductions += tax
                #other deductions
                #note: all other deductions will be save to TBL_TEMP_DEDUCTIONS
                #Truck Vale
                totaldeductions += truckvale
                other_deductions += truckvale
                netpay -= truckvale
                query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
                data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'NONFIXED', 'TRUCK VALE', truckvale, 0, username, company)
                cursor.execute(query,data_values)
                #nonfixed deductions
                cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'NONFIXED' and company = '{company}' order by priority")
                nonfixedlist = cursor.fetchall()
                for nonfixedlist in nonfixedlist:
                    dedname = nonfixedlist[1]
                    cursor.execute(f"SELECT * FROM TBL_NONFIXED_DEDUCTIONS where empno = '{empno}' and dedtype = '{dedname}' and \
                                periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                    nonfixed = cursor.fetchall()
                    for nonfixed in nonfixed:
                        if nonfixed[0] != None:
                            netpay -= float(nonfixed[3])
                            if netpay < 0:
                                netpay += float(nonfixed[3])
                                #############code to save in unapplied deductions########
                                query = ("insert into TBL_TEMP_UNAPPLIED (empno, DEDNAME, periodfrom, periodto, paymonth, payyeaR, amount, ID, COMPANY, DATEMAINTAINED, STATUS, USER)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
                                data_values = (empno, dedname, periodfrom, periodto, paymonth, payyear, nonfixed[3], nonfixed[0], company,dateprocess,0, username)
                                cursor.execute(query,data_values)
                            else:
                                totaldeductions += float(nonfixed[3])
                                other_deductions += float(nonfixed[3])
                                query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
                                data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'NONFIXED', nonfixed[2], nonfixed[3], nonfixed[0], username, company)
                                cursor.execute(query,data_values)
                #fixed deductions
                cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'FIXED' and company = '{company}' order by priority")
                fixedlist = cursor.fetchall()
                for fixedlist in fixedlist:
                    dedname = fixedlist[1]
                    cursor.execute(f"SELECT * FROM TBL_FIXED_DEDUCTIONS where empno = '{empno}' and dedtype = '{dedname}' and datestart < '{dateprocess}' and amtremaining > 0")
                    fixed = cursor.fetchall()
                    for fixed in fixed:
                        if fixed[0] != None:
                            if fixed[13] == "NO":
                                netpay -= float(fixed[5])
                                if netpay < 0:
                                    netpay += float(fixed[5])
                                    #############all unapplied fixed deductions will not be save to unapplied########
                                else:
                                    totaldeductions += float(fixed[5])
                                    other_deductions += float(fixed[5])
                                    query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                                    data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'FIXED', fixed[2], fixed[5], fixed[0], username, company)
                                    cursor.execute(query,data_values)
                #insert to TBL_TEMP_PAYROLL
                query = ("insert into TBL_TEMP_PAYROLL (empno, empname, periodfrom, periodto, paymonth, payyear, paytype, emptype, trips, basicpay, \
                        allowance, excesstrip, grosspay, sssee, dateprocess, user, company, otherearnings, netpay, phic, hdmf, tax, totaldeductions,atm,otherdeductions,location,ssser,ec)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
                data_values = (empno, empname, periodfrom, periodto, paymonth, payyear, paytype, position, trips, basicpay, \
                            allowance, excesstrip, grosspay, sssee, dateprocess, username, company, otherearnings,netpay, phicee, hdmf, tax, totaldeductions, atm,other_deductions,location,ssser,ec)
                cursor.execute(query,data_values)     
            return redirect("show_payroll_summary")
        else:
            messages.info(request, 'No record to generate.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required         
def generate_payroll_helper(request, rowid):
    username = request.session['username']
    company = request.session['company']
    location = request.POST['location']
    if request.method == 'POST':
        cursor = connection.cursor()
        # cursor.execute(f"delete from TBL_OTHER_EARNINGS where amount = 0")
        cursor.execute(f"delete from TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_DEDUCTIONS where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_UNAPPLIED where user = '{username}' and company = '{company}'")
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where rowid = {rowid}")
        
        payrollperiod = cursor.fetchall()
        empno = request.POST['empno']
        dateprocess = datetime.now()
        for row in payrollperiod:
            periodfrom = row[3]
            periodto = row[4]
            paymonth = row[6]
            payyear = row[7]
            position = row[2]
            paytype = row[1]
        if request.POST['range'] == "ALL":
            if location == "ALL":
                cursor.execute(f"SELECT TBL_HELPER.helperempno FROM TBL_HELPER inner join TBL_MASTERFILE on TBL_HELPER.helperempno = TBL_MASTERFILE.empno\
                                where position = '{position}' and periodfrom = '{periodfrom}' and periodto = '{periodto}' \
                            and TBL_HELPER.company = '{company}' group by TBL_HELPER.helperempno")
                request.session['generatebutton'] = "ACTIVE"
            else:
                cursor.execute(f"SELECT TBL_HELPER.helperempno FROM TBL_HELPER inner join TBL_MASTERFILE on TBL_HELPER.helperempno = TBL_MASTERFILE.empno\
                                where position = '{position}' and location = '{request.POST['location']}' and periodfrom = '{periodfrom}' and periodto = '{periodto}' \
                            and TBL_HELPER.company = '{company}' group by TBL_HELPER.helperempno")
        else:
            cursor.execute(f"SELECT helperempno FROM TBL_VTR where periodfrom = '{periodfrom}' and periodto = '{periodto}' and helperempno = {empno}")
        count  = cursor.rowcount
        if count > 0:
            emplist = cursor.fetchall()
            for row in emplist:
                empno = row[0]
                #collect data per employee, and get total of trips in the selected payroll period
                cursor.execute(f"SELECT lname,fname,mname,tripamount,helperallowance,excesstrip,truckvale,tripeq,atm,sssd,phicd,hdmfd,taxd,location FROM TBL_HELPER \
                            inner join TBL_MASTERFILE on TBL_HELPER.helperempno = TBL_MASTERFILE.empno \
                            where TBL_HELPER.helperempno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                result = cursor.fetchall()
                basicpay = 0
                allowance = 0
                excesstrip = 0
                truckvale = 0
                trips = 0
                otherearnings = 0
                # otherearnings1 = 0
                for result in result:
                    empname = result[0] + ", " + result[1] + ", " + result[2]
                    atm = result[8]
                    location = result[13]
                    basicpay += float(result[3])
                    allowance += float(result[4])
                    excesstrip += float(result[5])
                    truckvale += float(result[6])
                    trips += float(result[7])
                cursor.execute(f"SELECT sum(amount) as earnings from TBL_OTHER_EARNINGS where empno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                earnings = cursor.fetchall()
                for earnings in earnings:
                    if earnings[0] != None:
                        otherearnings = float(earnings[0])
                ######end of Other earnings computation##########
                # otherearnings = allowance + excesstrip + otherearnings1
                grosspay = basicpay + otherearnings
                totaldeductions = 0
                other_deductions = 0
                netpay = grosspay
                cursor.execute(f"SELECT sum(sssee) as totalsss, sum(phic) as totalphic, sum(hdmf) as totalhdmf, sum(grosspay) as \
                            totalgross, sum(ssser) as totalssser, sum(ec) as totalec  FROM TBL_FINAL_PAYROLL where empno = {empno} \
                            and paymonth = {paymonth} and payyear = {payyear}")
                govded = cursor.fetchall()
                for govded in govded:
                    if govded[0] == None: totalsss = 0
                    else: totalsss = float(govded[0])
                    if govded[4] == None: totalssser = 0
                    else: totalssser = float(govded[4])
                    if govded[5] == None: totalec = 0
                    else: totalec = float(govded[5])
                    if govded[1] == None: totalphic = 0
                    else: totalphic = float(govded[1])
                    if govded[3] == None: totalgross = 0
                    else: totalgross = float(govded[3])
                totalgross += grosspay
                #sss computation
                if result[9] == "YES":
                    cursor.execute(f"SELECT * FROM TBL_SSS where sssfrom <= {totalgross} and sssto >= {totalgross}")
                    sss = cursor.fetchall()
                    for sss in sss:
                        sssee = float(sss[3]) + float(sss[6]) - float(totalsss)
                        ssser = float(sss[4]) + float(sss[7]) - float(totalssser)
                        ec = float(sss[5]) - float(totalec)
                else:
                    sssee = 0
                    ssser = 0
                    ec = 0
                netpay -= sssee
                totaldeductions += sssee
                #phic vomputation
                if result[10] == "YES":
                    cursor.execute("SELECT * FROM TBL_PHICHDMF")
                    phichdmf = cursor.fetchall()
                    for phichdmf in phichdmf:
                        phicto1 = float(phichdmf[1])
                        phicfrom2 = float(phichdmf[2])
                        phicto2 = float(phichdmf[3])
                        phicfrom3 = float(phichdmf[4])
                        amt1 = float(phichdmf[5])
                        amt2 = float(phichdmf[6])
                        amt3 = float(phichdmf[7])
                        hdmfpercentage = float(phichdmf[8])
                    if totalgross <= phicto1:
                        phicee = round((amt1/2),2) - float(totalphic)
                    elif totalgross >= phicfrom2 and totalgross <= phicto2:
                        phicee = round((((totalgross * amt2)/100)/2),2) - totalphic
                    elif totalgross >= phicfrom3:
                        phicee = round((amt3/2),2)
                else:
                    phicee = 0
                netpay  -= phicee
                totaldeductions += phicee
                #HDMF computation
                if result[11] == "YES":
                    hdmf = round((((grosspay * hdmfpercentage) / 2)/100),2)
                else:
                    hdmf = 0
                netpay = netpay - hdmf
                totaldeductions += hdmf
                #tax
                if result[12] == "YES":
                    tax = 0
                else:
                    tax = 0
                netpay  -= tax
                totaldeductions += tax
                #other deductions
                #note: all other deductions will be save to TBL_TEMP_DEDUCTIONS
                #Truck Vale
                totaldeductions += truckvale
                other_deductions += truckvale
                netpay -= truckvale
                query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
                data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'NONFIXED', 'TRUCK VALE', truckvale, 0, username, company)
                cursor.execute(query,data_values)
                #nonfixed deductions
                cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'NONFIXED' and company = '{company}' order by priority")
                nonfixedlist = cursor.fetchall()
                for nonfixedlist in nonfixedlist:
                    dedname = nonfixedlist[1]
                    cursor.execute(f"SELECT * FROM TBL_NONFIXED_DEDUCTIONS where empno = '{empno}' and dedtype = '{dedname}' and \
                                periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                    nonfixed = cursor.fetchall()
                    for nonfixed in nonfixed:
                        if nonfixed[0] != None:
                            netpay -= float(nonfixed[3])
                            if netpay < 0:
                                netpay += float(nonfixed[3])
                                #############code to save in unapplied deductions########
                                query = ("insert into TBL_TEMP_UNAPPLIED (empno, DEDNAME, periodfrom, periodto, paymonth, payyeaR, amount, ID, COMPANY, DATEMAINTAINED, STATUS, USER)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
                                data_values = (empno, dedname, periodfrom, periodto, paymonth, payyear, nonfixed[3], nonfixed[0], company,dateprocess,0, username)
                                cursor.execute(query,data_values)
                            else:
                                totaldeductions += float(nonfixed[3])
                                other_deductions += float(nonfixed[3])
                                query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
                                data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'NONFIXED', nonfixed[2], nonfixed[3], nonfixed[0], username, company)
                                cursor.execute(query,data_values)
                #fixed deductions
                cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'FIXED' and company = '{company}' order by priority")
                fixedlist = cursor.fetchall()
                for fixedlist in fixedlist:
                    dedname = fixedlist[1]
                    cursor.execute(f"SELECT * FROM TBL_FIXED_DEDUCTIONS where empno = '{empno}' and dedtype = '{dedname}' and datestart < '{dateprocess}' and amtremaining > 0")
                    fixed = cursor.fetchall()
                    for fixed in fixed:
                        if fixed[0] != None:
                            if fixed[13] == "NO":
                                netpay -= float(fixed[5])
                                if netpay < 0:
                                    netpay += float(fixed[5])
                                    #############all unapplied fixed deductions will not be save to unapplied########
                                else:
                                    totaldeductions += float(fixed[5])
                                    other_deductions += float(fixed[5])
                                    query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                                    data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'FIXED', fixed[2], fixed[5], fixed[0], username, company)
                                    cursor.execute(query,data_values)
                #insert to TBL_TEMP_PAYROLL
                query = ("insert into TBL_TEMP_PAYROLL (empno, empname, periodfrom, periodto, paymonth, payyear, paytype, emptype, trips, basicpay, \
                        allowance, excesstrip, grosspay, sssee, dateprocess, user, company, otherearnings, netpay, phic, hdmf, tax, totaldeductions,atm,otherdeductions,location,ssser,ec)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
                data_values = (empno, empname, periodfrom, periodto, paymonth, payyear, paytype, position, trips, basicpay, \
                            allowance, excesstrip, grosspay, sssee, dateprocess, username, company, otherearnings,netpay, phicee, hdmf, tax, totaldeductions, atm,other_deductions,location,ssser,ec)
                cursor.execute(query,data_values)     
            return redirect("show_payroll_summary")
        else:
            messages.info(request, 'No record to generate.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required 
def generate_payroll_office(request, rowid):
    username = request.session['username']
    company = request.session['company']
    location = request.POST['location']
    if request.method == 'POST':
        cursor = connection.cursor()
        empno = request.POST['empno']
        # cursor.execute(f"delete from TBL_OTHER_EARNINGS where amount = 0")
        cursor.execute(f"delete from TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_DEDUCTIONS where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_UNAPPLIED where user = '{username}' and company = '{company}'")
        cursor.execute(f"SELECT * from TBL_RATE")
        rate = cursor.fetchall()
        for rate in rate:
            reg_holiday_rate = float(rate[1])
            reg_holiday_ot_rate = float(rate[2])
            spec_holiday_rate = float(rate[3])
            spec_holiday_ot_rate = float(rate[4])
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where rowid = {rowid}")
        payrollperiod = cursor.fetchall()
        dateprocess = datetime.now()
        for row in payrollperiod:
            periodfrom = row[3]
            periodto = row[4]
            paymonth = row[6]
            payyear = row[7]
            position = row[2]
            paytype = row[1]
        if request.POST['range'] == "ALL":
            if location == "ALL":
                cursor.execute(f"SELECT empno FROM TBL_ATTENDANCE where position = '{position}' and periodfrom = '{periodfrom}' and periodto = '{periodto}' and company = '{company}'")
                request.session['generatebutton'] = "ACTIVE"
            else:
                cursor.execute(f"SELECT TBL_ATTENDANCE.empno FROM TBL_ATTENDANCE inner join TBL_MASTERFILE on TBL_ATTENDANCE.empno = TBL_MASTERFILE.empno\
                                where location = '{request.POST['location']}' and periodfrom = '{periodfrom}' and periodto = '{periodto}' \
                            and TBL_ATTENDANCE.company = '{company}' group by TBL_ATTENDANCE.empno")
        else:
            cursor.execute(f"SELECT empno FROM TBL_ATTENDANCE where periodfrom = '{periodfrom}' and periodto = '{periodto}' and empno = {empno}")
        count  = cursor.rowcount
        if count > 0:
            emplist = cursor.fetchall()
            for row in emplist:
                empno = row[0]
                #collect data per employee, and get total of trips in the selected payroll period
                cursor.execute(f"SELECT lname,fname,mname,allowance,salary,REGDAY,REGOT,REGHOLIDAY,REGHOLIDAYOT,SPECHOLIDAY,SPECHOLIDAYOT,TARDY,atm,sssd,phicd,hdmfd,taxd,location FROM TBL_ATTENDANCE \
                            inner join TBL_MASTERFILE on TBL_ATTENDANCE.empno = TBL_MASTERFILE.empno \
                            where TBL_ATTENDANCE.empno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                result = cursor.fetchall()
                basic_pay = 0
                allowance = 0
                holiday = 0
                ot = 0
                no_days = 0
                daily_rate = 0
                other_earnings = 0
                total_gross = 0
                net_pay = 0
                taxable = 0
                tmp_tax = 0
                tmp_inc = 0
                tax = 0
                for result in result:
                    empname = result[0] + ", " + result[1] + ", " + result[2]
                    atm = result[12]
                    location = result[17]
                    monthly_salary = float(result[4])
                    allowance = float(result[3])
                    regular_day = float(result[5])
                    regular_day_ot = float(result[6])
                    regular_holiday = float(result[7])
                    regular_holiday_ot = float(result[8])
                    special_day = float(result[9])
                    special_day_ot = float(result[10])
                    tardy = float(result[11])
                    daily_rate = monthly_salary/26
                    hourly_rate = daily_rate/8
                    ############computation of salary###########
                    tardy = tardy * hourly_rate
                    no_days = regular_day + regular_holiday + special_day
                    basic_pay = (no_days * daily_rate) - tardy
                    allowance = allowance * no_days
                    holiday = (((reg_holiday_rate - 1) * daily_rate) * regular_holiday) + (((spec_holiday_rate - 1) * daily_rate) * special_day)
                    ot = (regular_day_ot * hourly_rate) + (reg_holiday_ot_rate * regular_holiday_ot * hourly_rate) + (spec_holiday_ot_rate * special_day_ot * hourly_rate)
                    ######Other earnings computation##########
                    cursor.execute(f"SELECT sum(amount) as earnings from TBL_OTHER_EARNINGS where empno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                    earnings = cursor.fetchall()
                    for earnings in earnings:
                        if earnings[0] != None:
                            other_earnings = float(earnings[0])
                    ######end of Other earnings computation##########
                    # otherearnings = allowance + excesstrip + otherearnings1
                    total_gross = basic_pay + holiday + ot + allowance
                    total_deductions = 0
                    other_deductions = 0
                    net_pay = total_gross
                    cursor.execute(f"SELECT sum(sssee) as totalsss, sum(phic) as totalphic, sum(hdmf) as totalhdmf, sum(grosspay) as \
                                totalgross, sum(ssser) as totalssser, sum(ec) as totalec  FROM TBL_FINAL_PAYROLL where empno = {empno} \
                                and paymonth = {paymonth} and payyear = {payyear}")
                    govded = cursor.fetchall()
                    for govded in govded:
                        if govded[0] == None: totalsss = 0
                        else: totalsss = float(govded[0])
                        if govded[4] == None: totalssser = 0
                        else: totalssser = float(govded[4])
                        if govded[5] == None: totalec = 0
                        else: totalec = float(govded[5])
                        if govded[1] == None: totalphic = 0
                        else: totalphic = float(govded[1])
                        if govded[3] == None: prev_total_gross = 0
                        else: prev_total_gross = float(govded[3])
                    prev_total_gross += total_gross
                    #sss computation
                    if result[13] == "YES":
                        cursor.execute(f"SELECT * FROM TBL_SSS where sssfrom <= {prev_total_gross} and sssto >= {prev_total_gross}")
                        sss = cursor.fetchall()
                        for sss in sss:
                            sssee = float(sss[3]) + float(sss[6]) - float(totalsss)
                            ssser = float(sss[4]) + float(sss[7]) - float(totalssser)
                            ec = float(sss[5]) - float(totalec)
                    else:
                        sssee = 0
                        ssser = 0
                        ec = 0
                    net_pay -= sssee
                    total_deductions += sssee
                    #phic vomputation
                    if result[14] == "YES":
                        cursor.execute("SELECT * FROM TBL_PHICHDMF")
                        phichdmf = cursor.fetchall()
                        for phichdmf in phichdmf:
                            phicto1 = float(phichdmf[1])
                            phicfrom2 = float(phichdmf[2])
                            phicto2 = float(phichdmf[3])
                            phicfrom3 = float(phichdmf[4])
                            amt1 = float(phichdmf[5])
                            amt2 = float(phichdmf[6])
                            amt3 = float(phichdmf[7])
                            hdmfpercentage = float(phichdmf[8])
                        if prev_total_gross <= phicto1:
                            phicee = round((amt1/2),2) - float(totalphic)
                        elif prev_total_gross >= phicfrom2 and prev_total_gross <= phicto2:
                            phicee = round((((prev_total_gross * amt2)/100)/2),2) - totalphic
                        elif prev_total_gross >= phicfrom3:
                            phicee = round((amt3/2),2)
                    else:
                        phicee = 0
                    net_pay  -= phicee
                    total_deductions += phicee
                    #HDMF computation
                    if result[15] == "YES":
                        hdmf = round((((prev_total_gross * hdmfpercentage) / 2)/100),2)
                    else:
                        hdmf = 0
                    net_pay -= hdmf
                    taxable = net_pay
                    total_deductions += hdmf
                    ###############tax##############
                    if result[16] == "YES":
                        if paytype == "SEMI-MONTHLY":
                            if taxable < 10418:
                                tax = 0
                            elif taxable > 10416 and taxable < 16667:
                                tmp_tax = 0
                                tmp_inc = (taxable - 10417) * 0.15
                                tax = tmp_tax + tmp_inc
                            elif taxable > 16666 and taxable < 33333:
                                tmp_tax = 937.5
                                tmp_inc = (taxable - 16667) * 0.2
                                tax = tmp_tax + tmp_inc
                            elif taxable > 33332 and taxable < 83333:
                                tmp_tax = 4270.7
                                tmp_inc = (taxable - 33333) * 0.25
                                tax = tmp_tax + tmp_inc
                            elif taxable > 83332 and taxable < 333333:
                                tmp_tax = 16770.7
                                tmp_inc = (taxable - 83333) * 0.3
                                tax = tmp_tax + tmp_inc
                    else:
                        tax = 0
                    net_pay  -= tax
                    total_deductions += tax
                    #nonfixed deductions
                    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'NONFIXED' and company = '{company}' order by priority")
                    nonfixedlist = cursor.fetchall()
                    for nonfixedlist in nonfixedlist:
                        dedname = nonfixedlist[1]
                        cursor.execute(f"SELECT * FROM TBL_NONFIXED_DEDUCTIONS where empno = '{empno}' and dedtype = '{dedname}' and \
                                    periodfrom = '{periodfrom}' and periodto = '{periodto}'")
                        nonfixed = cursor.fetchall()
                        for nonfixed in nonfixed:
                            if nonfixed[0] != None:
                                net_pay -= float(nonfixed[3])
                                if net_pay < 0:
                                    net_pay += float(nonfixed[3])
                                    #############code to save in unapplied deductions########
                                    query = ("insert into TBL_TEMP_UNAPPLIED (empno, DEDNAME, periodfrom, periodto, paymonth, payyeaR, amount, ID, COMPANY, DATEMAINTAINED, STATUS, USER)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                                    data_values = (empno, dedname, periodfrom, periodto, paymonth, payyear, nonfixed[3], nonfixed[0], company,dateprocess,0, username)
                                    cursor.execute(query,data_values)
                                else:
                                    total_deductions += float(nonfixed[3])
                                    other_deductions += float(nonfixed[3])
                                    query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                                    data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'NONFIXED', nonfixed[2], nonfixed[3], nonfixed[0], username, company)
                                    cursor.execute(query,data_values)
                    #fixed deductions
                    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'FIXED' and company = '{company}' order by priority")
                    fixedlist = cursor.fetchall()
                    for fixedlist in fixedlist:
                        dedname = fixedlist[1]
                        cursor.execute(f"SELECT * FROM TBL_FIXED_DEDUCTIONS where empno = '{empno}' and dedtype = '{dedname}' and datestart < '{dateprocess}' and amtremaining > 0")
                        fixed = cursor.fetchall()
                        for fixed in fixed:
                            if fixed[0] != None:
                                if fixed[13] == "NO":
                                    net_pay -= float(fixed[5])
                                    if net_pay < 0:
                                        net_pay += float(fixed[5])
                                        #############all unapplied fixed deductions will not be save to unapplied########
                                    else:
                                        total_deductions += float(fixed[5])
                                        other_deductions += float(fixed[5])
                                        query = ("insert into TBL_TEMP_DEDUCTIONS (empno, paytype, emptype, periodfrom, periodto, paymonth, payyear, dedtype, dedname, amount, id, USER, COMPANY)"
                                                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                            )
                                        data_values = (empno, paytype, position, periodfrom, periodto, paymonth, payyear, 'FIXED', fixed[2], fixed[5], fixed[0], username, company)
                                        cursor.execute(query,data_values)
                    #insert to TBL_TEMP_PAYROLL
                    query = ("insert into TBL_TEMP_PAYROLL (empno, empname, periodfrom, periodto, paymonth, payyear, paytype, emptype, trips, basicpay, \
                            allowance , grosspay, sssee, dateprocess, user, company, otherearnings, netpay, phic, hdmf, tax, totaldeductions,atm,ot,holiday,tardy,otherdeductions,location,ssser,ec)"
                                                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                    )
                    data_values = (empno, empname, periodfrom, periodto, paymonth, payyear, paytype, position, no_days, basic_pay, \
                                allowance, total_gross, sssee, dateprocess, username, company, other_earnings, net_pay, phicee, hdmf, tax, total_deductions, atm, ot, holiday, tardy,other_deductions,location,ssser,ec)
                    cursor.execute(query,data_values)     
            return redirect("show_payroll_summary_admin")
        else:
            messages.info(request, 'No record to generate.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required 
def show_payroll_summary(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
    data = cursor.fetchall()
    return render(request,'show_payroll_summary.html',{'data': data, 'generatebutton': request.session['generatebutton'], 'role1': request.session['role1']})

@login_required 
def show_payroll_summary_admin(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
    data = cursor.fetchall()
    return render(request,'show_payroll_summary_admin.html',{'data': data, 'generatebutton': request.session['generatebutton'], 'role1': request.session['role1']})

@login_required         
def exportvtr(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format()
        bold.set_align('center')
        bold.set_bottom(1)
        bold.set_top(1)
        dateformat = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        dateformat.set_align('left')
        dateformat1 = workbook.add_format({'num_format': 'd mmm yy'})
        dateformat1.set_align('left')
        money = workbook.add_format({'num_format': '#,##0.00'})
        worksheet.set_column('G:G',10,None)
        worksheet.set_column('D:D',10,None)
        cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        row2 = 0 # row count for left part of payslip
        row1 = 1 # use to check if top or bottom of the page
        row3 = 0 # row count for right part of payslip
        for data in data:
            row4 = row2
            period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            periodfrom = data[3]
            periodto = data[4]
            empno = data[1]
            position = data[8]
            if position == "DRIVER":
                ashelper = "(HELPER)"
            else:
                ashelper = ""
            #Left part of payslip
            worksheet.write(row2, 0, "VTR SUMMARY")
            row2+=1
            worksheet.write(row2, 0, "Payroll Period:")
            worksheet.write(row2, 2, period)
            worksheet.write(row2, 5, "Pay Date:")
            worksheet.write(row2, 6, datetime.now(), dateformat)
            row2+=1
            worksheet.write(row2, 0, "Employee No/Name:")
            worksheet.write(row2, 2, str(data[1]) + " - " + data[2])
            row2+=1
            worksheet.write(row2, 0, "VTR #", bold)
            worksheet.write(row2, 1, "PLATE #", bold)
            worksheet.write(row2, 2, "DATE", bold)
            worksheet.write(row2, 3, "DESTINATION", bold)
            worksheet.write(row2, 4, "TRIP EQ", bold)
            worksheet.write(row2, 5, "TRIP RATE", bold)
            worksheet.write(row2, 6, "ALLOWANCE", bold)
            worksheet.write(row2, 7, "TRUCK VALE", bold)
            row2+=1
            cursor.execute(f"SELECT * FROM TBL_VTR where driverempno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}' order by periodfrom")
            vtr = cursor.fetchall()
            for vtr in vtr:
                worksheet.write(row2, 0, vtr[1])
                worksheet.write(row2, 1, vtr[3])
                worksheet.write(row2, 2, vtr[2], dateformat1)
                worksheet.write(row2, 3, vtr[5])
                worksheet.write(row2, 4, vtr[6], money)
                worksheet.write(row2, 5, vtr[9], money)
                worksheet.write(row2, 6, vtr[10], money)
                worksheet.write(row2, 7, vtr[12], money)
                row2+=1
            cursor.execute(f"SELECT TBL_HELPER.vtrno, plateno, vtrdate, locationto, TBL_HELPER.tripeq, TBL_HELPER.tripamount, helperallowance, TBL_HELPER.truckvale \
                        FROM TBL_HELPER inner join  TBL_VTR on TBL_HELPER.vtrno = TBL_VTR.vtrno \
                        where helperempno = {empno} and TBL_HELPER.periodfrom = '{periodfrom}' and TBL_HELPER.periodto = '{periodto}' order by TBL_HELPER.periodfrom")
            vtr = cursor.fetchall()
            for vtr in vtr:
                worksheet.write(row2, 0, vtr[0])
                worksheet.write(row2, 1, vtr[1])
                worksheet.write(row2, 2, vtr[2], dateformat1)
                worksheet.write(row2, 3, vtr[3] + ashelper)
                worksheet.write(row2, 4, vtr[4], money)
                worksheet.write(row2, 5, vtr[5], money)
                worksheet.write(row2, 6, vtr[6], money)
                worksheet.write(row2, 7, vtr[7], money)
                row2+=1
            row2 = row4 + 21       
            if row1 % 2 == 0: #check if top or bottom part of payslip for new line spacing
                row2 += 1
                row3 += 1
            else:
                row2 += 2
                row3 += 2
            row1 += 1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='VTR.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required                               
def exportmbank(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT lname, fname, mname,  TBL_MASTERFILE.atm, netpay FROM TBL_TEMP_PAYROLL inner join TBL_MASTERFILE on \
                TBL_TEMP_PAYROLL.empno = TBL_MASTERFILE.empno where user = '{username}' and TBL_TEMP_PAYROLL.company = '{company}' and \
                TBL_MASTERFILE.atm > 0 order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        money = workbook.add_format({'num_format': '#,##0.00'})
        format = workbook.add_format({'num_format': '@'})
        # format.set_num_format(0) 
        worksheet.set_column('A:A',20,None)
        worksheet.set_column('B:B',20,None)
        worksheet.set_column('C:C',20,None)
        worksheet.set_column('D:D',27,None)
        worksheet.set_column('E:E',20,None)
        
        row2 = 0 # row count for left part of payslip

        worksheet.write(row2, 0, "Last Name")
        worksheet.write(row2, 1, "First Name")
        worksheet.write(row2, 2, "Middle Name")
        worksheet.write(row2, 3, "Employee Account Number")
        worksheet.write(row2, 4, "Amount")
        row2+=1
        for data in data:
            worksheet.write(row2, 0, data[0])
            worksheet.write(row2, 1, data[1])
            worksheet.write(row2, 2, data[2])
            worksheet.write(row2, 3, str(data[3]), format)
            worksheet.write(row2, 4, data[4], money)
            row2+=1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='METROBANK.xlsx')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required
def post_payroll(request):
    username = request.session['username']
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.execute(f"SELECT * from TBL_TEMP_PAYROLL where user = '{username}' limit 1")
        if cursor.rowcount > 0:
            data = cursor.fetchall()
            for data in data:
                position = data[8]
            if position == "DRIVER":
                return post_payroll_driver(request)
            elif position == "HELPER":
                return post_payroll_helper(request)
            elif position == "OFFICE STAFF" or position == "MAINTENANCE":
                return post_payroll_admin(request)
    response = "Payroll Period successfully Posted."
    return HttpResponse(response)
# def post_payroll(request):
#     username = request.session['username']
#     cursor = connection.cursor()
#     cursor.execute(f"SELECT * from TBL_TEMP_PAYROLL where user = '{username}' limit 1")
#     if cursor.rowcount > 0:
#         data = cursor.fetchall()
#         for data in data:
#             position = data[8]
#         if position == "DRIVER":
#             return post_payroll_driver(request)
#         elif position == "HELPER":
#             return post_payroll_helper(request)
#         elif position == "OFFICE STAFF" or position == "MAINTENANCE":
#             return post_payroll_admin(request)
#     return redirect('generate')

@login_required        
def post_payroll_driver(request):   
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * from TBL_TEMP_PAYROLL where user = '{username}'")
    if cursor.rowcount > 0:
        data = cursor.fetchall()
        for data in data:
            cursor.execute(f"UPDATE  TBL_OTHER_EARNINGS set status = 1 where empno = {data[1]} and periodfrom = '{data[3]}' and periodto = '{data[4]}'")
            # cursor.execute(f"UPDATE  TBL_VTR set status = 1 where driverempno = {data[1]} and periodfrom = '{data[3]}' and periodto = '{data[4]}' and company = '{company}'")
            query = ("insert into TBL_FINAL_PAYROLL (EMPNO, EMPNAME, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, PAYTYPE, EMPTYPE, TRIPS, BASICPAY, ALLOWANCE, \
                     EXCESSTRIP, GROSSPAY, WITHHOLDINGTAX, GROSSAFTERTAX, SSSEE, SSSER, PHIC, HDMF, TAX, TOTALDEDUCTIONS, NETPAY, DATEPROCESS, USER, COMPANY,\
                      OTHEREARNINGS,otherdeductions,atm,location,ec)"
                     "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
            data_values = (data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10], data[11], data[12], data[13], \
                           data[14], data[15], data[16], data[17], data[18], data[19], data[20], data[21], data[22], data[23], data[24], data[25], \
                            data[26], data[31], data[27],data[32],data[33])
            cursor.execute(query,data_values)
            cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where EMPNO = {data[1]} and dedtype = 'FIXED' and user = '{username}'")
            fixed = cursor.fetchall()
            if cursor.rowcount > 0:
                for fixed in fixed:
                    cursor.execute(f"SELECT * from TBL_FIXED_DEDUCTIONS where rowid = {fixed[11]}")
                    fixed1 = cursor.fetchall()
                    for fixed1 in fixed1:
                        amtremaining = float(fixed1[6]) - float(fixed[10])
                        periodremaining = float(fixed1[7]) - 1
                        cursor.execute(f"update TBL_FIXED_DEDUCTIONS  set amtremaining = {amtremaining}, periodremaining = {periodremaining} where rowid = {fixed[11]}")
            cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where EMPNO = {data[1]} and dedtype = 'NONFIXED' and user = '{username}'")
            nonfixed = cursor.fetchall()
            if cursor.rowcount > 0:
                for nonfixed in nonfixed:
                    cursor.execute(f"update TBL_NONFIXED_DEDUCTIONS  set status = 1 where rowid = {nonfixed[11]}")
                    cursor.execute(f"update TBL_FINAL_UNAPPLIED  set status = 1 where rowid = {nonfixed[11]}")
        cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where user = '{username}'")
        temp_ded = cursor.fetchall()
        if cursor.rowcount > 0:
            for temp_ded in temp_ded:
                query = ("insert into TBL_FINAL_DEDUCTIONS (EMPNO, PAYTYPE, EMPTYPE, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, DEDTYPE, DEDNAME, AMOUNT, ID, \
                        USER, COMPANY)"
                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                data_values = (temp_ded[1], temp_ded[2], temp_ded[3], temp_ded[4], temp_ded[5], temp_ded[6], temp_ded[7], temp_ded[8], temp_ded[9], temp_ded[10], \
                               temp_ded[11], temp_ded[12], temp_ded[13])
                cursor.execute(query,data_values)
        cursor.execute(f"update TBL_PAYROLL_PERIOD  set status = 'PAID' where periodfrom = '{data[3]}' and periodto = '{data[4]}' and employeetype = '{data[8]}' and company = '{company}'")
        cursor.execute(f"SELECT * from TBL_TEMP_UNAPPLIED where user = '{username}'")
        temp_ua = cursor.fetchall()
        if cursor.rowcount > 0:
            for temp_ua in temp_ua:
                cursor.execute(f"select * from TBL_FINAL_UNAPPLIED where id = {temp_ua[11]}")
                if cursor.rowcount < 1:
                    query = ("insert into TBL_FINAL_UNAPPLIED (EMPNO, DEDNAME, AMOUNT, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, COMPANY, STATUS, DATEMAINTAINED, ID, \
                            USER)"
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                    data_values = (temp_ua[1], temp_ua[2], temp_ua[3], temp_ua[4], temp_ua[5], temp_ua[6], temp_ua[7], temp_ua[8], temp_ua[9], temp_ua[10], \
                                temp_ua[11], temp_ua[12])
                    cursor.execute(query,data_values)
        cursor.execute(f"delete from TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_DEDUCTIONS where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_UNAPPLIED where user = '{username}' and company = '{company}'")    
    return redirect('post_payroll')

@login_required  
def post_payroll_helper(request):   
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * from TBL_TEMP_PAYROLL where user = '{username}'")
    if cursor.rowcount > 0:
        data = cursor.fetchall()
        for data in data:
            cursor.execute(f"UPDATE  TBL_OTHER_EARNINGS set status = 1 where empno = {data[1]} and periodfrom = '{data[3]}' and periodto = '{data[4]}'")
            # cursor.execute(f"UPDATE  TBL_VTR set status = 1 where driverempno = {data[1]} and periodfrom = '{data[3]}' and periodto = '{data[4]}' and company = '{company}'")
            query = ("insert into TBL_FINAL_PAYROLL (EMPNO, EMPNAME, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, PAYTYPE, EMPTYPE, TRIPS, BASICPAY, ALLOWANCE, \
                     EXCESSTRIP, GROSSPAY, WITHHOLDINGTAX, GROSSAFTERTAX, SSSEE, SSSER, PHIC, HDMF, TAX, TOTALDEDUCTIONS, NETPAY, DATEPROCESS, USER, COMPANY,\
                      OTHEREARNINGS,otherdeductions,atm,location,ec)"
                     "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
            data_values = (data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10], data[11], data[12], data[13], \
                           data[14], data[15], data[16], data[17], data[18], data[19], data[20], data[21], data[22], data[23], data[24], data[25], \
                            data[26],data[31],data[27],data[32],data[33])
            cursor.execute(query,data_values)
            cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where EMPNO = {data[1]} and dedtype = 'FIXED' and user = '{username}'")
            fixed = cursor.fetchall()
            if cursor.rowcount > 0:
                for fixed in fixed:
                    cursor.execute(f"SELECT * from TBL_FIXED_DEDUCTIONS where rowid = {fixed[11]}")
                    fixed1 = cursor.fetchall()
                    for fixed1 in fixed1:
                        amtremaining = float(fixed1[6]) - float(fixed[10])
                        periodremaining = float(fixed1[7]) - 1
                        cursor.execute(f"update TBL_FIXED_DEDUCTIONS  set amtremaining = {amtremaining}, periodremaining = {periodremaining} where rowid = {fixed[11]}")
            cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where EMPNO = {data[1]} and dedtype = 'NONFIXED' and user = '{username}'")
            nonfixed = cursor.fetchall()
            if cursor.rowcount > 0:
                for nonfixed in nonfixed:
                    cursor.execute(f"update TBL_NONFIXED_DEDUCTIONS  set status = 1 where rowid = {nonfixed[11]}")
                    cursor.execute(f"update TBL_FINAL_UNAPPLIED  set status = 1 where rowid = {nonfixed[11]}")
        cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where user = '{username}'")
        temp_ded = cursor.fetchall()
        if cursor.rowcount > 0:
            for temp_ded in temp_ded:
                query = ("insert into TBL_FINAL_DEDUCTIONS (EMPNO, PAYTYPE, EMPTYPE, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, DEDTYPE, DEDNAME, AMOUNT, ID, \
                        USER, COMPANY)"
                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                data_values = (temp_ded[1], temp_ded[2], temp_ded[3], temp_ded[4], temp_ded[5], temp_ded[6], temp_ded[7], temp_ded[8], temp_ded[9], temp_ded[10], \
                               temp_ded[11], temp_ded[12], temp_ded[13])
                cursor.execute(query,data_values)
        cursor.execute(f"update TBL_PAYROLL_PERIOD  set status = 'PAID' where periodfrom = '{data[3]}' and periodto = '{data[4]}' and employeetype = '{data[8]}' and company = '{company}'")
        cursor.execute(f"SELECT * from TBL_TEMP_UNAPPLIED where user = '{username}'")
        temp_ua = cursor.fetchall()
        if cursor.rowcount > 0:
            for temp_ua in temp_ua:
                cursor.execute(f"select * from TBL_FINAL_UNAPPLIED where id = {temp_ua[11]}")
                if cursor.rowcount < 1:
                    query = ("insert into TBL_FINAL_UNAPPLIED (EMPNO, DEDNAME, AMOUNT, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, COMPANY, STATUS, DATEMAINTAINED, ID, \
                            USER)"
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                    data_values = (temp_ua[1], temp_ua[2], temp_ua[3], temp_ua[4], temp_ua[5], temp_ua[6], temp_ua[7], temp_ua[8], temp_ua[9], temp_ua[10], \
                                temp_ua[11], temp_ua[12])
                    cursor.execute(query,data_values)
        cursor.execute(f"delete from TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_DEDUCTIONS where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_UNAPPLIED where user = '{username}' and company = '{company}'")       
    return redirect('post_payroll')

@login_required        
def post_payroll_admin(request):   
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * from TBL_TEMP_PAYROLL where user = '{username}'")
    if cursor.rowcount > 0:
        data = cursor.fetchall()
        for data in data:
            cursor.execute(f"UPDATE  TBL_OTHER_EARNINGS set status = 1 where empno = {data[1]} and periodfrom = '{data[3]}' and periodto = '{data[4]}'")
            # cursor.execute(f"UPDATE  TBL_VTR set status = 1 where driverempno = {data[1]} and periodfrom = '{data[3]}' and periodto = '{data[4]}' and company = '{company}'")
            query = ("insert into TBL_FINAL_PAYROLL (EMPNO, EMPNAME, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, PAYTYPE, EMPTYPE, TRIPS, BASICPAY, ALLOWANCE, \
                     EXCESSTRIP, GROSSPAY, WITHHOLDINGTAX, GROSSAFTERTAX, SSSEE, SSSER, PHIC, HDMF, TAX, TOTALDEDUCTIONS, NETPAY, DATEPROCESS, USER, COMPANY,\
                      OTHEREARNINGS,ot,holiday,tardy,otherdeductions,atm,location,ec)"
                     "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
            data_values = (data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8], data[9], data[10], data[11], data[12], data[13], \
                           data[14], data[15], data[16], data[17], data[18], data[19], data[20], data[21], data[22], data[23], data[24], data[25], data[26], \
                            data[28], data[29], data[30], data[31], data[27],data[32],data[33])
            cursor.execute(query,data_values)
            cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where EMPNO = {data[1]} and dedtype = 'FIXED' and user = '{username}'")
            fixed = cursor.fetchall()
            if cursor.rowcount > 0:
                for fixed in fixed:
                    cursor.execute(f"SELECT * from TBL_FIXED_DEDUCTIONS where rowid = {fixed[11]}")
                    fixed1 = cursor.fetchall()
                    for fixed1 in fixed1:
                        amtremaining = float(fixed1[6]) - float(fixed[10])
                        periodremaining = float(fixed1[7]) - 1
                        cursor.execute(f"update TBL_FIXED_DEDUCTIONS  set amtremaining = {amtremaining}, periodremaining = {periodremaining} where rowid = {fixed[11]}")
            cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where EMPNO = {data[1]} and dedtype = 'NONFIXED' and user = '{username}'")
            nonfixed = cursor.fetchall()
            if cursor.rowcount > 0:
                for nonfixed in nonfixed:
                    cursor.execute(f"update TBL_NONFIXED_DEDUCTIONS  set status = 1 where rowid = {nonfixed[11]}")
                    cursor.execute(f"update TBL_FINAL_UNAPPLIED  set status = 1 where rowid = {nonfixed[11]}")
        cursor.execute(f"SELECT * from TBL_TEMP_DEDUCTIONS where user = '{username}'")
        temp_ded = cursor.fetchall()
        if cursor.rowcount > 0:
            for temp_ded in temp_ded:
                query = ("insert into TBL_FINAL_DEDUCTIONS (EMPNO, PAYTYPE, EMPTYPE, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, DEDTYPE, DEDNAME, AMOUNT, ID, \
                        USER, COMPANY)"
                        "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                data_values = (temp_ded[1], temp_ded[2], temp_ded[3], temp_ded[4], temp_ded[5], temp_ded[6], temp_ded[7], temp_ded[8], temp_ded[9], temp_ded[10], \
                               temp_ded[11], temp_ded[12], temp_ded[13])
                cursor.execute(query,data_values)
        cursor.execute(f"update TBL_PAYROLL_PERIOD  set status = 'PAID' where periodfrom = '{data[3]}' and periodto = '{data[4]}' and employeetype = '{data[8]}' and company = '{company}'")
        cursor.execute(f"SELECT * from TBL_TEMP_UNAPPLIED where user = '{username}'")
        temp_ua = cursor.fetchall()
        if cursor.rowcount > 0:
            for temp_ua in temp_ua:
                cursor.execute(f"select * from TBL_FINAL_UNAPPLIED where id = {temp_ua[11]}")
                if cursor.rowcount < 1:
                    query = ("insert into TBL_FINAL_UNAPPLIED (EMPNO, DEDNAME, AMOUNT, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, COMPANY, STATUS, DATEMAINTAINED, ID, \
                            USER)"
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
                    data_values = (temp_ua[1], temp_ua[2], temp_ua[3], temp_ua[4], temp_ua[5], temp_ua[6], temp_ua[7], temp_ua[8], temp_ua[9], temp_ua[10], \
                                temp_ua[11], temp_ua[12])
                    cursor.execute(query,data_values)
        cursor.execute(f"delete from TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_DEDUCTIONS where user = '{username}' and company = '{company}'")
        cursor.execute(f"delete from TBL_TEMP_UNAPPLIED where user = '{username}' and company = '{company}'")     
    return redirect('post_payroll')

@login_required   
def export_summary(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where rowid = {rowid}")
    data = cursor.fetchall()
    for r in data:
        period = r[3].strftime("%m/%d/%Y") + "-" + r[4].strftime("%m/%d/%Y")
        rowid=r[0]
        position = r[2]
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE where position = '{position}' and company = '{company}'")
    data = cursor.fetchall()
    # return render(request, 'generate_payroll.html',{'position': position, 'rowid': rowid, 'period': period, 'data': data, 'company': company})

    return render(request,'export_summary.html',{'role1': request.session['role1']})

@login_required 
def atmpayrollsummary(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where atm > 0 and user = '{username}' and company = '{company}' limit 1")
    data = cursor.fetchall()
    for data in data:
        position = data[8]
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where atm > 0 and user = '{username}' and company = '{company}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        bold1 = workbook.add_format({"bold": True})
        money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        allborder = workbook.add_format({'border':1})
        worksheet.set_column('A:A',30,None)
        #excel header
        worksheet.write('A1', 'PAYROLL - ATM', bold1)
        worksheet.write('C2', 'WE HEREBY ACKNOWLEDGE TO HAVE RECEIVED FROM  the sum specified opposite our respective names as full compensation for our services rendered')
        worksheet.write('A4', 'ATM SUMMARY', bold1)
        worksheet.write('A5', 'Name Of', bold)
        worksheet.write('A6', 'Employee', bold)
        worksheet.write('B5', 'No Of', bold)
        if position == "OFFICE STAFF" or position == "MAINTENANCE":
            worksheet.write('B6', 'Days', bold)
            col = 7
        else:
            worksheet.write('B6', 'Trips', bold)
            col = 8
        worksheet.write('C5', 'Total', bold)
        worksheet.write('C6', 'Gross', bold)
        if position == "DRIVER" or position == "HELPER":
            worksheet.write('D5', 'Truck', bold)
            worksheet.write('D6', 'Vale', bold)
            worksheet.write('E6', 'SSS', bold)
            worksheet.write('F6', 'PHIC', bold)
            worksheet.write('G6', 'HDMF', bold)
            worksheet.write('H6', 'TAX', bold)
        else:
            worksheet.write('D6', 'SSS', bold)
            worksheet.write('E6', 'PHIC', bold)
            worksheet.write('F6', 'HDMF', bold)
            worksheet.write('G6', 'TAX', bold)

        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where company = '{company}' and deduction_type = 'FIXED' and deduction_name <> 'TRUCK VALE' order by priority")
        count = cursor.rowcount
        if count > 0:
            data = cursor.fetchall()
            for data in data:
                worksheet.write(5, col, data[1], bold)
                col+=1
        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where company = '{company}'  and deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' order by priority")
        count = cursor.rowcount
        if count > 0:
            data = cursor.fetchall()
            for data in data:
                worksheet.write(5, col, data[1], bold)
                col+=1
        worksheet.write(4, col, 'Total', bold)
        worksheet.write(5, col, 'Deductions', bold)
        col+=1
        worksheet.write(4, col, 'Total', bold)
        worksheet.write(5, col, 'Net Pay', bold)
        col+=1
        worksheet.write(4, col, 'Signature', bold)
        worksheet.write(5, col, 'Of Employee', bold)
        col+=1
        worksheet.write(4, col, 'Account', bold)
        worksheet.write(5, col, 'Number', bold)
        col+=1
        worksheet.write(5, col, 'EMP No.', bold)
        row = 6
        row1 = 7
        a = []
        cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where atm > 0 and user = '{username}' and company = '{company}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        for data in data:
            col=0
            period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            worksheet.write(row, col, data[2], allborder)
            col+=1
            worksheet.write(row, col, data[9], allborder)
            col+=1
            worksheet.write(row, col, float(data[10]) + float(data[26]), money)
            col+=1
            if position == "DRIVER" or position == "HELPER":
                cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where user = '{username}' and dedname ='TRUCK VALE' and empno = {data[1]}")
                truckvale = cursor.fetchall()
                for truckvale in truckvale:
                    if truckvale[0] != None:
                        worksheet.write(row, col, float(truckvale[0]), money)
                    else:
                        worksheet.write(row, col, "")
                
                col+=1
            worksheet.write(row, col, data[16], money) #sss
            col+=1
            worksheet.write(row, col, data[18], money)  #phic
            col+=1
            worksheet.write(row, col, data[19], money) #hdmf
            col+=1
            worksheet.write(row, col, data[20], money) #tax
            col+=1
            cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'FIXED' and deduction_name <> 'TRUCK VALE' and company = '{company}' order by priority")
            count = cursor.rowcount
            if count > 0:
                fixeddedlist = cursor.fetchall()
                for fixeddedlist in fixeddedlist:
                    cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where dedtype = 'FIXED' and user = '{username}' and dedname ='{fixeddedlist[1]}' and empno = {data[1]}")
                    fixedded = cursor.fetchall()
                    for fixedded in fixedded:
                        if fixedded[0] != None:
                            worksheet.write(row, col, float(fixedded[0]), money)
                        else:
                            worksheet.write(row, col, "", money)
                        col+=1
            cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' and company = '{company}' order by priority")
            count = cursor.rowcount
            if count > 0:
                nonfixeddedlist = cursor.fetchall()
                for nonfixeddedlist in nonfixeddedlist:
                    cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where dedtype = 'NONFIXED' and user = '{username}' and dedname ='{nonfixeddedlist[1]}' and empno = {data[1]}")
                    nonfixedded = cursor.fetchall()
                    for nonfixedded in nonfixedded:
                        if nonfixedded[0] != None:
                            worksheet.write(row, col, float(nonfixedded[0]), money)
                        else:
                            worksheet.write(row, col, "", money)
                        col+=1  
            worksheet.write(row, col, float(data[21]), money) #total deductions
            col+=1  
            worksheet.write(row, col, float(data[22]), money) #net pay  
            col+=1 
            worksheet.write(row, col, "", allborder)
            col+=1 
            worksheet.write(row, col, data[27], allborder)#bank account number
            col+=1  
            worksheet.write(row, col, data[1], allborder)  #empno
            row+=1
            row1+=1
            if row1 == 34:
                row+=1
                a.append(str(row))
                row1 = 1
        if payrollcount <= 27 :
            for i in range(2, col-2,1):
                cell_range = xl_range(6,i,row-1,i)
                worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)
        else:
            for value in a:
                # worksheet.write('A2', str(a), bold1)
                if value == "34":
                    for i in range(2, col-2,1):
                        cell_range = xl_range(6,i,32,i)
                        worksheet.write(33, i, '=SUM('+ str(cell_range) +')', money1)
                else:
                    rowbottom = int(value)
                    for i in range(2, col-2,1):
                        cell_range = xl_range(rowbottom-35,i,rowbottom-2,i)
                        worksheet.write(rowbottom-1, i, '=SUM('+ str(cell_range) +')', money1)
            if row1 < 34:
                rowbottom = row1
                for i in range(2, col-2,1):
                    cell_range = xl_range(row-rowbottom,i,row-1,i)
                    worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)             
        worksheet.write('G1', 'Payroll Period: ' + period, bold1)
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='atm payroll summary.xlsx')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required 
def cashpayrollsummary(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where atm < 1 and user = '{username}' and company = '{company}' limit 1")
    data = cursor.fetchall()
    for data in data:
        position = data[8]
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where atm < 1 and user = '{username}' and company = '{company}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        bold1 = workbook.add_format({"bold": True})
        money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        allborder = workbook.add_format({'border':1})
        worksheet.set_column('A:A',30,None)
        #excel header
        worksheet.write('A1', 'PAYROLL - CASH', bold1)
        worksheet.write('C2', 'WE HEREBY ACKNOWLEDGE TO HAVE RECEIVED FROM  the sum specified opposite our respective names as full compensation for our services rendered')
        worksheet.write('A4', 'CASH SUMMARY', bold1)
        worksheet.write('A5', 'Name Of', bold)
        worksheet.write('A6', 'Employee', bold)
        worksheet.write('B5', 'No Of', bold)
        if position == "OFFICE STAFF" or position == "MAINTENANCE":
            worksheet.write('B6', 'Days', bold)
            col = 7
        else:
            worksheet.write('B6', 'Trips', bold)
            col = 8
        worksheet.write('C5', 'Total', bold)
        worksheet.write('C6', 'Gross', bold)
        if position == "DRIVER" or position == "HELPER":
            worksheet.write('D5', 'Truck', bold)
            worksheet.write('D6', 'Vale', bold)
            worksheet.write('E6', 'SSS', bold)
            worksheet.write('F6', 'PHIC', bold)
            worksheet.write('G6', 'HDMF', bold)
            worksheet.write('H6', 'TAX', bold)
        else:
            worksheet.write('D6', 'SSS', bold)
            worksheet.write('E6', 'PHIC', bold)
            worksheet.write('F6', 'HDMF', bold)
            worksheet.write('G6', 'TAX', bold)
        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where company = '{company}' and deduction_type = 'FIXED' and deduction_name <> 'TRUCK VALE' order by priority")
        count = cursor.rowcount
        if count > 0:
            data = cursor.fetchall()
            for data in data:
                worksheet.write(5, col, data[1], bold)
                col+=1
        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where company = '{company}'  and deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' order by priority")
        count = cursor.rowcount
        if count > 0:
            data = cursor.fetchall()
            for data in data:
                worksheet.write(5, col, data[1], bold)
                col+=1
        worksheet.write(4, col, 'Total', bold)
        worksheet.write(5, col, 'Deductions', bold)
        col+=1
        worksheet.write(4, col, 'Total', bold)
        worksheet.write(5, col, 'Net Pay', bold)
        col+=1
        worksheet.write(4, col, 'Signature', bold)
        worksheet.write(5, col, 'Of Employee', bold)
        col+=1
        worksheet.write(5, col, 'EMP No.', bold)
        row = 6
        row1 = 7
        a = []
        cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where atm < 1 and user = '{username}' and company = '{company}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        for data in data:
            col=0
            period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            worksheet.write(row, col, data[2], allborder)
            col+=1
            worksheet.write(row, col, data[9], allborder)
            col+=1
            worksheet.write(row, col, float(data[10]) + float(data[26]), money)
            col+=1
            if position == "DRIVER" or position == "HELPER":
                cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where user = '{username}' and dedname ='TRUCK VALE' and empno = {data[1]}")
                truckvale = cursor.fetchall()
                for truckvale in truckvale:
                    if truckvale[0] != None:
                        worksheet.write(row, col, float(truckvale[0]), money)
                    else:
                        worksheet.write(row, col, "")
                
                col+=1
            worksheet.write(row, col, data[16], money) #sss
            col+=1
            worksheet.write(row, col, data[18], money)  #phic
            col+=1
            worksheet.write(row, col, data[19], money) #hdmf
            col+=1
            worksheet.write(row, col, data[20], money) #tax
            col+=1
            cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'FIXED' and deduction_name <> 'TRUCK VALE' and company = '{company}' order by priority")
            count = cursor.rowcount
            if count > 0:
                fixeddedlist = cursor.fetchall()
                for fixeddedlist in fixeddedlist:
                    cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where dedtype = 'FIXED' and user = '{username}' and dedname ='{fixeddedlist[1]}' and empno = {data[1]}")
                    fixedded = cursor.fetchall()
                    for fixedded in fixedded:
                        if fixedded[0] != None:
                            worksheet.write(row, col, float(fixedded[0]), money)
                        else:
                            worksheet.write(row, col, "", money)
                        col+=1
            cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' and company = '{company}' order by priority")
            count = cursor.rowcount
            if count > 0:
                nonfixeddedlist = cursor.fetchall()
                for nonfixeddedlist in nonfixeddedlist:
                    cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where dedtype = 'NONFIXED' and user = '{username}' and dedname ='{nonfixeddedlist[1]}' and empno = {data[1]}")
                    nonfixedded = cursor.fetchall()
                    for nonfixedded in nonfixedded:
                        if nonfixedded[0] != None:
                            worksheet.write(row, col, float(nonfixedded[0]), money)
                        else:
                            worksheet.write(row, col, "", money)
                        col+=1  
            worksheet.write(row, col, float(data[21]), money) #total deductions
            col+=1  
            worksheet.write(row, col, float(data[22]), money) #net pay  
            col+=1 
            worksheet.write(row, col, "", allborder)
            col+=1  
            worksheet.write(row, col, data[1], allborder)  #empno
            row+=1
            row1+=1
            if row1 == 34:
                row+=1
                a.append(str(row))
                row1 = 1
        if payrollcount <= 27 :
            for i in range(2, col-2,1):
                cell_range = xl_range(6,i,row-1,i)
                worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)
        else:
            for value in a:
                # worksheet.write('A2', str(a), bold1)
                if value == "34":
                    for i in range(2, col-2,1):
                        cell_range = xl_range(6,i,32,i)
                        worksheet.write(33, i, '=SUM('+ str(cell_range) +')', money1)
                else:
                    rowbottom = int(value)
                    for i in range(2, col-2,1):
                        cell_range = xl_range(rowbottom-35,i,rowbottom-2,i)
                        worksheet.write(rowbottom-1, i, '=SUM('+ str(cell_range) +')', money1)
            if row1 < 34:
                rowbottom = row1
                for i in range(2, col-2,1):
                    cell_range = xl_range(row-rowbottom,i,row-1,i)
                    worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)             
        worksheet.write('G1', 'Payroll Period: ' + period, bold1)
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='cash payroll summary.xlsx')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required 
def exportpayslip(request):
    username = request.session['username']
    company = request.session['company']

    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        dateformat = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        dateformat.set_align('left')
        bold1 = workbook.add_format({"bold": True})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        money2 = workbook.add_format({'num_format': '#,##0.00'})
        money2.set_bottom(3)
        worksheet.set_column('H:H',10,None)
        
        cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        row2 = 0 # row count for left part of payslip
        row1 = 1 # use to check if top or bottom of the page
        row3 = 0 # row count for right part of payslip
        for data in data:
            row4 = row2
            period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            #Left part of payslip
            worksheet.write(row2, 0, company, bold1)
            row2+=1
            worksheet.write(row2, 0, "Payroll Period:")
            worksheet.write(row2, 2, period)
            row2+=3
            worksheet.write(row2, 0, "Employee No/Name:")
            worksheet.write(row2, 2, str(data[1]) + " - " + data[2])
            row2+=1
            worksheet.write(row2, 0, "Earnings:")
            row2+=1
            worksheet.write(row2, 0, "No Of Trips")
            worksheet.write(row2, 3, data[9], money2)
            row2+=1
            worksheet.write(row2, 0, "Basic Pay")
            worksheet.write(row2, 3, data[10], money1)
            row2+=1
            if float(data[26]) > 0:
                worksheet.write(row2, 0, "Other Earnings")
                worksheet.write(row2, 3, data[26], money1)
                row2+=1
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money1)
            row2 = row4 + 19
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money2)
            row2+=1
            worksheet.write(row2, 0, "Net Pay")
            worksheet.write(row2, 3, data[22], money2)
            row2+=1
            worksheet.write(row2, 0, "Received By:")
            
            #right part of payslip
            row3+=1
            worksheet.write(row3, 5, "Pay Date:")
            worksheet.write(row3, 7, datetime.now(), dateformat)
            row3+=3
            cursor.execute(f"SELECT * FROM TBL_MASTERFILE where empno = {data[1]}")
            emp = cursor.fetchall()
            dept = "None"
            for emp in emp:
                dept = emp[23]
            worksheet.write(row3, 6, "Dept:")
            worksheet.write(row3, 7, dept)
            row3+=1
            worksheet.write(row3, 5, "Deductions")
            row3+=1
            if data[16] != 0:
                worksheet.write(row3, 5, "SSS CONTRIBUTION")
                worksheet.write(row3, 8, data[16], money1)
                row3+=1
            if data[18] != 0:
                worksheet.write(row3, 5, "PH CONTRIBUTION")
                worksheet.write(row3, 8, data[18], money1)
                row3+=1
            if data[19] != 0:
                worksheet.write(row3, 5, "PAGIBIG CONTRI")
                worksheet.write(row3, 8, data[19], money1)
                row3+=1
            if data[20] != 0:
                worksheet.write(row3, 5, "TAX")
                worksheet.write(row3, 8, data[20], money1)
                row3+=1
            cursor.execute(f"SELECT dedname FROM TBL_TEMP_DEDUCTIONS where empno = {data[1]} and user = '{username}' and company = '{company}' group by dedname order by dedname")
            deductions = cursor.fetchall()
            count = cursor.rowcount
            if count > 0:
                for deductions in deductions:
                    dedname = deductions[0]
                    cursor.execute(f"SELECT sum(amount) as amount FROM TBL_TEMP_DEDUCTIONS where empno = {data[1]} and user = '{username}' and dedname = '{dedname}'")
                    dedamount = cursor.fetchall()
                    for dedamount in dedamount:
                        if dedamount != None:
                            worksheet.write(row3, 5, dedname)
                            worksheet.write(row3, 8, dedamount[0], money1)
                            row3+=1
            row3 = row4 + 19
            worksheet.write(row3, 5, "Total Deductions")
            worksheet.write(row3, 8, data[21], money2)
            row3 += 2
            worksheet.write(row3, 5, "Date Received:")        
            if row1 % 2 == 0: #check if top or bottom part of payslip for new line spacing
                row2 += 1
                row3 += 1
            else:
                row2 += 2
                row3 += 2
            row1 += 1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='Payslip.xlsx')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required 
def exportpayslipadmin(request):
    username = request.session['username']
    company = request.session['company']

    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        dateformat = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        dateformat.set_align('left')
        bold1 = workbook.add_format({"bold": True})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        money2 = workbook.add_format({'num_format': '#,##0.00'})
        money2.set_bottom(3)
        worksheet.set_column('H:H',10,None)
        
        cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        row2 = 0 # row count for left part of payslip
        row1 = 1 # use to check if top or bottom of the page
        row3 = 0 # row count for right part of payslip
        for data in data:
            row4 = row2
            period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            #Left part of payslip
            worksheet.write(row2, 0, company, bold1)
            row2+=1
            worksheet.write(row2, 0, "Payroll Period:")
            worksheet.write(row2, 2, period)
            row2+=3
            worksheet.write(row2, 0, "Employee No/Name:")
            worksheet.write(row2, 2, str(data[1]) + " - " + data[2])
            row2+=1
            worksheet.write(row2, 0, "Earnings:")
            row2+=1
            worksheet.write(row2, 0, "No Of Days")
            worksheet.write(row2, 3, data[9], money2)
            row2+=1
            worksheet.write(row2, 0, "Basic Pay")
            worksheet.write(row2, 3, data[10], money1)
            row2+=1
            if float(data[11]) > 0:
                worksheet.write(row2, 0, "Allowance")
                worksheet.write(row2, 3, data[11], money1)
                row2+=1
            if float(data[29]) > 0:
                worksheet.write(row2, 0, "Holiday")
                worksheet.write(row2, 3, data[29], money1)
                row2+=1
            if float(data[28]) > 0:
                worksheet.write(row2, 0, "OT")
                worksheet.write(row2, 3, data[28], money1)
                row2+=1
            if float(data[26]) > 0:
                worksheet.write(row2, 0, "Other Earnings")
                worksheet.write(row2, 3, data[26], money1)
                row2+=1
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money1)
            if float(data[30]) > 0:
                row2+=2
                worksheet.write(row2, 0, "Tardy")
                worksheet.write(row2, 3, data[30], money1)
            row2 = row4 + 19
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money2)
            row2+=1
            worksheet.write(row2, 0, "Net Pay")
            worksheet.write(row2, 3, data[22], money2)
            row2+=1
            worksheet.write(row2, 0, "Received By:")
            
            #right part of payslip
            row3+=1
            worksheet.write(row3, 5, "Pay Date:")
            worksheet.write(row3, 7, datetime.now(), dateformat)
            row3+=3
            cursor.execute(f"SELECT * FROM TBL_MASTERFILE where empno = {data[1]}")
            emp = cursor.fetchall()
            dept = "None"
            for emp in emp:
                dept = emp[23]
            worksheet.write(row3, 6, "Dept:")
            worksheet.write(row3, 7, dept)
            row3+=1
            worksheet.write(row3, 5, "Deductions")
            row3+=1
            if data[16] != 0:
                worksheet.write(row3, 5, "SSS CONTRIBUTION")
                worksheet.write(row3, 8, data[16], money1)
                row3+=1
            if data[18] != 0:
                worksheet.write(row3, 5, "PH CONTRIBUTION")
                worksheet.write(row3, 8, data[18], money1)
                row3+=1
            if data[19] != 0:
                worksheet.write(row3, 5, "PAGIBIG CONTRI")
                worksheet.write(row3, 8, data[19], money1)
                row3+=1
            if data[20] != 0:
                worksheet.write(row3, 5, "TAX")
                worksheet.write(row3, 8, data[20], money1)
                row3+=1
            cursor.execute(f"SELECT dedname FROM TBL_TEMP_DEDUCTIONS where empno = {data[1]} and user = '{username}' and company = '{company}' group by dedname order by dedname")
            deductions = cursor.fetchall()
            count = cursor.rowcount
            if count > 0:
                for deductions in deductions:
                    dedname = deductions[0]
                    cursor.execute(f"SELECT sum(amount) as amount FROM TBL_TEMP_DEDUCTIONS where empno = {data[1]} and user = '{username}' and dedname = '{dedname}'")
                    dedamount = cursor.fetchall()
                    for dedamount in dedamount:
                        if dedamount != None:
                            worksheet.write(row3, 5, dedname)
                            worksheet.write(row3, 8, dedamount[0], money1)
                            row3+=1
            row3 = row4 + 19
            worksheet.write(row3, 5, "Total Deductions")
            worksheet.write(row3, 8, data[21], money2)
            row3 += 2
            worksheet.write(row3, 5, "Date Received:")        
            if row1 % 2 == 0: #check if top or bottom part of payslip for new line spacing
                row2 += 1
                row3 += 1
            else:
                row2 += 2
                row3 += 2
            row1 += 1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='Payslip.xlsx')
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


@login_required                                 
def unapplied_deductions(request):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT TBL_TEMP_UNAPPLIED.empno, lname,fname,mname,dedname,amount FROM TBL_TEMP_UNAPPLIED \
                            inner join TBL_MASTERFILE on TBL_TEMP_UNAPPLIED.empno = TBL_MASTERFILE.empno \
                            where user = '{username}' and TBL_TEMP_UNAPPLIED.company = '{company}'")
    data = cursor.fetchall()
    return render(request, 'unapplied_deductions.html',{'data': data, 'role1': request.session['role1']})

@login_required 
def vtr(request):
    username = request.session['username']
    company = request.session['company']
    
    if request.method == 'POST':
        if request.session['vtrcheck'] != 0: #check if vtrno exist, if yes, upload vtr details to the page
            messages.info(request, 'VTR Exist!')
            return redirect(f"edit_vtr/{request.session['vtrcheck']}")#,{'data': data,'data1': data1, 'company': company, 'period': period, 'expenseslist': expenseslist, 'periodvalue': periodvalue})
        
        cursor = connection.cursor()
        vtrno =request.POST['vtrno']
        vtrdate = request.POST['vtrdate']
        plateno = request.POST['plateno'].upper()
        locationfrom = request.POST['locationfrom'].upper()
        locationto = request.POST['locationto'].upper()
        driverempno = request.POST['driverempno']
        if request.POST['drivertripamount'] == "":
            tripamount = 0
        else:
            tripamount = request.POST['drivertripamount']

        if request.POST['truckallowance'] == "":
            allowance = 0
        else:
            allowance = request.POST['truckallowance']

        if request.POST['tripequivalent'] == "":
                tripeq = 0
        else:
            tripeq = request.POST['tripequivalent']

        if request.POST['driverallowance'] == "":
            driverallowance = 0
        else:
            driverallowance = request.POST['driverallowance']

        if request.POST['driverexcesstrip'] == "":
            excesstrip = 0
        else:
            excesstrip = request.POST['driverexcesstrip']
                
        if request.POST['drivertruckvale'] == "":
            truckvale = 0
        else:
            truckvale = request.POST['drivertruckvale']
                
        payrollperiod = request.POST['payrollperiod']
        periodfrom = payrollperiod[6:10]+"-"+payrollperiod[0:2]+"-"+payrollperiod[3:5]
        periodto = payrollperiod[17:21]+"-"+payrollperiod[11:13]+"-"+payrollperiod[14:16]
        dh = 0
        datemaintained = datetime.now()
        #############check vtrdate if under the selected payroll period############
        #############start - get payroll payrollperiod details - start###############
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE periodfrom = '{periodfrom}' and periodto = '{periodto}' and employeetype = 'DRIVER'")
        data = cursor.fetchall()
        for r in data:
            paymonth = r[6]
            payyear = r[7]
        #############end - get payroll payrollperiod details - end###############
        helpers = "(" + request.POST['helpers'] + ")"
        helpers = eval(helpers)
        expenseslist = "(" + request.POST['expenseslist'] + ")"
        expenseslist = eval(expenseslist)
                    
        ############START - SAVE HELPER DETAILS TO TBL_HELPER - START###########
        i = 1
        if helpers != ('m','m','m','m','m','m'):
            while i < len(helpers):
                helperempno = helpers[i][0]
                helpername = helpers[i][1]
                helpertripamount= helpers[i][2]
                helperallowance = helpers[i][3]
                helperexcesstrip = helpers[i][4]
                helpertruckvale = helpers[i][5]
                ############start - check helper if driver - start##################
                cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE empno = {helperempno}")
                data = cursor.fetchall()
                for r in data:
                        position = r[24]
                if position == 'DRIVER':
                    dh += 1
                ############end - check helper if driver - end##################
                    
                query = ("insert into TBL_HELPER (VTRNO, HELPEREMPNO, helpername, TRIPAMOUNT, HELPERALLOWANCE, EXCESSTRIP, TRUCKVALE, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, USER, COMPANY, tripeq)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
                data_values = (vtrno, helperempno, helpername, helpertripamount, helperallowance, helperexcesstrip, helpertruckvale, periodfrom, periodto, paymonth, payyear, username, company, tripeq )
                cursor.execute(query,data_values)
                if helperallowance != "0":
                    query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate, status)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                    data_values = (vtrno,'ALLOWANCE', helperempno, helperallowance, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
                    cursor.execute(query,data_values)
                if helperexcesstrip != "0":
                    query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate,status)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                    data_values = (vtrno,'EXCESS TRIP', helperempno, helperexcesstrip, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
                    cursor.execute(query,data_values)
                i += 1
                ############END - SAVE HELPER DETAILS TO TBL_HELPER - END######
                ############START - SAVE TRIP EXPENSES TO TBL_VTR_EXPENSES - START############
        i = 1
        if expenseslist != ('m','m'):
            while i < len(expenseslist):
                expensesname = expenseslist[i][0]
                expensesamount= expenseslist[i][1]
                            
                query = ("insert into TBL_VTR_EXPENSES (vtrno, EXPENSES, AMOUNT, COMPANY)"
                                                    "VALUES (%s, %s, %s, %s)"
                                                )
                data_values = (vtrno, expensesname, expensesamount, company)
                cursor.execute(query,data_values)
                i += 1
            ############END - SAVE TRIP EXPENSES TO TBL_VTR_EXPENSES-END############
            ############start - save vtr to TBL_VTR - start########################
        query = ("insert into TBL_VTR (VTRNO, vtrdate, plateno, locationfrom,locationto, tripeq, allowance, driverempno, tripamount, driverallowance, excesstrip, truckvale, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, DH)"
                                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                            )
        data_values = (vtrno, vtrdate, plateno, locationfrom,locationto, tripeq, allowance, driverempno, tripamount, driverallowance, excesstrip, truckvale, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, dh)
        cursor.execute(query,data_values)
        ############end - save vtr to TBL_VTR - end##########################
        ##############save to other earnings#################################
        if driverallowance != 0:
            query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate, status)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
            data_values = (vtrno,'ALLOWANCE', driverempno, driverallowance, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
            cursor.execute(query,data_values)
        if excesstrip != 0:
            query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate,status)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
            data_values = (vtrno,'EXCESS TRIP', driverempno, excesstrip, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
            cursor.execute(query,data_values)
        ############3########################################################
        cursor.execute(f"SELECT * FROM TBL_TRICKS where username = '{username}'")
        count  = cursor.rowcount
        if count > 0 :
            query = f"update TBL_TRICKS set payrollperiod = '{payrollperiod}'"
            cursor.execute(query)
        else:
            query = ("insert into TBL_TRICKS (username, payrollperiod)"
                                                "VALUES (%s, %s)"
                                            )
            data_values = (username, payrollperiod)
            cursor.execute(query,data_values)
            ###################################################################
        messages.info(request, 'VTR successfully added!')
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    request.session['vtrcheck'] = 0
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE COMPANY = '{company}' and status = 'ACTIVE' and position = 'DRIVER' order by lname")
    data = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE COMPANY = '{company}' and status = 'ACTIVE' and (position = 'DRIVER' or Position = 'HELPER') order by lname")
    data1 = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE employeetype = 'Driver' and company = '{company}' and status = 'ACTIVE'")
    period = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_EXPENSES_LIST where company = '{company}'")
    expenseslist = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_TRICKS where username = '{username}'")
    periodlist = cursor.fetchall()
    count  = cursor.rowcount
    if count > 0 :
        for r in periodlist:
            periodvalue = r[2]
    else:
        periodvalue = "None"
    return render(request, 'vtr.html',{'data': data,'data1': data1, 'company': company, 'period': period, 'expenseslist': expenseslist, 'periodvalue': periodvalue, 'role1': request.session['role1']}) #,{'data': data, 'company': company}

@login_required 
def testcall(request):
    #Get the variable text
    company = request.session['company']
    vtrno = request.POST['text']
    #Do whatever with the input variable text
    request.session['vtrcheck'] = 0
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_VTR where VTRNO = {vtrno} and company = '{company}'")
    count  = cursor.rowcount
    if count > 0 :
        response = "YES"
        request.session['vtrcheck'] = vtrno
    else:
        response = "NO"
        request.session['vtrcheck'] = 0
   #Send the response 
    return HttpResponse(response)

@login_required 
def testvtr(request):
    #Get the variable text
    company = request.session['company']
    vtrno = request.POST['text']
    #Do whatever with the input variable text
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_VTR where VTRNO = {vtrno} and company = '{company}'")
    count  = cursor.rowcount
    if count > 0 :
        response = "YES"
    else:
        response = "NO"
   #Send the response 
    return HttpResponse(response)

@login_required 
def getrate(request):
    #Get the variable text
    empno = request.POST['text']
    #Do whatever with the input variable text
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE where empno = {empno}")
    data = cursor.fetchall()
    for data in data:
        response = int(data[26])
   #Send the response 
    return HttpResponse(response)

@login_required 
def edit_vtr(request,vtrcheck):
    username = request.session['username']
    company = request.session['company']
    if request.method == 'POST':
        cursor = connection.cursor()
        vtrno =request.POST['vtrno']
        vtrdate = request.POST['vtrdate']
        plateno = request.POST['plateno'].upper()
        locationfrom = request.POST['locationfrom'].upper()
        locationto = request.POST['locationto'].upper()
        driverempno = request.POST['driverempno']
        if request.POST['drivertripamount'] == "":
            tripamount = 0
        else:
            tripamount = request.POST['drivertripamount']

        if request.POST['truckallowance'] == "":
            allowance = 0
        else:
            allowance = request.POST['truckallowance']

        if request.POST['tripequivalent'] == "":
            tripeq = 0
        else:
            tripeq = request.POST['tripequivalent']

        if request.POST['driverallowance'] == "":
            driverallowance = 0
        else:
            driverallowance = request.POST['driverallowance']

        if request.POST['driverexcesstrip'] == "":
            excesstrip = 0
        else:
            excesstrip = request.POST['driverexcesstrip']
            
        if request.POST['drivertruckvale'] == "":
            truckvale = 0
        else:
            truckvale = request.POST['drivertruckvale']

        payrollperiod = request.POST['payrollperiod']
        periodfrom = payrollperiod[6:10]+"-"+payrollperiod[0:2]+"-"+payrollperiod[3:5]
        periodto = payrollperiod[17:21]+"-"+payrollperiod[11:13]+"-"+payrollperiod[14:16]
        dh = 0
        datemaintained = datetime.now()
        #############check vtrdate if under the selected payroll period############
        #############start - get payroll payrollperiod details - start###############
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE periodfrom = '{periodfrom}' and periodto = '{periodto}' and employeetype = 'DRIVER'")
        data = cursor.fetchall()
        for r in data:
            paymonth = r[6]
            payyear = r[7]
        #############end - get payroll payrollperiod details - end###############
        helpers = "(" + request.POST['helpers'] + ")"
        helpers = eval(helpers)
        expenseslist = "(" + request.POST['expenseslist'] + ")"
        expenseslist = eval(expenseslist)
                
        ############START - SAVE HELPER DETAILS TO TBL_HELPER - START###########
        cursor.execute(f"delete FROM TBL_HELPER WHERE vtrno = {vtrcheck} and company = '{company}'")
        cursor.execute(f"delete FROM TBL_OTHER_EARNINGS WHERE vtrno = {vtrcheck} and company = '{company}'")
        i = 1
        if helpers != ('m','m','m','m','m','m'):
            while i < len(helpers):
                helperempno = helpers[i][0]
                helpername = helpers[i][1]
                helpertripamount= helpers[i][2]
                helperallowance = helpers[i][3]
                helperexcesstrip = helpers[i][4]
                helpertruckvale = helpers[i][5]
                ############start - check helper if driver - start##################
                cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE empno = {helperempno}")
                data = cursor.fetchall()
                for r in data:
                        position = r[24]
                if position == 'DRIVER':
                    dh += 1
                ############end - check helper if driver - end##################

                query = ("insert into TBL_HELPER (VTRNO, HELPEREMPNO, helpername, TRIPAMOUNT, HELPERALLOWANCE, EXCESSTRIP, TRUCKVALE, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, USER, COMPANY, tripeq)"
                                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                            )
                data_values = (vtrno, helperempno, helpername, helpertripamount, helperallowance, helperexcesstrip, helpertruckvale, periodfrom, periodto, paymonth, payyear, username, company, tripeq )
                cursor.execute(query,data_values)
                if helperallowance != 0:
                    query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate, status)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                    data_values = (vtrno,'ALLOWANCE', helperempno, helperallowance, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
                    cursor.execute(query,data_values)
                if helperexcesstrip != 0:
                    query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate,status)"
                                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                        )
                    data_values = (vtrno,'EXCESS TRIP', helperempno, helperexcesstrip, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
                    cursor.execute(query,data_values)
                i += 1
                ############END - SAVE HELPER DETAILS TO TBL_HELPER - END######
                ############START - SAVE TRIP EXPENSES TO TBL_VTR_EXPENSES - START############
        cursor.execute(f"delete FROM TBL_VTR_EXPENSES WHERE vtrno = {vtrcheck}  and company = '{company}'")
        i = 1
        if expenseslist != ('m','m'):
            while i < len(expenseslist):
                expensesname = expenseslist[i][0]
                expensesamount= expenseslist[i][1]
                        
                query = ("insert into TBL_VTR_EXPENSES (vtrno, EXPENSES, AMOUNT, company)"
                                                "VALUES (%s, %s, %s, %s)"
                                            )
                data_values = (vtrno, expensesname, expensesamount,company)
                cursor.execute(query,data_values)
                i += 1
            ############END - SAVE TRIP EXPENSES TO TBL_VTR_EXPENSES-END############
            ############start - save vtr to TBL_VTR - start########################
        cursor.execute(f"delete FROM TBL_VTR WHERE vtrno = {vtrcheck} and company = '{company}'")
        query = ("insert into TBL_VTR (VTRNO, vtrdate, plateno, locationfrom,locationto, tripeq, allowance, driverempno, tripamount, driverallowance, excesstrip, truckvale, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, DH)"
                                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                        )
        data_values = (vtrno, vtrdate, plateno, locationfrom,locationto, tripeq, allowance, driverempno, tripamount, driverallowance, excesstrip, truckvale, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, dh)
        cursor.execute(query,data_values)
        if driverallowance != 0:
            query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate, status)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
            data_values = (vtrno,'ALLOWANCE', driverempno, driverallowance, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
            cursor.execute(query,data_values)
        if excesstrip != 0:
            query = ("insert into TBL_OTHER_EARNINGS (VTRNO, earningstype, empno, amount, periodfrom, periodto, paymonth, payyear, user, company, datemaintained, datelastupdate,status)"
                                                    "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                                                )
            data_values = (vtrno,'EXCESS TRIP', driverempno, excesstrip, periodfrom, periodto, paymonth, payyear, username, company, datemaintained, datemaintained, 0 )
            cursor.execute(query,data_values)
        ############end - save vtr to TBL_VTR - end##########################
        messages.info(request, 'VTR successfully updated')
        return redirect("vtr")
        ##################
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE COMPANY = '{company}' and status = 'ACTIVE' and position = 'DRIVER' order by lname")
    data = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE COMPANY = '{company}' and status = 'ACTIVE' and (position = 'DRIVER' or Position = 'HELPER') order by lname")
    data1 = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_EXPENSES_LIST where company = '{company}'")
    expenseslist = cursor.fetchall()
    cursor.execute(f"SELECT * FROM TBL_TRICKS where username = '{username}'")
    periodlist = cursor.fetchall()
    count  = cursor.rowcount
    if count > 0 :
        for r in periodlist:
            periodvalue = r[2]
    else:
        periodvalue = "None"
    cursor.execute(f"SELECT * FROM TBL_HELPER WHERE vtrno = {vtrcheck} and company = '{company}'")
    helperlist = cursor.fetchall()
    allhelper = "('m','m','m','m','m','m')"
    for row in helperlist:
        helper = "('" + str(row[2]) + "','" + row[3]  + "','" + str(row[4])  + "','" + str(row[5])  + "','" + str(row[6]) + "','" + str(row[7])  + "')"
        allhelper = allhelper + "," + helper
    allhelper = str(allhelper)
    cursor.execute(f"SELECT * FROM TBL_VTR_EXPENSES WHERE vtrno = {vtrcheck} and company = '{company}'")
    expenseslist1 = cursor.fetchall()
    period = [""]
    cursor.execute(f"SELECT * FROM TBL_VTR WHERE vtrno = {vtrcheck} and company = '{company}'")
    vtr = cursor.fetchall()
    for row in vtr:
        vtrdate = str(row[2])
        vtrdate = vtrdate[0:10]
        empno = row[8]
        period.append(row[13].strftime("%m/%d/%Y") + "-" + row[14].strftime("%m/%d/%Y"))
        oldperiod = row[13].strftime("%m/%d/%Y") + "-" + row[14].strftime("%m/%d/%Y")
    cursor.execute(f"SELECT * FROM TBL_MASTERFILE WHERE empno = {empno}")
    masterfile = cursor.fetchall()
    for row in masterfile:
        drivername = row[2] + ", " + row[3]
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE employeetype = 'Driver' and company = '{company}' and status = 'ACTIVE'")
    period = cursor.fetchall()

    return render(request, 'edit_vtr.html',{'data': data,'data1': data1, 'company': company, 'period': period,
                                            'expenseslist': expenseslist, 'periodvalue': periodvalue,
                                            'helperlist': helperlist, 'expenseslist1': expenseslist1,
                                                'vtr': vtr, 'vtrdate': vtrdate, 'drivername': drivername,
                                                'allhelper': allhelper, 'oldperiod': oldperiod, 'role1': request.session['role1'] }) #,{'data': data, 'company': company}

@login_required 
def delete_vtr(request,vtrno):
    company = request.session['company']
    if request.method == 'POST':
        cursor = connection.cursor()
        cursor.execute(f"delete FROM TBL_VTR where vtrno = {vtrno} AND COMPANY = '{company}'")
        cursor.execute(f"delete FROM TBL_HELPER where vtrno = {vtrno} AND COMPANY = '{company}'")
        cursor.execute(f"delete FROM TBL_VTR_EXPENSES where vtrno = {vtrno} AND COMPANY = '{company}'")
        cursor.execute(f"delete FROM TBL_OTHER_EARNINGS where vtrno = {vtrno} AND COMPANY = '{company}'")
        return redirect("vtr")

@login_required
def attendance(request):   
    company = request.session['company']
    cursor = connection.cursor()
    if request.method == "POST" :
        emptype = request.POST['dept']
        period = request.POST['period']
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE payrollperiod = '{period}' and employeetype = '{emptype}' and company = '{company}' and status = 'ACTIVE'")
        period = cursor.fetchall()
        for periods in period:
            request.session['rowid'] = periods[0]
        cursor.execute(f"SELECT rowid, empno, lname, fname, mname FROM TBL_MASTERFILE WHERE POSITION = '{emptype}' and company = '{company}' and status = 'ACTIVE'")
        emp = cursor.fetchall()
        request.session['emptype'] = emptype        
        return render(request, 'add_attendance.html',{'emp': emp, 'period': period, 'role1': request.session['role1']})
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE employeetype = 'OFFICE STAFF' and company = '{company}' and status = 'ACTIVE'")
    period = cursor.fetchall()
    return render(request, 'select_attendance.html',{'period': period, 'role1': request.session['role1']})

@login_required
def closeattendance(request):   
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE employeetype = 'OFFICE STAFF' and company = '{company}' and status = 'ACTIVE'")
    period = cursor.fetchall()
    return render(request, 'select_attendance.html',{'period': period, 'role1': request.session['role1']})

@login_required
def add_attendance(request):   
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    if request.method == "POST":
        if request.POST['empno'] == "":
            messages.error(request, "No employee selected.")
            return redirect("add_attendance")
        empno = request.POST['empno']
        if request.POST['regday'] == "": regday = 0
        else: regday = request.POST['regday']
        if request.POST['regot'] == "": regot = 0
        else: regot = request.POST['regot']
        if request.POST['reghday'] == "": reghday = 0
        else: reghday = request.POST['reghday']
        if request.POST['reghot'] == "": reghot = 0
        else: reghot = request.POST['reghot'] 
        if request.POST['specday'] == "": specday = 0
        else: specday = request.POST['specday'] 
        if request.POST['specot'] == "": specot = 0
        else: specot = request.POST['specot'] 
        if request.POST['tardy'] == "": tardy = 0
        else: tardy = request.POST['tardy']
        emptype = request.session['emptype']
        rowid = request.session['rowid']
        datemaintained = datetime.now()
        cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE rowid = {rowid}")
        period = cursor.fetchall()
        for period in period:
            periodfrom = period[3]
            periodto = period[4]
            paymonth = period[6]
            payyear = period[7]
            position = period[2]
        cursor.execute(f"SELECT * FROM TBL_ATTENDANCE WHERE EMPNO = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
        if cursor.rowcount > 0:
            attendance = cursor.fetchall()
            for attendance in attendance:
                id = attendance[0]
            cursor.execute(f"update TBL_ATTENDANCE  set REGDAY = {regday}, REGOT = {regot}, REGHOLIDAY = {reghday}, REGHOLIDAYOT = {reghot}, \
                        SPECHOLIDAY = {specday}, SPECHOLIDAYOT = {specot}, TARDY = {tardy}, PERIODFROM = '{periodfrom}', PERIODTO = '{periodto}', \
                        PAYMONTH = {paymonth}, PAYYEAR = {payyear}, DATEMAINTAINED = '{datemaintained}', USER = '{username}', COMPANY = '{company}', \
                           position = '{position}' WHERE rowid = {id}")
            messages.error(request, "Attendance updated.")
        else:
            query = ("insert into TBL_ATTENDANCE (empno, REGDAY, REGOT, REGHOLIDAY, REGHOLIDAYOT, SPECHOLIDAY, SPECHOLIDAYOT, TARDY, PERIODFROM, PERIODTO, \
                        PAYMONTH, PAYYEAR, DATEMAINTAINED, USER, COMPANY, position)"
                            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                        )
            data_values = (empno, regday, regot, reghday, reghot, specday, specot, tardy, periodfrom, periodto, paymonth, payyear, datemaintained, username, company, position )
            cursor.execute(query,data_values)
            messages.error(request, "Attendance added.")
        return redirect("add_attendance")
    
    emptype = request.session['emptype']
    rowid = request.session['rowid']
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE rowid = {rowid}")
    period = cursor.fetchall()
    cursor.execute(f"SELECT rowid, empno, lname, fname, mname FROM TBL_MASTERFILE WHERE POSITION = '{emptype}' and company = '{company}' and status = 'ACTIVE'")
    emp = cursor.fetchall()
    return render(request, 'add_attendance.html',{'emp': emp, 'period': period, 'emptype': emptype, 'role1': request.session['role1']})

@login_required
def deleteattendance(request):
    cursor = connection.cursor()
    if request.method == "POST":
        id = request.POST['checkdelete']
        cursor.execute(f"delete FROM TBL_ATTENDANCE WHERE rowid = {id}")
        messages.error(request, "Attendance deleted.")
        return redirect("add_attendance")

@login_required
def checkattendance(request):
    #Get the variable text
    rowid = int(request.POST['rowid'])
    text = request.POST['text']
    #Do whatever with the input variable text
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD WHERE rowid = {rowid}")
    period = cursor.fetchall()
    for period in period:
        periodfrom = period[3]
        periodto = period[4]
    cursor.execute(f"SELECT * FROM TBL_ATTENDANCE WHERE EMPNO = {text} and periodfrom = '{periodfrom}' and periodto = '{periodto}'")
    if cursor.rowcount > 0:
        attendance = cursor.fetchall()
        for attendance in attendance:
            id = attendance[0]
            regday = attendance[2]
            regot = attendance[3] 
            reghday = attendance[4] 
            reghot = attendance[5] 
            specday = attendance[6] 
            specot = attendance[7] 
            tardy = attendance[8]
            checkvalue = 'YES'
    else:
        id = 0
        regday = 0
        regot = 0
        reghday = 0 
        reghot = 0
        specday = 0 
        specot = 0
        tardy = 0
        checkvalue = 'NO'
   #Send the response 
    return JsonResponse({'id': id,'regday': regday, 'regot': regot, 'reghday': reghday, 'reghot': reghot, 'specday': specday, 'specot': specot, 'tardy': tardy, 'checkvalue': checkvalue})


@login_required
def getperiod(request):
    #Get the variable text
    emptype = request.POST['text']
    company = request.session['company']
    response = []
    i = 1
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_PAYROLL_PERIOD where employeetype = '{emptype}' and company = '{company}' and status ='ACTIVE'")
    period = cursor.fetchall()
    for period in period:
        if i == 1:
            response = period[11]
        else:
            response = response + "," + period[11]
        i += 1
    return HttpResponse(response)

@login_required
def rate(request):
    if request.method == "POST" :
        regday = request.POST['regday']
        regot = request.POST['regot']
        specday = request.POST['specday']
        specot = request.POST['specot']
        cursor = connection.cursor()
        cursor.execute(f"update TBL_RATE set regholiday = '{regday}', regholidayot = '{regot}', specialday = '{specday}', \
                       specialdayot = '{specot}'")
        messages.error(request, "Rate successfully updated.")
        return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_RATE")
    rate = cursor.fetchall()
    return render(request, 'rate.html',{'rate': rate, 'role1': request.session['role1']}) #,{'data': data, 'company': company}

def simulatesummary(request):
    username = request.session['username']
    company = request.session['company']
    role1 = request.session['role1']
    cursor = connection.cursor()
    
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()
    
    bold = workbook.add_format({"bold": True, 'border':1})
    bold.set_align('center')
    bold1 = workbook.add_format({"bold": True})
    money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
    money1 = workbook.add_format({'num_format': '#,##0.00'})
    allborder = workbook.add_format({'border':1})
    worksheet.set_column('A:A',30,None)
    #excel header
    worksheet.write('A1', 'PAYROLL - ATM', bold1)
    worksheet.write('C2', 'WE HEREBY ACKNOWLEDGE TO HAVE RECEIVED FROM  the sum specified opposite our respective names as full compensation for our services rendered')
    worksheet.write('A4', 'ATM SUMMARY', bold1)
    worksheet.write('A5', 'Name Of', bold)
    worksheet.write('A6', 'Employee', bold)
    worksheet.write('B5', 'No Of', bold)
    worksheet.write('B6', 'Trips', bold)
    worksheet.write('C5', 'Total', bold)
    worksheet.write('C6', 'Gross', bold)
    worksheet.write('D5', 'Truck', bold)
    worksheet.write('D6', 'Vale', bold)
    worksheet.write('E6', 'SSS', bold)
    worksheet.write('F6', 'PHIC', bold)
    worksheet.write('G6', 'HDMF', bold)
    worksheet.write('H6', 'TAX', bold)
    col = 8
    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where company = '{company}' and deduction_type = 'FIXED' and deduction_name <> 'TRUCK VALE' order by priority")
    count = cursor.rowcount
    if count > 0:
        data = cursor.fetchall()
        for data in data:
            worksheet.write(5, col, data[1], bold)
            col+=1
    cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where company = '{company}'  and deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' order by priority")
    count = cursor.rowcount
    if count > 0:
        data = cursor.fetchall()
        for data in data:
            worksheet.write(5, col, data[1], bold)
            col+=1
    worksheet.write(4, col, 'Total', bold)
    worksheet.write(5, col, 'Deductions', bold)
    col+=1
    worksheet.write(4, col, 'Total', bold)
    worksheet.write(5, col, 'Net Pay', bold)
    col+=1
    worksheet.write(4, col, 'Signature', bold)
    worksheet.write(5, col, 'Of Employee', bold)
    col+=1
    worksheet.write(4, col, 'Account', bold)
    worksheet.write(5, col, 'Number', bold)
    col+=1
    worksheet.write(5, col, 'EMP No.', bold)
    row = 6
    row1 = 7
    a = []
    # cursor.execute(f"SELECT * FROM TBL_TEMP_PAYROLL where user = '{username}' and company = '{company}' order by empname")
    # data = cursor.fetchall()
    payrollcount = 67
    for i in range(1,payrollcount,1):
        col=0
        # period = 123
        worksheet.write(row, col, 123, allborder)
        col+=1
        worksheet.write(row, col, 123, allborder)
        col+=1
        worksheet.write(row, col, 123, money)
        col+=1
        # cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where user = '{username}' and dedname ='TRUCK VALE' and empno = {data[1]}")
        # truckvale = cursor.fetchall()
        # for truckvale in truckvale:
        #     if truckvale[0] != None:
        #         worksheet.write(row, col, float(truckvale[0]), money)
        #     else:
        worksheet.write(row, col, 123)
        
        col+=1
        worksheet.write(row, col, 123, money) #sss
        col+=1
        worksheet.write(row, col, 123, money)  #phic
        col+=1
        worksheet.write(row, col, 123, money) #hdmf
        col+=1
        worksheet.write(row, col, 123, money) #tax
        col+=1
        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'FIXED' and deduction_name <> 'TRUCK VALE' and company = '{company}' order by priority")
        count = cursor.rowcount
        if count > 0:
            fixeddedlist = cursor.fetchall()
            for fixeddedlist in fixeddedlist:
                # cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where dedtype = 'FIXED' and user = '{username}' and dedname ='{fixeddedlist[1]}' and empno = {data[1]}")
                # fixedded = cursor.fetchall()
                # for fixedded in fixedded:
                    # if fixedded[0] != None:
                    #     worksheet.write(row, col, float(fixedded[0]), money)
                    # else:
                    worksheet.write(row, col, 123, money)
                    col+=1
        cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' and company = '{company}' order by priority")
        count = cursor.rowcount
        if count > 0:
            nonfixeddedlist = cursor.fetchall()
            for nonfixeddedlist in nonfixeddedlist:
                # cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_TEMP_DEDUCTIONS where dedtype = 'NONFIXED' and user = '{username}' and dedname ='{nonfixeddedlist[1]}' and empno = {data[1]}")
                # nonfixedded = cursor.fetchall()
                # for nonfixedded in nonfixedded:
                #     if nonfixedded[0] != None:
                #         worksheet.write(row, col, float(nonfixedded[0]), money)
                #     else:
                    worksheet.write(row, col, 123, money)
                    col+=1  
        worksheet.write(row, col, 123, money) #total deductions
        col+=1  
        worksheet.write(row, col, 123, money) #net pay  
        col+=1 
        worksheet.write(row, col, "", allborder)
        col+=1 
        worksheet.write(row, col, 123, allborder)#bank account number
        col+=1  
        worksheet.write(row, col, 123, allborder)  #empno
        row+=1
        row1+=1
        if row1 == 34:
            row+=1
            a.append(str(row))
            row1 = 1
    if payrollcount <= 27 :
        for i in range(2, col-2,1):
            cell_range = xl_range(6,i,row-1,i)
            worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)
    else:
        for value in a:
            # worksheet.write('A2', str(a), bold1)
            if value == "34":
                for i in range(2, col-2,1):
                    cell_range = xl_range(6,i,32,i)
                    worksheet.write(33, i, '=SUM('+ str(cell_range) +')', money1)
            else:
                rowbottom = int(value)
                
                for i in range(2, col-2,1):
                    cell_range = xl_range(rowbottom-35,i,rowbottom-2,i)
                    worksheet.write(rowbottom-1, i, '=SUM('+ str(cell_range) +')', money1)
        if row1 < 34:
            rowbottom = row1
            for i in range(2, col-2,1):
                cell_range = xl_range(row-rowbottom,i,row-1,i)
                worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)        
    worksheet.write('G1', "123", bold1)
    workbook.close()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='atm payroll summary.xlsx')

@login_required 
def clear_table(request):
    if request.method == 'POST':
        if request.POST['tablename'] != "":
            tablename = request.POST['tablename'].upper()
            cursor = connection.cursor()
            cursor.execute(f"delete from {tablename}")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if request.POST['password2'] != "":
            empno = request.POST['password2']
            cursor = connection.cursor()
            cursor.execute(f"delete from TBL_MASTERFILE where empno = {empno}")
            cursor.execute(f"delete from TBL_FINAL_DEDUCTIONS where empno = {empno}")
            cursor.execute(f"delete from TBL_FINAL_PAYROLL where empno = {empno}")
            cursor.execute(f"delete from TBL_FIXED_DEDUCTIONS where empno = {empno}")
            cursor.execute(f"delete from TBL_HELPER where helperempno = {empno}")
            cursor.execute(f"delete from TBL_OTHER_EARNINGS where empno = {empno}")
            cursor.execute(f"delete from TBL_VTR where driverempno = {empno}")
            
    
    return render(request, 'clear_table.html',{'role1': request.session['role1']})

@login_required 
def gov_ded(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"Select payyear from TBL_PAYROLL_PERIOD where company = '{company}' and status = 'PAID' group by payyear order by payyear desc")
    year = cursor.fetchall()
    month = []
    for i in range(1, 13,1):
        month.append(str(i))
    
    return render(request, 'gov_ded.html',{'month': month, 'year': year, 'role1': request.session['role1']})

@login_required 
def deduction_reports(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"Select payyear from TBL_PAYROLL_PERIOD where company = '{company}' and status = 'PAID' group by payyear order by payyear desc")
    year = cursor.fetchall()
    cursor.execute(f"Select deduction_name from TBL_DEDUCTION_TYPE where company = '{company}' group by deduction_name order by deduction_name")
    dedtype = cursor.fetchall()
    month = []
    for i in range(1, 13,1):
        month.append(str(i))
    return render(request, 'deduction_reports.html',{'month': month, 'year': year, 'dedtype': dedtype, 'role1': request.session['role1']})

@login_required 
def payroll_reports(request):
    company = request.session['company']
    cursor = connection.cursor()
    cursor.execute(f"Select location from TBL_LOCATION where company = '{company}' order by location")
    location = cursor.fetchall()
    # month = []
    # for i in range(1, 13,1):
    #     month.append(str(i))
    return render(request, 'payroll_reports.html',{'location': location, 'role1': request.session['role1']})

@login_required 
def get_year(request):
    #Get the variable text
    company = request.session['company']
    emp_type = request.POST['text']
    #Do whatever with the input variable text
    response = []
    i = 1
    cursor = connection.cursor()
    cursor.execute(f"SELECT PAYYEAR FROM TBL_PAYROLL_PERIOD where status = 'PAID' and employeetype = '{emp_type}' and company = '{company}' group by payyear order by payyear desc")
    pay_year = cursor.fetchall()
    for pay_year in pay_year:
        if i == 1:
            response = pay_year[0]
        else:
            response = response + "," + pay_year[0]
        i += 1
    return HttpResponse(response)

@login_required 
def get_month(request):
    #Get the variable text
    company = request.session['company']
    emp_type = request.POST['text']
    pay_year = request.POST['pay_year']
    #Do whatever with the input variable text
    response = []
    i = 1
    cursor = connection.cursor()
    cursor.execute(f"SELECT paymonth FROM TBL_PAYROLL_PERIOD where status = 'PAID' and payyear = {pay_year} and employeetype = '{emp_type}' and company = '{company}' group by paymonth order by paymonth desc")
    pay_year = cursor.fetchall()
    for pay_year in pay_year:
        if i == 1:
            response = pay_year[0]
        else:
            response = str(response) + "," + str(pay_year[0])
        i += 1
    return HttpResponse(response)

@login_required 
def get_payroll_period(request):
    #Get the variable text
    company = request.session['company']
    emp_type = request.POST['text']
    pay_year = request.POST['pay_year']
    pay_month = request.POST['pay_month']
    #Do whatever with the input variable text
    period = []
    i = 1
    cursor = connection.cursor()
    cursor.execute(f"SELECT rowid, payrollperiod FROM TBL_PAYROLL_PERIOD where status = 'PAID' and paymonth = {pay_month} and payyear = {pay_year} and employeetype = '{emp_type}' and company = '{company}' order by rowid")
    pay_period = cursor.fetchall()
    for pay_period in pay_period:
        if i == 1:
            period = pay_period[1]
        else:
            period = period + "," + pay_period[1]
        i += 1
    return JsonResponse({'period': period})

@login_required 
def posted_summary_reports(request):
    company = request.session['company']
    position = request.POST['emptype']
    period = request.POST['period']
    location = request.POST['location']
    cursor = connection.cursor()
    cursor.execute(f"SELECT periodfrom, periodto FROM TBL_PAYROLL_PERIOD where payrollperiod = '{period}' and employeetype = '{position}' and company = '{company}' limit 1")
    data = cursor.fetchall()
    for data in data:
        periodfrom = data[0]
        periodto = data[1]
    if location == "ALL":
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
    else:
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        bold1 = workbook.add_format({"bold": True})
        money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        allborder = workbook.add_format({'border':1})
        worksheet.set_column('A:A',30,None)
        #excel header
        worksheet.write('A1', 'PAYROLL - SUMMARY', bold1)
        worksheet.write('C2', 'WE HEREBY ACKNOWLEDGE TO HAVE RECEIVED FROM  the sum specified opposite our respective names as full compensation for our services rendered')
        worksheet.write('A4', 'SUMMARY', bold1)
        worksheet.write('A5', 'Name Of', bold)
        worksheet.write('A6', 'Employee', bold)
        worksheet.write('B5', 'No Of', bold)
        if position == "OFFICE STAFF" or position == "MAINTENANCE":
            worksheet.write('B6', 'Days', bold)
            col = 7
        else:
            worksheet.write('B6', 'Trips', bold)
            col = 8
        worksheet.write('C5', 'Total', bold)
        worksheet.write('C6', 'Gross', bold)
        if position == "DRIVER" or position == "HELPER":
            worksheet.write('D5', 'Truck', bold)
            worksheet.write('D6', 'Vale', bold)
            worksheet.write('E6', 'SSS', bold)
            worksheet.write('F6', 'PHIC', bold)
            worksheet.write('G6', 'HDMF', bold)
            worksheet.write('H6', 'TAX', bold)
        else:
            worksheet.write('D6', 'SSS', bold)
            worksheet.write('E6', 'PHIC', bold)
            worksheet.write('F6', 'HDMF', bold)
            worksheet.write('G6', 'TAX', bold)

        cursor.execute(f"SELECT deduction_name FROM TBL_DEDUCTION_TYPE where company = '{company}' and deduction_name <> 'TRUCK VALE' group by deduction_name order by deduction_name")
        count = cursor.rowcount
        if count > 0:
            data = cursor.fetchall()
            for data in data:
                worksheet.write(5, col, data[0], bold)
                col+=1
        # cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where company = '{company}'  and deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' order by priority")
        # count = cursor.rowcount
        # if count > 0:
        #     data = cursor.fetchall()
        #     for data in data:
        #         worksheet.write(5, col, data[1], bold)
        #         col+=1
        worksheet.write(4, col, 'Total', bold)
        worksheet.write(5, col, 'Deductions', bold)
        col+=1
        worksheet.write(4, col, 'Total', bold)
        worksheet.write(5, col, 'Net Pay', bold)
        col+=1
        worksheet.write(4, col, 'Signature', bold)
        worksheet.write(5, col, 'Of Employee', bold)
        col+=1
        worksheet.write(4, col, 'Account', bold)
        worksheet.write(5, col, 'Number', bold)
        col+=1
        worksheet.write(5, col, 'EMP No.', bold)
        row = 6
        row1 = 7
        a = []
        if location == "ALL":
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
        else:
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        for data in data:
            col=0
            # period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            worksheet.write(row, col, data[2], allborder)
            col+=1
            worksheet.write(row, col, data[9], allborder)
            col+=1
            worksheet.write(row, col, float(data[13]), money)
            col+=1
            if position == "DRIVER" or position == "HELPER":
                cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_FINAL_DEDUCTIONS where periodfrom = '{periodfrom}' and periodto = '{periodto}' and dedname ='TRUCK VALE' and empno = {data[1]}")
                truckvale = cursor.fetchall()
                for truckvale in truckvale:
                    if truckvale[0] != None:
                        worksheet.write(row, col, float(truckvale[0]), money)
                    else:
                        worksheet.write(row, col, "")
                
                col+=1
            worksheet.write(row, col, data[16], money) #sss
            col+=1
            worksheet.write(row, col, data[18], money)  #phic
            col+=1
            worksheet.write(row, col, data[19], money) #hdmf
            col+=1
            worksheet.write(row, col, data[20], money) #tax
            col+=1
            cursor.execute(f"SELECT deduction_name FROM TBL_DEDUCTION_TYPE where company = '{company}' and deduction_name <> 'TRUCK VALE' group by deduction_name order by deduction_name")
            count = cursor.rowcount
            if count > 0:
                fixeddedlist = cursor.fetchall()
                for fixeddedlist in fixeddedlist:
                    cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_FINAL_DEDUCTIONS where periodfrom = '{periodfrom}' and periodto = '{periodto}' and dedname ='{fixeddedlist[0]}' and empno = {data[1]}")
                    fixedded = cursor.fetchall()
                    for fixedded in fixedded:
                        if fixedded[0] != None:
                            worksheet.write(row, col, float(fixedded[0]), money)
                        else:
                            worksheet.write(row, col, "", money)
                        col+=1
            # cursor.execute(f"SELECT * FROM TBL_DEDUCTION_TYPE where deduction_type = 'NONFIXED' and deduction_name <> 'TRUCK VALE' and company = '{company}' order by priority")
            # count = cursor.rowcount
            # if count > 0:
            #     nonfixeddedlist = cursor.fetchall()
            #     for nonfixeddedlist in nonfixeddedlist:
            #         cursor.execute(f"SELECT sum(AMOUNT) as amount FROM TBL_FINAL_DEDUCTIONS where dedtype = 'NONFIXED' and periodfrom = '{periodfrom}' and periodto = '{periodto}' and dedname ='{nonfixeddedlist[1]}' and empno = {data[1]}")
            #         nonfixedded = cursor.fetchall()
            #         for nonfixedded in nonfixedded:
            #             if nonfixedded[0] != None:
            #                 worksheet.write(row, col, float(nonfixedded[0]), money)
            #             else:
            #                 worksheet.write(row, col, "", money)
            #             col+=1  
            worksheet.write(row, col, float(data[21]), money) #total deductions
            col+=1  
            worksheet.write(row, col, float(data[22]), money) #net pay  
            col+=1 
            worksheet.write(row, col, "", allborder)
            col+=1 
            worksheet.write(row, col, data[27], allborder)#bank account number
            col+=1  
            worksheet.write(row, col, data[1], allborder)  #empno
            row+=1
            row1+=1
            if row1 == 34:
                row+=1
                a.append(str(row))
                row1 = 1
        if payrollcount <= 27 :
            for i in range(2, col-2,1):
                cell_range = xl_range(6,i,row-1,i)
                worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)
        else:
            for value in a:
                # worksheet.write('A2', str(a), bold1)
                if value == "34":
                    for i in range(2, col-2,1):
                        cell_range = xl_range(6,i,32,i)
                        worksheet.write(33, i, '=SUM('+ str(cell_range) +')', money1)
                else:
                    rowbottom = int(value)
                    for i in range(2, col-2,1):
                        cell_range = xl_range(rowbottom-35,i,rowbottom-2,i)
                        worksheet.write(rowbottom-1, i, '=SUM('+ str(cell_range) +')', money1)
            if row1 < 34:
                rowbottom = row1
                for i in range(2, col-2,1):
                    cell_range = xl_range(row-rowbottom,i,row-1,i)
                    worksheet.write(row, i, '=SUM('+ str(cell_range) +')', money1)             
        worksheet.write('G1', 'Payroll Period: ' + period, bold1)
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='payroll summary.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required 
def posted_admin_payslip_reports(request):
    company = request.session['company']
    position = request.POST['emptype']
    period = request.POST['period']
    location = request.POST['location']
    cursor = connection.cursor()
    cursor.execute(f"SELECT periodfrom, periodto FROM TBL_PAYROLL_PERIOD where payrollperiod = '{period}' and employeetype = '{position}' and company = '{company}' limit 1")
    data = cursor.fetchall()
    for data in data:
        periodfrom = data[0]
        periodto = data[1]
    if location == "ALL":
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
    else:
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        dateformat = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        dateformat.set_align('left')
        bold1 = workbook.add_format({"bold": True})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        money2 = workbook.add_format({'num_format': '#,##0.00'})
        money2.set_bottom(3)
        worksheet.set_column('H:H',10,None)
        
        if location == "ALL":
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
        else:
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        row2 = 0 # row count for left part of payslip
        row1 = 1 # use to check if top or bottom of the page
        row3 = 0 # row count for right part of payslip
        for data in data:
            row4 = row2
            # period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            #Left part of payslip
            worksheet.write(row2, 0, company, bold1)
            row2+=1
            worksheet.write(row2, 0, "Payroll Period:")
            worksheet.write(row2, 2, period)
            row2+=3
            worksheet.write(row2, 0, "Employee No/Name:")
            worksheet.write(row2, 2, str(data[1]) + " - " + data[2])
            row2+=1
            worksheet.write(row2, 0, "Earnings:")
            row2+=1
            worksheet.write(row2, 0, "No Of Days")
            worksheet.write(row2, 3, data[9], money2)
            row2+=1
            worksheet.write(row2, 0, "Basic Pay")
            worksheet.write(row2, 3, data[10], money1)
            row2+=1
            if float(data[11]) > 0:
                worksheet.write(row2, 0, "Allowance")
                worksheet.write(row2, 3, data[11], money1)
                row2+=1
            if float(data[29]) > 0:
                worksheet.write(row2, 0, "Holiday")
                worksheet.write(row2, 3, data[29], money1)
                row2+=1
            if float(data[28]) > 0:
                worksheet.write(row2, 0, "OT")
                worksheet.write(row2, 3, data[28], money1)
                row2+=1
            if float(data[26]) > 0:
                worksheet.write(row2, 0, "Other Earnings")
                worksheet.write(row2, 3, data[26], money1)
                row2+=1
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money1)
            if float(data[30]) > 0:
                row2+=2
                worksheet.write(row2, 0, "Tardy")
                worksheet.write(row2, 3, data[30], money1)
            row2 = row4 + 19
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money2)
            row2+=1
            worksheet.write(row2, 0, "Net Pay")
            worksheet.write(row2, 3, data[22], money2)
            row2+=1
            worksheet.write(row2, 0, "Received By:")
            
            #right part of payslip
            row3+=1
            worksheet.write(row3, 5, "Pay Date:")
            worksheet.write(row3, 7, datetime.now(), dateformat)
            row3+=3
            cursor.execute(f"SELECT * FROM TBL_MASTERFILE where empno = {data[1]}")
            emp = cursor.fetchall()
            dept = "None"
            for emp in emp:
                dept = emp[23]
            worksheet.write(row3, 6, "Dept:")
            worksheet.write(row3, 7, dept)
            row3+=1
            worksheet.write(row3, 5, "Deductions")
            row3+=1
            if data[16] != 0:
                worksheet.write(row3, 5, "SSS CONTRIBUTION")
                worksheet.write(row3, 8, data[16], money1)
                row3+=1
            if data[18] != 0:
                worksheet.write(row3, 5, "PH CONTRIBUTION")
                worksheet.write(row3, 8, data[18], money1)
                row3+=1
            if data[19] != 0:
                worksheet.write(row3, 5, "PAGIBIG CONTRI")
                worksheet.write(row3, 8, data[19], money1)
                row3+=1
            if data[20] != 0:
                worksheet.write(row3, 5, "TAX")
                worksheet.write(row3, 8, data[20], money1)
                row3+=1
            cursor.execute(f"SELECT dedname FROM TBL_FINAL_DEDUCTIONS where empno = {data[1]} and periodfrom = '{periodfrom}' and periodto = '{periodto}' group by dedname order by dedname")
            deductions = cursor.fetchall()
            count = cursor.rowcount
            if count > 0:
                for deductions in deductions:
                    dedname = deductions[0]
                    cursor.execute(f"SELECT sum(amount) as amount FROM TBL_FINAL_DEDUCTIONS where empno = {data[1]} and periodfrom = '{periodfrom}' and periodto = '{periodto}' and dedname = '{dedname}'")
                    dedamount = cursor.fetchall()
                    for dedamount in dedamount:
                        if dedamount != None:
                            worksheet.write(row3, 5, dedname)
                            worksheet.write(row3, 8, dedamount[0], money1)
                            row3+=1
            row3 = row4 + 19
            worksheet.write(row3, 5, "Total Deductions")
            worksheet.write(row3, 8, data[21], money2)
            row3 += 2
            worksheet.write(row3, 5, "Date Received:")        
            if row1 % 2 == 0: #check if top or bottom part of payslip for new line spacing
                row2 += 1
                row3 += 1
            else:
                row2 += 2
                row3 += 2
            row1 += 1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='Payslip.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required 
def posted_delivery_payslip_reports(request):
    company = request.session['company']
    position = request.POST['emptype']
    period = request.POST['period']
    location = request.POST['location']
    cursor = connection.cursor()
    cursor.execute(f"SELECT periodfrom, periodto FROM TBL_PAYROLL_PERIOD where payrollperiod = '{period}' and employeetype = '{position}' and company = '{company}' limit 1")
    data = cursor.fetchall()
    for data in data:
        periodfrom = data[0]
        periodto = data[1]
    if location == "ALL":
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
    else:
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        dateformat = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        dateformat.set_align('left')
        bold1 = workbook.add_format({"bold": True})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        money2 = workbook.add_format({'num_format': '#,##0.00'})
        money2.set_bottom(3)
        worksheet.set_column('H:H',10,None)
        
        if location == "ALL":
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
        else:
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        row2 = 0 # row count for left part of payslip
        row1 = 1 # use to check if top or bottom of the page
        row3 = 0 # row count for right part of payslip
        for data in data:
            row4 = row2
            # period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            #Left part of payslip
            worksheet.write(row2, 0, company, bold1)
            row2+=1
            worksheet.write(row2, 0, "Payroll Period:")
            worksheet.write(row2, 2, period)
            row2+=3
            worksheet.write(row2, 0, "Employee No/Name:")
            worksheet.write(row2, 2, str(data[1]) + " - " + data[2])
            row2+=1
            worksheet.write(row2, 0, "Earnings:")
            row2+=1
            worksheet.write(row2, 0, "No Of Trips")
            worksheet.write(row2, 3, data[9], money2)
            row2+=1
            worksheet.write(row2, 0, "Basic Pay")
            worksheet.write(row2, 3, data[10], money1)
            row2+=1
            if float(data[26]) > 0:
                worksheet.write(row2, 0, "Other Earnings")
                worksheet.write(row2, 3, data[26], money1)
                row2+=1
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money1)
            row2 = row4 + 19
            worksheet.write(row2, 0, "Total Gross Pay")
            worksheet.write(row2, 3, data[13], money2)
            row2+=1
            worksheet.write(row2, 0, "Net Pay")
            worksheet.write(row2, 3, data[22], money2)
            row2+=1
            worksheet.write(row2, 0, "Received By:")
            
            #right part of payslip
            row3+=1
            worksheet.write(row3, 5, "Pay Date:")
            worksheet.write(row3, 7, datetime.now(), dateformat)
            row3+=3
            cursor.execute(f"SELECT * FROM TBL_MASTERFILE where empno = {data[1]}")
            emp = cursor.fetchall()
            dept = "None"
            for emp in emp:
                dept = emp[23]
            worksheet.write(row3, 6, "Dept:")
            worksheet.write(row3, 7, dept)
            row3+=1
            worksheet.write(row3, 5, "Deductions")
            row3+=1
            if data[16] != 0:
                worksheet.write(row3, 5, "SSS CONTRIBUTION")
                worksheet.write(row3, 8, data[16], money1)
                row3+=1
            if data[18] != 0:
                worksheet.write(row3, 5, "PH CONTRIBUTION")
                worksheet.write(row3, 8, data[18], money1)
                row3+=1
            if data[19] != 0:
                worksheet.write(row3, 5, "PAGIBIG CONTRI")
                worksheet.write(row3, 8, data[19], money1)
                row3+=1
            if data[20] != 0:
                worksheet.write(row3, 5, "TAX")
                worksheet.write(row3, 8, data[20], money1)
                row3+=1
            cursor.execute(f"SELECT dedname FROM TBL_FINAL_DEDUCTIONS where empno = {data[1]} and periodfrom = '{periodfrom}' and periodto = '{periodto}' and company = '{company}' group by dedname order by dedname")
            deductions = cursor.fetchall()
            count = cursor.rowcount
            if count > 0:
                for deductions in deductions:
                    dedname = deductions[0]
                    cursor.execute(f"SELECT sum(amount) as amount FROM TBL_FINAL_DEDUCTIONS where empno = {data[1]} and periodfrom = '{periodfrom}' and periodto = '{periodto}' and dedname = '{dedname}'")
                    dedamount = cursor.fetchall()
                    for dedamount in dedamount:
                        if dedamount != None:
                            worksheet.write(row3, 5, dedname)
                            worksheet.write(row3, 8, dedamount[0], money1)
                            row3+=1
            row3 = row4 + 19
            worksheet.write(row3, 5, "Total Deductions")
            worksheet.write(row3, 8, data[21], money2)
            row3 += 2
            worksheet.write(row3, 5, "Date Received:")        
            if row1 % 2 == 0: #check if top or bottom part of payslip for new line spacing
                row2 += 1
                row3 += 1
            else:
                row2 += 2
                row3 += 2
            row1 += 1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='Payslip.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required         
def posted_driver_vtr_reports(request):
    company = request.session['company']
    position = request.POST['emptype']
    period = request.POST['period']
    location = request.POST['location']
    cursor = connection.cursor()
    cursor.execute(f"SELECT periodfrom, periodto FROM TBL_PAYROLL_PERIOD where payrollperiod = '{period}' and employeetype = '{position}' and company = '{company}' limit 1")
    data = cursor.fetchall()
    for data in data:
        periodfrom = data[0]
        periodto = data[1]
    if location == "ALL":
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
    else:
        cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format()
        bold.set_align('center')
        bold.set_bottom(1)
        bold.set_top(1)
        dateformat = workbook.add_format({'num_format': 'mm/dd/yyyy'})
        dateformat.set_align('left')
        dateformat1 = workbook.add_format({'num_format': 'd mmm yy'})
        dateformat1.set_align('left')
        money = workbook.add_format({'num_format': '#,##0.00'})
        worksheet.set_column('G:G',10,None)
        worksheet.set_column('D:D',10,None)
        if location == "ALL":
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' order by empname")
        else:
            cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where periodfrom = '{periodfrom}' and periodto = '{periodto}' and emptype = '{position}' \
                       and company = '{company}' and location = '{location}' order by empname")
        data = cursor.fetchall()
        payrollcount = cursor.rowcount
        row2 = 0 # row count for left part of payslip
        row1 = 1 # use to check if top or bottom of the page
        row3 = 0 # row count for right part of payslip
        for data in data:
            row4 = row2
            period = data[3].strftime("%m/%d/%Y") + "-" + data[4].strftime("%m/%d/%Y")
            periodfrom = data[3]
            periodto = data[4]
            empno = data[1]
            position = data[8]
            if position == "DRIVER":
                ashelper = "(HELPER)"
            else:
                ashelper = ""
            #Left part of payslip
            worksheet.write(row2, 0, "VTR SUMMARY")
            row2+=1
            worksheet.write(row2, 0, "Payroll Period:")
            worksheet.write(row2, 2, period)
            worksheet.write(row2, 5, "Pay Date:")
            worksheet.write(row2, 6, datetime.now(), dateformat)
            row2+=1
            worksheet.write(row2, 0, "Employee No/Name:")
            worksheet.write(row2, 2, str(data[1]) + " - " + data[2])
            row2+=1
            worksheet.write(row2, 0, "VTR #", bold)
            worksheet.write(row2, 1, "PLATE #", bold)
            worksheet.write(row2, 2, "DATE", bold)
            worksheet.write(row2, 3, "DESTINATION", bold)
            worksheet.write(row2, 4, "TRIP EQ", bold)
            worksheet.write(row2, 5, "TRIP RATE", bold)
            worksheet.write(row2, 6, "ALLOWANCE", bold)
            worksheet.write(row2, 7, "TRUCK VALE", bold)
            row2+=1
            cursor.execute(f"SELECT * FROM TBL_VTR where driverempno = {empno} and periodfrom = '{periodfrom}' and periodto = '{periodto}' order by periodfrom")
            vtr = cursor.fetchall()
            for vtr in vtr:
                worksheet.write(row2, 0, vtr[1])
                worksheet.write(row2, 1, vtr[3])
                worksheet.write(row2, 2, vtr[2], dateformat1)
                worksheet.write(row2, 3, vtr[5])
                worksheet.write(row2, 4, vtr[6], money)
                worksheet.write(row2, 5, vtr[9], money)
                worksheet.write(row2, 6, vtr[10], money)
                worksheet.write(row2, 7, vtr[12], money)
                row2+=1
            cursor.execute(f"SELECT TBL_HELPER.vtrno, plateno, vtrdate, locationto, TBL_HELPER.tripeq, TBL_HELPER.tripamount, helperallowance, TBL_HELPER.truckvale \
                        FROM TBL_HELPER inner join  TBL_VTR on TBL_HELPER.vtrno = TBL_VTR.vtrno \
                        where helperempno = {empno} and TBL_HELPER.periodfrom = '{periodfrom}' and TBL_HELPER.periodto = '{periodto}' order by TBL_HELPER.periodfrom")
            vtr = cursor.fetchall()
            for vtr in vtr:
                worksheet.write(row2, 0, vtr[0])
                worksheet.write(row2, 1, vtr[1])
                worksheet.write(row2, 2, vtr[2], dateformat1)
                worksheet.write(row2, 3, vtr[3] + ashelper)
                worksheet.write(row2, 4, vtr[4], money)
                worksheet.write(row2, 5, vtr[5], money)
                worksheet.write(row2, 6, vtr[6], money)
                worksheet.write(row2, 7, vtr[7], money)
                row2+=1
            row2 = row4 + 21       
            if row1 % 2 == 0: #check if top or bottom part of payslip for new line spacing
                row2 += 1
                row3 += 1
            else:
                row2 += 2
                row3 += 2
            row1 += 1
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='VTR.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required    
def sss_ded_reports(request):
    company = request.session['company']
    paymonth = request.POST['paymonth1']
    payyear = request.POST['payyear1']
    gov_ded_type = request.POST['dedtype']
    month = calendar.month_name[int(paymonth)]
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where paymonth = {paymonth} and payyear = '{payyear}' and company = '{company}'")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        bold1 = workbook.add_format({"bold": True})
        money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        format = workbook.add_format({'num_format': '@','border':1})
        allborder = workbook.add_format({'border':1})
        worksheet.set_column('A:A',15,None)
        worksheet.set_column('B:B',30,None)
        #excel header
        worksheet.write('A1', gov_ded_type + ' CONTRIBUTION REPORTS  -  (' + month + " " + payyear + ')', bold1)
        worksheet.write('C1', 'SALARY', bold1)
        cursor.execute(f"SELECT periodfrom, periodto, paytype FROM TBL_PAYROLL_PERIOD where company = '{company}' and (employeetype = 'OFFICE STAFF' or employeetype ='DRIVER') and paymonth = {paymonth} and \
                       payyear = {payyear} and status = 'PAID' order by periodfrom")
        col_count = cursor.rowcount
        col = 2
        row = 1
        i = 1
        period = []
        period1 = []
        if col_count > 0:
            data = cursor.fetchall()
            for data in data:
                period.append(data[0])
                period1.append(data[1])
                periodfrom = parse_datetime(str(data[0]))
                periodfrom = periodfrom.strftime("%d")
                periodto = parse_datetime(str(data[1]))
                periodto = periodto.strftime("%d")
                worksheet.write(row, col, i, bold)
                worksheet.write(row+1, col, periodfrom + "-" + periodto, bold)
                worksheet.write(row, col + col_count + 1 , i, bold)
                worksheet.write(row+1, col + col_count + 1 ,  periodfrom + "-" + periodto, bold)
                col+=1
                i+=1
            worksheet.write(row, col, "Total", bold)
            worksheet.write(row-1, col+1, "CONTRIBUTIONS", bold1)
            worksheet.write(row+1, col, "", bold)
            worksheet.write(row, col + col_count + 1 , "Total", bold)
            worksheet.write(row+1, col + col_count + 1 ,  "", bold)
            worksheet.write(row, col + col_count + 2 , "SSSER", bold)
            worksheet.write(row+1, col + col_count + 2 ,  "", bold)
            worksheet.write(row, col + col_count + 3 , "EC", bold)
            worksheet.write(row+1, col + col_count + 3 ,  "", bold)
        cursor.execute(f"SELECT empno, empname FROM TBL_FINAL_PAYROLL where company = '{company}' and paymonth = {paymonth} and \
                       payyear = {payyear} group by empno,empname order by empname")
        row_count = cursor.rowcount
        data = cursor.fetchall()
        for y in range(3, row_count+3, 1):
            for z in range(2, (col_count * 2) + 4, 1):
                worksheet.write(y, z, "",format)

        row = 3
        for data in data: 
            col = 0
            empno = data[0]
            cursor.execute(f"SELECT lname, fname, sss FROM TBL_MASTERFILE where empno = {empno}")
            emprec = cursor.fetchall()
            sssno = ""
            empname = ""
            for emprec in emprec:
                empname = emprec[0] + ", " + emprec[1]
                sssno = emprec[2]
            worksheet.write(row, col, str(sssno) , format)
            col+=1
            worksheet.write(row, col, empname, format)
            cursor.execute(f"SELECT grosspay, sssee, ssser, periodfrom, periodto, ec FROM TBL_FINAL_PAYROLL where empno = {empno} and paymonth = {paymonth} and \
                       payyear = {payyear} order by periodfrom")
            sss_result = cursor.fetchall()
            total_salary = 0
            total_sssee = 0
            total_ssser = 0
            total_ec = 0
            for sss_result in sss_result:
                x = 0
                col = 2
                for x in range(0,col_count,1):
                    if period[x] == sss_result[3] and period1[x] == sss_result[4]:
                        worksheet.write(row, col, sss_result[0] , money)
                        worksheet.write(row, col + col_count + 1 ,sss_result[1] , money) 
                    # else:
                    #     worksheet.write(row, col, "", money)
                    #     worksheet.write(row, col + count + 1 ,"" , money)
                    x+=1
                    col+=1
                total_salary += float(sss_result[0])
                total_sssee += float(sss_result[1])
                total_ssser += float(sss_result[2])
                total_ec += float(sss_result[5])
            worksheet.write(row, col_count + 2, total_salary , money)
            worksheet.write(row, (col_count * 2) + 3 ,total_sssee , money)
            worksheet.write(row, (col_count * 2) + 4 ,total_ssser , money)
            worksheet.write(row, (col_count * 2) + 5 ,total_ec , money)
            row+=1       
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='SSS DED summary.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    
@login_required    
def gov_ded_reports(request):
    company = request.session['company']
    paymonth = request.POST['paymonth1']
    payyear = request.POST['payyear1']
    if request.POST['dedtype'] == "PHILHEALTH":
        gov_ded_type = "PHIC"
    elif request.POST['dedtype'] == "PAGIBIG":
        gov_ded_type = "HDMF"
    elif request.POST['dedtype'] == "TAX":
        gov_ded_type = "TAX"
    else:
        gov_ded_type = "SSS"
    month = calendar.month_name[int(paymonth)]
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_FINAL_PAYROLL where paymonth = {paymonth} and payyear = '{payyear}' and company = '{company}'")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        bold1 = workbook.add_format({"bold": True})
        money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        format = workbook.add_format({'num_format': '@','border':1})
        allborder = workbook.add_format({'border':1})
        worksheet.set_column('A:A',15,None)
        worksheet.set_column('B:B',30,None)
        #excel header
        worksheet.write('A1', gov_ded_type + ' CONTRIBUTION REPORTS  -  (' + month + " " + payyear + ')', bold1)
        worksheet.write('C1', 'SALARY', bold1)
        cursor.execute(f"SELECT periodfrom, periodto, paytype FROM TBL_PAYROLL_PERIOD where company = '{company}' and (employeetype = 'OFFICE STAFF' or employeetype ='DRIVER') and paymonth = {paymonth} and \
                       payyear = {payyear} and status = 'PAID' order by periodfrom")
        col_count = cursor.rowcount
        col = 2
        row = 1
        i = 1
        period = []
        period1 = []
        if col_count > 0:
            data = cursor.fetchall()
            for data in data:
                period.append(data[0])
                period1.append(data[1])
                periodfrom = parse_datetime(str(data[0]))
                periodfrom = periodfrom.strftime("%d")
                periodto = parse_datetime(str(data[1]))
                periodto = periodto.strftime("%d")
                worksheet.write(row, col, i, bold)
                worksheet.write(row+1, col, periodfrom + "-" + periodto, bold)
                worksheet.write(row, col + col_count + 1 , i, bold)
                worksheet.write(row+1, col + col_count + 1 ,  periodfrom + "-" + periodto, bold)
                col+=1
                i+=1
            worksheet.write(row, col, "Total", bold)
            worksheet.write(row-1, col+1, "CONTRIBUTIONS", bold1)
            worksheet.write(row+1, col, "", bold)
            worksheet.write(row, col + col_count + 1 , "Total", bold)
            worksheet.write(row+1, col + col_count + 1 ,  "", bold)
        cursor.execute(f"SELECT empno, empname FROM TBL_FINAL_PAYROLL where company = '{company}' and paymonth = {paymonth} and \
                       payyear = {payyear} group by empno,empname order by empname")
        row_count = cursor.rowcount
        data = cursor.fetchall()
        for y in range(3, row_count+3, 1):
            for z in range(2, (col_count * 2) + 4, 1):
                worksheet.write(y, z, "",format)

        row = 3
        for data in data: 
            col = 0
            empno = data[0]
            cursor.execute(f"SELECT lname, fname, {gov_ded_type} FROM TBL_MASTERFILE where empno = {empno}")
            emprec = cursor.fetchall()
            empname = ""
            govno = ""
            for emprec in emprec:
                empname = emprec[0] + ", " + emprec[1]
                govno = emprec[2]
            worksheet.write(row, col, str(govno) , format)
            col+=1
            worksheet.write(row, col, empname, format)
            cursor.execute(f"SELECT grosspay, {gov_ded_type}, periodfrom, periodto FROM TBL_FINAL_PAYROLL where empno = {empno} and paymonth = {paymonth} and \
                       payyear = {payyear} order by periodfrom")
            gov_result = cursor.fetchall()
            total_salary = 0
            total_gov = 0
            for gov_result in gov_result:
                x = 0
                col = 2
                for x in range(0,col_count,1):
                    if period[x] == gov_result[2] and period1[x] == gov_result[3]:
                        worksheet.write(row, col, gov_result[0] , money)
                        worksheet.write(row, col + col_count + 1 ,gov_result[1] , money)
                    x+=1
                    col+=1
                total_salary += float(gov_result[0])
                total_gov += float(gov_result[1])
            worksheet.write(row, col_count + 2, total_salary , money)
            worksheet.write(row, (col_count * 2) + 3 ,total_gov , money)
            row+=1       
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename=f'{gov_ded_type} DED summary.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required    
def ded_reports(request):
    company = request.session['company']
    paymonth = request.POST['paymonth1']
    payyear = request.POST['payyear1']
    ded_type = request.POST['dedtype']
    month = calendar.month_name[int(paymonth)]
    cursor = connection.cursor()
    cursor.execute(f"SELECT * FROM TBL_FINAL_DEDUCTIONS where DEDNAME = '{ded_type}' AND paymonth = {paymonth} and payyear = '{payyear}' and company = '{company}'")
    data = cursor.fetchall()
    payrollcount = cursor.rowcount
    if payrollcount > 0:
        buffer = io.BytesIO()
        workbook = xlsxwriter.Workbook(buffer)
        worksheet = workbook.add_worksheet()
        bold = workbook.add_format({"bold": True, 'border':1})
        bold.set_align('center')
        bold1 = workbook.add_format({"bold": True})
        money = workbook.add_format({'num_format': '#,##0.00', 'border':1})
        money1 = workbook.add_format({'num_format': '#,##0.00'})
        format = workbook.add_format({'num_format': '@','border':1})
        worksheet.set_column('A:A',30,None)
        #excel header
        worksheet.write('A1', ded_type + ' DEDUCTION REPORTS  -  (' + month + " " + payyear + ')', bold1)
        cursor.execute(f"SELECT periodfrom, periodto, paytype FROM TBL_PAYROLL_PERIOD where company = '{company}' and (employeetype = 'OFFICE STAFF' or employeetype ='DRIVER') and paymonth = {paymonth} and \
                       payyear = {payyear} and status = 'PAID' order by periodfrom")
        count = cursor.rowcount
        col = 1
        row = 1
        i = 1
        period = []
        emptype = []
        if count > 0:
            data = cursor.fetchall()
            for data in data:
                period.append(data[0])
                emptype.append(data[2])
                periodfrom = parse_datetime(str(data[0]))
                periodfrom = periodfrom.strftime("%d")
                periodto = parse_datetime(str(data[1]))
                periodto = periodto.strftime("%d")
                worksheet.write(row, col, i, bold)
                worksheet.write(row+1, col, periodfrom + "-" + periodto, bold)
                col+=1
                i+=1
            worksheet.write(row, col, "Total", bold)
            worksheet.write(row+1, col, "", bold)
        cursor.execute(f"SELECT TBL_FINAL_DEDUCTIONS.empno, lname, fname FROM TBL_FINAL_DEDUCTIONS INNER JOIN TBL_MASTERFILE ON \
                       TBL_FINAL_DEDUCTIONS.empno = TBL_MASTERFILE.EMPNO where TBL_FINAL_DEDUCTIONS.company = '{company}' and TBL_FINAL_DEDUCTIONS.paymonth = {paymonth} and \
                       TBL_FINAL_DEDUCTIONS.payyear = {payyear} and dedname = '{ded_type}'group by TBL_FINAL_DEDUCTIONS.empno,lname,fname order by lname")
        data = cursor.fetchall()
        row = 3
        for data in data:
            x = 0
            empno = data[0]
            empname = data[1] + ", " + data[2]
            worksheet.write(row, 0, empname, format)
            total_amount = 0
            for x in range(0,count,1):
                cursor.execute(f"SELECT SUM(amount) AS ded_amount from TBL_FINAL_DEDUCTIONS where empno = {empno} and \
                               periodfrom = '{period[x]}' and paymonth = {paymonth} and payyear = {payyear} and dedname = '{ded_type}'\
                                and paytype = '{emptype[x]}'")
                ded_amount = cursor.fetchall()
                for ded_amount in ded_amount:
                    if ded_amount[0] != None:
                        amount = ded_amount[0]
                        total_amount += amount
                        worksheet.write(row, x+1, amount, money)
                    else:
                        worksheet.write(row, x+1, "", format)
            worksheet.write(row, x+2, total_amount, money)    
            row += 1
                       
        workbook.close()
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename='DEDUCTION summary.xlsx')
    else:
            messages.info(request, 'No record to export.')
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

@login_required       
def other_reports(request,rowid):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    if request.method == 'POST':
        if rowid == 1:
            cursor.execute(f"Select location from TBL_LOCATION where company = '{company}' order by location")
            location = cursor.fetchall()
            return render(request, 'reports_1.html',{'location': location, 'role1': request.session['role1']})
        if rowid == 2:
            # cursor.execute(f"Select location from TBL_LOCATION where company = '{company}' order by location")
            # location = cursor.fetchall()
            return render(request, 'reports_2.html',{ 'role1': request.session['role1']})
    return render(request, 'other_reports.html',{ 'role1': request.session['role1']})

@login_required   
def upload_data(request,rowid):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    return render(request, 'upload_data.html',{'role1': request.session['role1']})

@login_required   
def submit_upload(request,rowid):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    if request.method == 'POST':
        if rowid == 1:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/Masterfile.xlsx")
            sh = wb.active
            
            for r in range(2, 155, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = sh.cell(row=r,column=7).value
                c8 = sh.cell(row=r,column=8).value
                c9 = sh.cell(row=r,column=9).value
                c10 = sh.cell(row=r,column=10).value
                c11 = sh.cell(row=r,column=11).value
                c12 = sh.cell(row=r,column=12).value
                c13 = sh.cell(row=r,column=13).value
                c14 = sh.cell(row=r,column=14).value
                c15 = sh.cell(row=r,column=15).value
                c16 = sh.cell(row=r,column=16).value
                c17 = sh.cell(row=r,column=17).value
                c18 = sh.cell(row=r,column=18).value
                c19 = sh.cell(row=r,column=19).value
                c20 = sh.cell(row=r,column=20).value
                c21 = sh.cell(row=r,column=21).value
                c22 = sh.cell(row=r,column=22).value
                c23 = sh.cell(row=r,column=23).value
                c24 = sh.cell(row=r,column=24).value
                c25 = sh.cell(row=r,column=25).value
                c26 = sh.cell(row=r,column=26).value
                c27 = sh.cell(row=r,column=27).value
                c28 = sh.cell(row=r,column=28).value
                c29 = sh.cell(row=r,column=29).value
                c30 = sh.cell(row=r,column=30).value
                c31 = sh.cell(row=r,column=31).value
                c32 = sh.cell(row=r,column=32).value
                c33 = sh.cell(row=r,column=33).value
                c34 = sh.cell(row=r,column=34).value
                c35 = sh.cell(row=r,column=35).value
                c36 = sh.cell(row=r,column=36).value
                c37 = sh.cell(row=r,column=37).value
                c38 = username
                c39 = datetime.now()
                c40 = company
                c41 = username
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_MASTERFILE (empno,lname, fname, mname, birthdate, address1, address2, \
                        contact, gender, civil, spouse, child1, child2, sss, phic, hdmf, tax, sssd, phicd, hdmfd, \
                        taxd, hdmfpay, department, position, location, triprate, salary, allowance, datehired, datepermanent,\
                        status, password, showpayslip, payrolltype, atm, mother, emergency, userlastupdate, datelastupdated, company,username)"
                                "VALUES (%s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,\
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16, c17, c18, c19, c20, c21, c22, \
                                c23, c24, c25, c26, c27, c28, c29, c30, c31, c32, c33, c34, c35, c36, c37, c38, c39, c40, c41)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 2:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/final_payroll.xlsx")
            sh = wb.active
            
            for r in range(2, 2443, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = ""
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = sh.cell(row=r,column=7).value
                c8 = sh.cell(row=r,column=8).value
                c9 = sh.cell(row=r,column=9).value
                c10 = sh.cell(row=r,column=10).value
                c11 = 0
                c12 = 0
                c13 = sh.cell(row=r,column=13).value
                c14 = 0
                c15 = 0
                c16 = sh.cell(row=r,column=16).value
                c17 = sh.cell(row=r,column=17).value
                c18 = sh.cell(row=r,column=18).value
                c19 = sh.cell(row=r,column=19).value
                c20 = 0
                c21 = sh.cell(row=r,column=21).value
                c22 = sh.cell(row=r,column=22).value
                c23 = sh.cell(row=r,column=23).value
                c24 = sh.cell(row=r,column=24).value
                c25 = "FCMC"
                c26 = 0
                c27 = 0
                c28 = 0
                c29 = 0
                c30 = 0
                c31 = 0
                c32 = "SANTOLAN"
                c33 = 0
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_FINAL_PAYROLL (empno,EMPNAME, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, PAYTYPE, \
                        EMPTYPE, TRIPS, BASICPAY, ALLOWANCE, EXCESSTRIP, GROSSPAY, WITHHOLDINGTAX, GROSSAFTERTAX, SSSEE, SSSER, PHIC, HDMF, TAX, \
                        TOTALDEDUCTIONS, NETPAY, DATEPROCESS, USER, COMPANY, OTHEREARNINGS, atm, OT, HOLIDAY, TARDY,\
                        OTHERDEDUCTIONS, LOCATION, ec)"
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,\
                                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16, c17, c18, c19, c20, c21, c22, \
                                c23, c24, c25, c26, c27, c28, c29, c30, c31, c32, c33)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 3:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/VTR.xlsx")
            sh = wb.active
            
            for r in range(2, 8674, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = sh.cell(row=r,column=7).value
                c8 = sh.cell(row=r,column=8).value
                c9 = sh.cell(row=r,column=9).value
                c10 = sh.cell(row=r,column=10).value
                c11 = sh.cell(row=r,column=11).value
                c12 = sh.cell(row=r,column=12).value
                c13 = sh.cell(row=r,column=13).value
                c14 = sh.cell(row=r,column=14).value
                c15 = 0
                c16 = 0
                c17 = sh.cell(row=r,column=17).value
                c18 = "FCMC"
                c19 = sh.cell(row=r,column=19).value
                if sh.cell(row=r,column=20).value == "YES": c20 = 1
                else: c20 = 0
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_VTR (VTRNO, VTRDATE, PLATENO, LOCATIONFROM, LOCATIONTO, TRIPEQ, ALLOWANCE, \
                        DRIVEREMPNO, TRIPAMOUNT, DRIVERALLOWANCE, EXCESSTRIP, TRUCKVALE, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, USER, COMPANY, DATEMAINTAINED, DH)"
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,\
                                    %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14, c15, c16, c17, c18, c19, c20)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 4:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/VTR_HELPER.xlsx")
            sh = wb.active
            
            for r in range(2, 9002, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = sh.cell(row=r,column=7).value
                # c8 = sh.cell(row=r,column=8).value
                # c9 = sh.cell(row=r,column=9).value
                c10 = 0
                c11 = 0
                c12 = sh.cell(row=r,column=12).value
                c13 = "FCMC"
                c14 = 0
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_HELPER(VTRNO, HELPEREMPNO, helpername, TRIPAMOUNT, HELPERALLOWANCE, EXCESSTRIP, TRUCKVALE, \
                        PAYMONTH, PAYYEAR, USER, COMPANY, TRIPEQ)"
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c10, c11, c12, c13, c14)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 5:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/PERIOD.xlsx")
            sh = wb.active
            
            for r in range(2, 44, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = sh.cell(row=r,column=7).value
                c8 = sh.cell(row=r,column=8).value
                c9 = sh.cell(row=r,column=9).value
                c10 = "FCMC"
                c11 = ""
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_PAYROLL_PERIOD(paytype, employeetype, periodfrom, periodto, status, paymonth, payyear, \
                        datemaintained, user, company, PAYROLLPERIOD)"
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 6:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/FINAL_DED.xlsx")
            sh = wb.active
            
            for r in range(2, 22727, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = sh.cell(row=r,column=7).value
                c8 = sh.cell(row=r,column=8).value
                c9 = sh.cell(row=r,column=9).value
                c10 = sh.cell(row=r,column=10).value
                c11 = sh.cell(row=r,column=11).value
                c12 = sh.cell(row=r,column=12).value
                c13 = "FCMC"
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_FINAL_DEDUCTIONS(EMPNO, PAYTYPE, EMPTYPE, PERIODFROM, PERIODTO, PAYMONTH, PAYYEAR, \
                        DEDTYPE, DEDNAME, AMOUNT, ID, USER, COMPANY)"
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 7:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/EARNINGS.xlsx")
            sh = wb.active
            
            for r in range(2, 1896, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = datetime.now()
                c8 = datetime.now()
                c9 = username
                c10 = "FCMC"
                c11 = 1
                c12 = sh.cell(row=r,column=12).value
                c13 = sh.cell(row=r,column=13).value
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_OTHER_EARNINGS(EARNINGSTYPE, VTRNO, AMOUNT, EMPNO, PERIODFROM, PERIODTO, DATEMAINTAINED, \
                        DATELASTUPDATE, USER, COMPANY, STATUS, PAYMONTH, PAYYEAR)"
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 8:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/fixed_ded.xlsx")
            sh = wb.active
            
            for r in range(2, 243, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = sh.cell(row=r,column=4).value
                c5 = sh.cell(row=r,column=5).value
                c6 = sh.cell(row=r,column=6).value
                c7 = sh.cell(row=r,column=7).value
                c8 = sh.cell(row=r,column=8).value
                c9 = sh.cell(row=r,column=9).value
                c10 = sh.cell(row=r,column=10).value
                if sh.cell(row=r,column=11).value == "ACTIVE": c11 = 0
                else: c11 = 1
                c12 = username
                if sh.cell(row=r,column=13).value == "TRUE": c13 = "YES"
                else: c13 = "NO"
                c14 = "FCMC"
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_FIXED_DEDUCTIONS(EMPNO, DEDTYPE, DEDAMOUNT, NOOFPERIOD, PERIODDEDAMOUNT, AMTREMAINING, PERIODREMAINING, \
                        DATESTART, DATEMAINTAINED, DATELASTUPDATE, STATUS, USER, DEFERREDPAY, COMPANY)"
                                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4, c5, c6, c7, c8, c9, c10, c11, c12, c13, c14)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 9:
            wb = openpyxl.load_workbook("/Users/lorilenmedrano/Desktop/vtr_exp.xlsx")
            sh = wb.active
            
            for r in range(2, 9002, 1):
                c1 = sh.cell(row=r,column=1).value
                c2 = sh.cell(row=r,column=2).value
                c3 = sh.cell(row=r,column=3).value
                c4 = "FCMC"
                cursor = connection.cursor()
                # insert new data to TBL_MASTERFILE table
                query = ("insert into TBL_VTR_EXPENSES(VTRNO, EXPENSES, AMOUNT, COMPANY)"
                                "VALUES (%s, %s, %s, %s)"
                            )
                data_values = (c1, c2, c3, c4)
                cursor.execute(query,data_values)
            messages.info(request,"Done")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    return render(request, 'upload_data.html',{'role1': request.session['role1']})

@login_required   
def submit_fix(request,rowid):
    username = request.session['username']
    company = request.session['company']
    cursor = connection.cursor()
    if request.method == 'POST':
        if rowid == 1:
            cursor.execute(f"Select * from TBL_PAYROLL_PERIOD")
            data = cursor.fetchall()
            for r in data:
                periodfrom = r[3]
                periodto = r[4]
                paytype = r[1]
                cursor.execute(f"Update TBL_FINAL_DEDUCTIONS set paytype = '{paytype}' where periodfrom = '{periodfrom}' and periodto = '{periodto}'")
            messages.info(request,"Done fix")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 2:
            cursor.execute(f"Select * from TBL_PAYROLL_PERIOD")
            data = cursor.fetchall()
            for r in data:
                periodfrom = r[3]
                periodto = r[4]
                payrollperiod = periodfrom.strftime("%m/%d/%Y") + "-" + periodto.strftime("%m/%d/%Y")
                cursor.execute(f"Update TBL_PAYROLL_PERIOD set payrollperiod = '{payrollperiod}' where periodfrom = '{periodfrom}' and periodto = '{periodto}'")
            messages.info(request,"Done fix")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 3:
            cursor.execute(f"Select * from TBL_PAYROLL_PERIOD")
            data = cursor.fetchall()
            for r in data:
                periodfrom = r[3]
                periodto = r[4]
                paymonth = r[6]
                payyear = r[7]
                cursor.execute(f"Update TBL_VTR set paymonth = {paymonth}, payyear = '{payyear}' where periodfrom = '{periodfrom}' and periodto = '{periodto}'")
            messages.info(request,"Done fix")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 4:
            cursor.execute(f"Select * from TBL_VTR")
            data = cursor.fetchall()
            for r in data:
                periodfrom = r[13]
                periodto = r[14]
                paymonth = r[15]
                payyear = r[16]
                vtrno = r[1]
                cursor.execute(f"Update TBL_HELPER set paymonth = {paymonth}, payyear = '{payyear}', periodfrom = '{periodfrom}', periodto = '{periodto}' where \
                               vtrno = '{vtrno}'")
            messages.info(request,"Done fix")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
        if rowid == 5:
            cursor.execute(f"Select * from TBL_MASTERFILE")
            data = cursor.fetchall()
            for r in data:
                empname = r[2] + ", " + r[3]
                empno = r[1]
                cursor.execute(f"Update TBL_FINAL_PAYROLL set empname = '{empname}' where empno = {empno}")
            messages.info(request,"Done fix")
            return HttpResponseRedirect(request.META.get('HTTP_REFERER'))
    return render(request, 'fix_data.html',{'role1': request.session['role1']})
